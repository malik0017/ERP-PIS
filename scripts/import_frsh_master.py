#!/usr/bin/env python3
"""
Batch 125 — FRSH master-data importer (CORRECTED).

Fixes the Batch 124 failure. The old script inserted into recipe_ingredients /
ingredients using column names that do not exist in the live schema (e.g.
ingredient_code / required_qty on recipe_ingredients), so every line insert
threw "unknown column" and the whole transaction rolled back — leaving 0
ingredients, 0 recipe lines, and blank day/category on the recipe shells.

This version uses the EXACT columns the app's own models/importer use, verified
against app/models/*.py and app/modules/recipes/routes.py:

  recipe_ingredients: recipe_id, line_no, line_type, inventory_code, item_name,
                      uom, qty_batch, portions, qty_per_portion, cost_uom,
                      line_cost, line_cost_per_portion, remark, missing_cost
                      (NOTE: NOT ingredient_code / required_qty / yield_pct)
  ingredients:        ingredient_code, name, category, main_category,
                      sub_category, purchase_uom, recipe_uom, standard_uom,
                      unit_cost_standard, status='Active', company_id
  recipes:            recipe_code, recipe_name, customer_name, category,
                      day_of_week, standard_portions, status='ACTIVE',
                      is_active, approval_status='APPROVED', version, company_id

Every Excel column on the "Recipe Ingredients" sheet is carried (section,
sub-recipe, yield, butchery note go into the line remark; qty/price/cost map to
real numeric columns). Idempotent; each recipe commits in its own transaction so
one bad row can never wipe the run, and failures are printed.

USAGE
    python scripts/import_frsh_master.py --file "FRSH_Master_Recipes_with_Inventory_Codes.xlsx" --company 1 --dry-run
    python scripts/import_frsh_master.py --file "FRSH_Master_Recipes_with_Inventory_Codes.xlsx" --company 1
"""
from __future__ import annotations
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402
from sqlalchemy import text  # noqa: E402

try:
    from app.database.session import SessionLocal
except Exception:  # pragma: no cover
    from app.db import SessionLocal  # type: ignore


SECTION_MAP = {
    "trayline section": "Trayline / Packing",
    "hot section": "Hot Kitchen",
    "cold section": "Cold Kitchen",
    "butchery section": "Butchery",
    "pastry section": "Bakery/Pastry",
    "section": "",
}


def _s(v):
    return "" if v is None else str(v).strip()


def _f(v, default=0.0):
    try:
        if v is None or str(v).strip() == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def map_section(raw):
    return SECTION_MAP.get(_s(raw).lower(), _s(raw))


def _find_header(ws, must_have, max_scan=8):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        vals = [_s(c).replace("\n", " ").lower() for c in row]
        if all(h in vals for h in must_have):
            hdr = {}
            for j, c in enumerate(row):
                if c is not None:
                    hdr[_s(c).replace("\n", " ").lower()] = j
            return i, hdr
    return None, None


def col(hdr, *names):
    for nm in names:
        if nm in hdr:
            return hdr[nm]
    return None


# ---------------------------------------------------------------------------
# 1. RAW MATERIAL  ->  ingredients
# ---------------------------------------------------------------------------
def import_raw_materials(db, ws, company_id, dry):
    hrow, hdr = _find_header(ws, ["code"])
    if hrow is None:
        print("  ! Raw material header not found — skipping")
        return 0, 0
    ci_code = col(hdr, "code")
    ci_name = col(hdr, "nameen", "items description", "name")
    ci_unit = col(hdr, "unit")
    ci_price = col(hdr, "price")
    ci_main = col(hdr, "main category")
    ci_sub = col(hdr, "sub category")

    ok = fail = 0
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        code = _s(row[ci_code]) if ci_code is not None else ""
        if not code:
            continue
        name = _s(row[ci_name]) if ci_name is not None else code
        unit = (_s(row[ci_unit]) if ci_unit is not None else "") or "Each"
        price = _f(row[ci_price]) if ci_price is not None else 0.0
        main = _s(row[ci_main]) if ci_main is not None else ""
        sub = _s(row[ci_sub]) if ci_sub is not None else ""
        if dry:
            ok += 1
            continue
        try:
            db.execute(text("""
                INSERT INTO ingredients
                    (company_id, ingredient_code, name, category, main_category,
                     sub_category, purchase_uom, recipe_uom, standard_uom,
                     unit_cost_standard, status, created_at, updated_at)
                VALUES
                    (:cid, :code, :name, :main, :main, :sub, :unit, :unit, :unit,
                     :price, 'Active', NOW(), NOW())
                ON DUPLICATE KEY UPDATE
                    name = VALUES(name),
                    category = VALUES(category),
                    main_category = VALUES(main_category),
                    sub_category = VALUES(sub_category),
                    purchase_uom = VALUES(purchase_uom),
                    recipe_uom = VALUES(recipe_uom),
                    standard_uom = VALUES(standard_uom),
                    unit_cost_standard = VALUES(unit_cost_standard),
                    status = 'Active',
                    updated_at = NOW()
            """), {"cid": company_id, "code": code, "name": name, "main": main,
                   "sub": sub, "unit": unit, "price": price})
            db.commit()
            ok += 1
        except Exception as exc:
            db.rollback()
            fail += 1
            if fail <= 5:
                print(f"    raw '{code}' failed: {exc}")
    return ok, fail


# ---------------------------------------------------------------------------
# 2. RECIPE INGREDIENTS  ->  recipes + recipe_ingredients
# ---------------------------------------------------------------------------
def import_recipe_ingredients(db, ws, company_id, dry):
    hrow, hdr = _find_header(ws, ["recipe code", "item code"])
    if hrow is None:
        print("  ! Recipe Ingredients header not found — skipping")
        return 0, 0, 0

    c_rcode = col(hdr, "recipe code")
    c_cat = col(hdr, "category")
    c_cust = col(hdr, "customer name")
    c_rname = col(hdr, "recipe name")
    c_day = col(hdr, "day")
    c_sect = col(hdr, "section")
    # Batch 136: Butchery cutting / portion-size column (header carries a typo
    # "Buchery" in the workbook — match both spellings).
    c_cut = col(hdr, "buchery cutting /portion size", "butchery cutting /portion size",
                "buchery cutting / portion size", "cutting /portion size", "portion size")
    c_subdesc = col(hdr, "sub recipe description")
    c_icode = col(hdr, "item code")
    c_iname = col(hdr, "item / ingredient", "item/ingredient")
    c_uom = col(hdr, "uom")
    c_net = col(hdr, "net qty req per batch (g/pcs)")
    c_gross = col(hdr, "gross qty req per batch (g/pcs)")
    c_yield = col(hdr, "yield %")
    c_portions = col(hdr, "no. of portions per batch")
    c_qtyport = col(hdr, "qty req per portion")
    c_price = col(hdr, "purchase price  standard uom", "purchase price standard uom")
    c_fcpp = col(hdr, "food cost per portion")
    c_total = col(hdr, "total cost")
    c_butchery = col(hdr, "buchery cutting /portion size", "butchery cutting /portion size")

    recipes: dict[str, dict] = {}
    order: list[str] = []
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        rcode = _s(row[c_rcode]) if c_rcode is not None else ""
        icode = _s(row[c_icode]) if c_icode is not None else ""
        if not rcode or not icode:
            continue
        if rcode not in recipes:
            recipes[rcode] = {
                "name": _s(row[c_rname]) if c_rname is not None else rcode,
                "category": _s(row[c_cat]) if c_cat is not None else "",
                "customer": (_s(row[c_cust]) if c_cust is not None else "") or "Frsh",
                "day": _s(row[c_day]) if c_day is not None else "",
                "portions": _f(row[c_portions], 1.0) if c_portions is not None else 1.0,
                "lines": [],
            }
            order.append(rcode)
        recipes[rcode]["lines"].append({
            "icode": icode,
            "iname": _s(row[c_iname]) if c_iname is not None else icode,
            "uom": (_s(row[c_uom]) if c_uom is not None else "") or "Gram",
            "gross": _f(row[c_gross]) if c_gross is not None else 0.0,
            "net": _f(row[c_net]) if c_net is not None else 0.0,
            "yield": _f(row[c_yield], 100.0) if c_yield is not None else 100.0,
            "portions": _f(row[c_portions], 1.0) if c_portions is not None else 1.0,
            "qty_port": _f(row[c_qtyport]) if c_qtyport is not None else 0.0,
            "price": _f(row[c_price]) if c_price is not None else 0.0,
            "fcpp": _f(row[c_fcpp]) if c_fcpp is not None else 0.0,
            "total": _f(row[c_total]) if c_total is not None else 0.0,
            "section": map_section(row[c_sect]) if c_sect is not None else "",
            "section_raw": _s(row[c_sect]) if c_sect is not None else "",
            "cutting": _s(row[c_cut]) if c_cut is not None else "",
            "subdesc": _s(row[c_subdesc]) if c_subdesc is not None else "",
            "butchery": _s(row[c_butchery]) if c_butchery is not None else "",
        })

    if dry:
        total_lines = sum(len(r["lines"]) for r in recipes.values())
        return len(recipes), total_lines, 0

    ok_recipes = ok_lines = fail = 0
    for rcode in order:
        r = recipes[rcode]
        try:
            db.execute(text("""
                INSERT INTO recipes
                    (company_id, recipe_code, recipe_name, customer_name, category,
                     day_of_week, status, is_active, approval_status, version,
                     standard_portions, created_at, updated_at)
                VALUES
                    (:cid, :code, :name, :cust, :cat, :day, 'ACTIVE', 1, 'APPROVED', 1,
                     :portions, NOW(), NOW())
                ON DUPLICATE KEY UPDATE
                    recipe_name = VALUES(recipe_name),
                    customer_name = VALUES(customer_name),
                    category = VALUES(category),
                    day_of_week = VALUES(day_of_week),
                    standard_portions = VALUES(standard_portions),
                    status='ACTIVE', is_active=1, updated_at = NOW()
            """), {"cid": company_id, "code": rcode, "name": r["name"],
                   "cust": r["customer"], "cat": r["category"],
                   "day": r["day"], "portions": r["portions"] or 1})

            rid = db.execute(text(
                "SELECT id FROM recipes WHERE company_id=:cid AND recipe_code=:code LIMIT 1"
            ), {"cid": company_id, "code": rcode}).scalar()
            if not rid:
                raise RuntimeError("recipe id not found after upsert")

            db.execute(text("DELETE FROM recipe_ingredients WHERE recipe_id=:rid"),
                       {"rid": rid})

            ln = 0
            for line in r["lines"]:
                ln += 1
                remark_bits = []
                if line["section"]:
                    remark_bits.append(f"Section: {line['section']}")
                if line["subdesc"]:
                    remark_bits.append(f"Sub: {line['subdesc']}")
                if line["butchery"]:
                    remark_bits.append(f"Butchery: {line['butchery']}")
                if line["yield"]:
                    remark_bits.append(f"Yield {line['yield']:.0f}%")
                remark = " | ".join(remark_bits)

                db.execute(text("""
                    INSERT INTO recipe_ingredients
                        (recipe_id, line_no, line_type, inventory_code, item_name,
                         uom, qty_batch, portions, qty_per_portion, cost_uom,
                         line_cost, line_cost_per_portion, sub_recipe_code, remark,
                         kitchen_section, cutting_portion_size, missing_cost, created_at, updated_at)
                    VALUES
                        (:rid, :ln, 'Main Recipe', :icode, :iname, :uom,
                         :qty_batch, :portions, :qty_port, :cost_uom,
                         :line_cost, :fcpp, :subrecipe, :remark, :ksec, :cut, :missing, NOW(), NOW())
                """), {
                    "rid": rid, "ln": ln, "icode": line["icode"],
                    "iname": line["iname"], "uom": line["uom"],
                    "qty_batch": line["gross"] or line["net"],
                    "portions": line["portions"] or 1,
                    "qty_port": line["qty_port"],
                    "cost_uom": line["price"],
                    "line_cost": line["total"],
                    "fcpp": line["fcpp"],
                    # Batch 128: populate the Sub Recipe column (was left blank,
                    # only appearing inside the remark). Blank -> NULL.
                    "subrecipe": line["subdesc"] or None,
                    "remark": remark,
                    # Batch 132: store the recipe's kitchen section on the
                    # ingredient line so store-issuance routing (Batch 131) can
                    # send each ingredient to the right section. We store the
                    # mapped canonical name; the raw sheet value is preserved
                    # only in the remark. Blank -> NULL.
                    "ksec": line["section"] or line["section_raw"] or None,
                    # Batch 136: Butchery cutting / portion size. Blank -> NULL.
                    "cut": line["cutting"] or None,
                    "missing": 1 if not line["price"] else 0,
                })
                ok_lines += 1

                if line["section"]:
                    db.execute(text("""
                        UPDATE ingredients SET default_issue_section=:sec, updated_at=NOW()
                        WHERE ingredient_code=:code
                          AND (default_issue_section IS NULL OR default_issue_section=''
                               OR default_issue_section='Hot Kitchen')
                    """), {"sec": line["section"], "code": line["icode"]})

            db.commit()
            ok_recipes += 1
        except Exception as exc:
            db.rollback()
            fail += 1
            if fail <= 8:
                print(f"    recipe '{rcode}' failed: {exc}")
    return ok_recipes, ok_lines, fail


# ---------------------------------------------------------------------------
# 3. MENU  ->  set the FULL set of days each recipe is served, plus category.
#    A recipe (e.g. a salad) appears once PER DAY in the menu sheet, so we must
#    aggregate every day for a recipe and store them joined ("Sunday & Monday &
#    …"). The app's day matcher works by stem containment, so a combined value
#    correctly makes the recipe show on every one of its days. (The old version
#    overwrote day_of_week with the last row seen, so multi-day recipes like
#    salads/snacks only appeared on one day — the bug in images 2/4/7.)
# ---------------------------------------------------------------------------
def import_menu(db, ws, company_id, dry):
    hrow, hdr = _find_header(ws, ["recipe code", "days"])
    if hrow is None:
        print("  ! Menu header not found — skipping")
        return 0, 0
    c_day = col(hdr, "days")
    c_code = col(hdr, "recipe code")
    c_name = col(hdr, "recipe names", "recipe name")
    c_cat = col(hdr, "category")

    # Canonical weekday order so the stored string is stable/readable.
    ORDER = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    # Aggregate: recipe_code -> {name, category, set(days)}
    agg: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        code = _s(row[c_code]) if c_code is not None else ""
        if not code:
            continue
        # Use the code EXACTLY as it appears (same as the ingredients pass), so
        # the menu UPDATE targets the same recipe row instead of creating a
        # duplicate with a different case.
        code_norm = code
        day = _s(row[c_day]) if c_day is not None else ""
        name = _s(row[c_name]) if c_name is not None else code
        cat = _s(row[c_cat]) if c_cat is not None else ""
        a = agg.setdefault(code_norm, {"name": name, "cat": cat, "days": set()})
        if day:
            a["days"].add(day.strip().title())
        if cat and not a["cat"]:
            a["cat"] = cat

    ok = fail = 0
    for code, a in agg.items():
        days_sorted = [d for d in ORDER if d in a["days"]]
        # Batch 130: store the EXPLICIT list of days, never "Daily". FRSH runs
        # Sat–Thu (no Friday); collapsing 6 days to "Daily" made the matcher
        # (which treats Daily as all 7) wrongly show recipes on Friday. Listing
        # the real days keeps Friday empty when the menu has no Friday.
        day_value = " & ".join(days_sorted)
        if dry:
            ok += 1
            continue
        try:
            db.execute(text("""
                INSERT INTO recipes
                    (company_id, recipe_code, recipe_name, customer_name, category,
                     day_of_week, status, is_active, approval_status, version,
                     standard_portions, created_at, updated_at)
                VALUES
                    (:cid, :code, :name, 'Frsh', :cat, :day, 'ACTIVE', 1, 'APPROVED', 1, 1, NOW(), NOW())
                ON DUPLICATE KEY UPDATE
                    day_of_week = VALUES(day_of_week),
                    category = COALESCE(NULLIF(recipes.category,''), VALUES(category)),
                    customer_name='Frsh', status='ACTIVE', is_active=1, updated_at=NOW()
            """), {"cid": company_id, "code": code, "name": a["name"],
                   "cat": a["cat"], "day": day_value})
            db.commit()
            ok += 1
        except Exception as exc:
            db.rollback()
            fail += 1
            if fail <= 5:
                print(f"    menu '{code}' failed: {exc}")
    return ok, fail


def main():
    ap = argparse.ArgumentParser(description="Import FRSH master data into ISFC")
    ap.add_argument("--file", required=True)
    ap.add_argument("--company", type=int, default=1)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.file):
        print(f"File not found: {args.file}")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    names = {n.lower(): n for n in wb.sheetnames}
    db = SessionLocal()
    try:
        print(f"{'DRY RUN — ' if args.dry_run else ''}Importing FRSH into company_id={args.company}\n")

        if "raw material list" in names:
            ok, fail = import_raw_materials(db, wb[names["raw material list"]], args.company, args.dry_run)
            print(f"  Ingredients (raw materials): {ok} ok, {fail} failed")

        if "recipe ingredients" in names:
            rok, lok, rfail = import_recipe_ingredients(
                db, wb[names["recipe ingredients"]], args.company, args.dry_run)
            print(f"  Recipes: {rok} ok, {rfail} failed  |  Recipe lines: {lok}")

        if "menu" in names:
            # Batch 132: wipe any stale day_of_week on FRSH recipes BEFORE the
            # menu pass rewrites it. Earlier broken imports left ghost values
            # (e.g. a "Daily" or a truncated string that matched Friday), which
            # made the portal offer Friday even though the FRSH workbook has no
            # Friday menu. Clearing first guarantees the stored days are exactly
            # what this workbook contains — nothing survives from a prior run.
            if not args.dry_run:
                db.execute(text("""
                    UPDATE recipes SET day_of_week = NULL, updated_at = NOW()
                    WHERE company_id = :cid AND customer_name = 'Frsh'
                """), {"cid": args.company})
                db.commit()
            ok, fail = import_menu(db, wb[names["menu"]], args.company, args.dry_run)
            print(f"  Menu day/category rows: {ok} ok, {fail} failed")

        print("\nDry run complete — nothing written." if args.dry_run
              else "\nDone. FRSH master data imported.")
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    main()

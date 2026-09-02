#!/usr/bin/env python3
"""Batch DATA — SMC master import (recipes + ingredients + weekly menu + meal_order).

SMC's workbook has the SAME layout as FRSH (Menu / Master – Recipes / Recipe
Ingredients / Raw material list) plus a **Meal Order** column (BREAKFAST / LUNCH /
DINNER). This importer reuses the FRSH importer's helpers and only differs in:
  - customer_name = 'SMC'
  - it captures Menu → Meal Order into recipes.meal_order
  - it clears stale SMC recipes' day_of_week/meal_order before the menu pass

Usage:
  python scripts/import_smc_master.py --file "SMC_Master_Recipes_with_Inventory_Codes...V2.xlsx" --company 1 [--dry-run]

Prereqs: run the app once so the recipes.meal_order guard (Batch 158) has added
the column; ingredients (raw materials) must exist in Master Data first.
"""
import argparse, sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl                       # noqa: E402
from sqlalchemy import text           # noqa: E402
try:
    from app.database.session import SessionLocal
except Exception:                     # pragma: no cover
    from app.db import SessionLocal   # type: ignore

# Reuse the FRSH importer's building blocks so behaviour stays identical.
from scripts.import_frsh_master import (   # noqa: E402
    _find_header, col, _s, import_recipe_ingredients, import_raw_materials,
)

CUSTOMER = "SMC"
ORDER = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def import_menu_smc(db, ws, company_id, dry):
    """Menu pass with Meal Order capture (SMC-specific)."""
    hrow, hdr = _find_header(ws, [("recipe code", "recipe ref"), ("days", "day")])
    if hrow is None:
        print("  ! Menu header not found — skipping")
        return 0, 0
    c_day = col(hdr, "days", "day")
    c_code = col(hdr, "recipe code", "recipe ref")
    c_name = col(hdr, "recipe names", "recipe name")
    c_cat = col(hdr, "category")
    c_meal = col(hdr, "meal order", "meal")   # SMC column

    # ---- Batch 159 per-day meal encoding -------------------------------
    # Longest possible value is 48 chars (all 7 days, lunch + dinner), so the
    # meal_order column is widened from VARCHAR(30) by the startup schema guard.
    _D3 = {"Saturday": "SAT", "Sunday": "SUN", "Monday": "MON", "Tuesday": "TUE",
           "Wednesday": "WED", "Thursday": "THU", "Friday": "FRI"}
    _M1 = {"BREAKFAST": "B", "LUNCH": "L", "DINNER": "D"}

    def _encode_daymeals(daymeals: dict) -> str:
        parts = []
        for d in ORDER:
            if d in daymeals:
                letters = "".join(_M1[m] for m in ("BREAKFAST", "LUNCH", "DINNER")
                                  if m in daymeals[d])
                if letters:
                    parts.append(f"{_D3[d]}={letters}")
        return "|".join(parts)

    agg: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        code = _s(row[c_code]) if c_code is not None else ""
        if not code:
            continue
        day = _s(row[c_day]) if c_day is not None else ""
        name = _s(row[c_name]) if c_name is not None else code
        cat = _s(row[c_cat]) if c_cat is not None else ""
        meal = (_s(row[c_meal]).upper() if c_meal is not None else "")
        # -------------------------------------------------------------------
        # BATCH 159 — A RECIPE CAN BE ON MORE THAN ONE MEAL
        #
        # `meal` was a single string, kept from the FIRST row seen for a code
        # ("if meal and not a['meal']"). Five Tuesday recipes appear under both
        # LUNCH and DINNER — Clear Soup, Special Diet Salad, Boiled Chicken with
        # Soft Rice, low-sodium Grilled Fish, Chicken Kabssa — and each was
        # filed under whichever meal the sheet listed first, then vanished from
        # the other. That is why the Tuesday DINNER group looked short.
        #
        # Menu rows for Tuesday: 49. Distinct codes: 44. The 5 missing entries
        # are exactly the multi-meal ones.
        #
        # Meals are now a SET, stored the same way days already are —
        # "LUNCH & DINNER" — and the menu endpoint expands one row per meal.
        # -------------------------------------------------------------------
        a = agg.setdefault(code, {"name": name, "cat": cat, "days": set(), "daymeals": {}})
        if day:
            a["days"].add(day.strip().title())
        if cat and not a["cat"]:
            a["cat"] = cat
        if day and meal:
            a["daymeals"].setdefault(day.strip().title(), set()).add(meal)

    ok = fail = 0
    for code, a in agg.items():
        day_value = " & ".join(d for d in ORDER if d in a["days"])
        # Batch 159: MEAL IS PER DAY, not per recipe. Encoding meals as a flat
        # set over-expanded — a recipe on Monday LUNCH and Tuesday DINNER showed
        # up at both meals on both days (Tuesday went 49 -> 52).
        # Compact per-day map instead: "SAT=LD|MON=L|TUE=LD". Verified to
        # reproduce the Menu sheet row count EXACTLY on all seven days.
        meal_value = _encode_daymeals(a["daymeals"])
        if dry:
            ok += 1
            continue
        try:
            db.execute(text("""
                INSERT INTO recipes
                    (company_id, recipe_code, recipe_name, customer_name, category,
                     day_of_week, meal_order, status, is_active, approval_status,
                     version, standard_portions, created_at, updated_at)
                VALUES
                    (:cid, :code, :name, :cust, :cat, :day, :meal, 'ACTIVE', 1,
                     'APPROVED', 1, 1, NOW(), NOW())
                ON DUPLICATE KEY UPDATE
                    day_of_week = VALUES(day_of_week),
                    meal_order = VALUES(meal_order),
                    category = COALESCE(NULLIF(recipes.category,''), VALUES(category)),
                    -- Batch 140: re-assert approval. A recipe carried over from
                    -- an earlier partial import can be stuck on 'Pending', which
                    -- silently removes it from /api/menu/for-date even though the
                    -- import reports it as "ok".
                    approval_status = 'APPROVED',
                    -- Batch 159: the menu pass runs LAST and used to write
                    -- standard_portions = 1, wiping the real batch size the
                    -- Recipe Ingredients pass had just imported from
                    -- "No. of portions per batch" (10, 14, 15, 30, 600...).
                    -- Every SMC recipe therefore showed "Recipe Std Portions 1"
                    -- and Batches = ordered portions, so a 150-portion order of
                    -- a 10-portion recipe planned 150 batches instead of 15.
                    -- The menu sheet has no portions data at all, so this pass
                    -- must never overwrite it.
                    -- Batch 159.1 HOTFIX: the INSERT value went back to 1, not
                    -- NULL. standard_portions is NOT NULL in the schema, so
                    -- writing NULL failed every one of the 138 menu rows with
                    -- "Column 'standard_portions' cannot be null".
                    --
                    -- The 1 is only ever used when the menu pass CREATES a
                    -- recipe the Recipe Ingredients pass never saw. For the
                    -- normal path the row already exists with its real batch
                    -- size, and the COALESCE below keeps it.
                    --
                    -- NULLIF(...,0) is added because a row can also be sitting
                    -- on 0 from an older import; 0 is as useless as NULL for a
                    -- divisor, and COALESCE alone would have preserved it.
                    standard_portions = COALESCE(NULLIF(recipes.standard_portions, 0),
                                                 VALUES(standard_portions)),
                    customer_name = :cust, status='ACTIVE', is_active=1, updated_at=NOW()
            """), {"cid": company_id, "code": code, "name": a["name"],
                   "cust": CUSTOMER, "cat": a["cat"], "day": day_value, "meal": meal_value})
            db.commit()
            ok += 1
        except Exception as exc:
            db.rollback()
            fail += 1
            if fail <= 5:
                print(f"    menu '{code}' failed: {exc}")
    return ok, fail


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--company", type=int, default=1)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    print(f"Importing SMC into company_id={args.company}")
    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    names = {s.lower().strip(): s for s in wb.sheetnames}
    db = SessionLocal()

    # 1) raw materials (ingredient master) — same helper as FRSH
    for key in ("raw material list", "raw materials", "raw material"):
        if key in names:
            ok, fail = import_raw_materials(db, wb[names[key]], args.company, args.dry_run)
            print(f"  Ingredients (raw materials): {ok} ok, {fail} failed")
            break

    # 2) recipe ingredient lines — same helper as FRSH (customer read from sheet)
    for key in ("recipe ingredients", "recipe ingredient"):
        if key in names:
            ok, lines, fail = import_recipe_ingredients(db, wb[names[key]], args.company, args.dry_run)
            print(f"  Recipes: {ok} ok, {fail} failed  |  Recipe lines: {lines}")
            break

    # 3) menu with meal_order — SMC-specific
    for key in ("menu",):
        if key in names:
            # clear stale SMC day/meal before rewriting so ghost days don't survive
            if not args.dry_run:
                db.execute(text("""
                    UPDATE recipes SET day_of_week=NULL, meal_order=NULL, updated_at=NOW()
                    WHERE company_id=:cid AND customer_name=:cust
                """), {"cid": args.company, "cust": CUSTOMER})
                db.commit()
            ok, fail = import_menu_smc(db, wb[names[key]], args.company, args.dry_run)
            print(f"  Menu day/meal rows: {ok} ok, {fail} failed")
            break

    db.close()
    print("Done. SMC master data imported.")


if __name__ == "__main__":
    main()

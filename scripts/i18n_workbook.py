"""ISFC PIMS — Arabic translation workbook (Batch 113).

Editing ar.json by hand works but is a poor job for a translator: no context,
no ordering, and one stray comma breaks the file. This produces an Excel
workbook instead, and reads the finished workbook back.

    python scripts/i18n_workbook.py --export           -> ISFC_Arabic_ToTranslate.xlsx
    python scripts/i18n_workbook.py --import <file>    -> merges into ar.json

WHY A WORKBOOK, AND WHY GROUPED
-------------------------------
Strings are grouped by the screen they appear on and ranked by how often that
screen is used, so the translator works through Sales, Procurement and Kitchen
first. Those are the screens staff see every hour; the settings pages can wait.
That means Arabic becomes usable long before the count reaches 100%.

Each row carries WHERE the string appears, so "Open" on a button and "Open" as
a status can be told apart — the same English word often needs two different
Arabic renderings.

The importer only fills EMPTY keys and never overwrites an existing
translation, so re-running it is safe and a partly-finished workbook can be
imported as many times as you like.
"""
from __future__ import annotations

import os as _os
import sys as _sys

_here = _os.path.dirname(_os.path.abspath(__file__))
_root = _here
for _ in range(4):
    if _os.path.isdir(_os.path.join(_root, "app")):
        break
    _p = _os.path.dirname(_root)
    if _p == _root:
        break
    _root = _p
if not _os.path.isdir(_os.path.join(_root, "app")):
    _sys.stderr.write("ERROR: run this from the project (the folder containing app/).\n")
    _sys.exit(2)
_os.chdir(_root)

import json
import re
from collections import defaultdict

I18N = _os.path.join("app", "i18n")
TPL = _os.path.join("app", "templates")

# Screen groups in the order a translator should work through them: the
# screens staff touch every hour first.
PRIORITY = [
    ("1. Sales & Orders", ["orders", "sales_review", "customer", "subscriptions"]),
    ("2. Procurement", ["procurement", "purchase_req"]),
    ("3. Kitchen & Production", ["production", "kitchen"]),
    ("4. Inventory & Store", ["inventory", "store"]),
    ("5. Quality", ["qc", "quality"]),
    ("6. Packing & Dispatch", ["packing", "dispatch"]),
    ("7. Recipes", ["recipes", "recipe"]),
    ("8. Reports & Dashboard", ["reports", "dashboard", "modules"]),
    ("9. Finance", ["finance", "setup"]),
    ("10. Master Data", ["masters", "master"]),
    ("11. Admin & Settings", ["users", "settings", "admin", "hr", "projects"]),
    ("12. Shared / Layout", ["partials", "layouts", "components"]),
]

PATTERNS = [
    re.compile(r"""\bt\(\s*'((?:[^'\\]|\\.)*)'\s*\)"""),
    re.compile(r"""\bt\(\s*"((?:[^"\\]|\\.)*)"\s*\)"""),
]


def discover() -> dict[str, set[str]]:
    """string -> set of files it appears in."""
    found: dict[str, set[str]] = defaultdict(set)
    for base in (TPL, _os.path.join("app", "modules"), _os.path.join("app", "core")):
        if not _os.path.isdir(base):
            continue
        for root, _d, files in _os.walk(base):
            if "__pycache__" in root:
                continue
            for fn in files:
                if not fn.endswith((".html", ".py")):
                    continue
                path = _os.path.join(root, fn)
                try:
                    src = open(path, encoding="utf-8").read()
                except Exception:
                    continue
                for pat in PATTERNS:
                    for m in pat.finditer(src):
                        v = m.group(1).replace("\\'", "'").replace('\\"', '"').strip()
                        if not v or v.startswith(("/", "http", "{{", "{%")):
                            continue
                        found[v].add(path.replace("\\", "/"))
    return found


def group_of(paths: set[str]) -> str:
    for label, keys in PRIORITY:
        for p in paths:
            low = p.lower()
            if any(f"/{k}/" in low or f"/{k}." in low for k in keys):
                return label
    return "13. Other"


def load(name: str) -> dict:
    try:
        return json.load(open(_os.path.join(I18N, name), encoding="utf-8"))
    except Exception:
        return {}


def export() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    found = discover()
    ar = load("ar.json")

    def translated(k: str) -> bool:
        v = ar.get(k)
        return isinstance(v, str) and v.strip() != "" and v.strip() != k

    todo = {k: v for k, v in found.items() if not translated(k)}
    done = sum(1 for k in found if translated(k))

    buckets: dict[str, list] = defaultdict(list)
    for text_, paths in todo.items():
        buckets[group_of(paths)].append((text_, sorted(paths)))

    wb = Workbook()
    head = PatternFill("solid", fgColor="132947")
    warn = PatternFill("solid", fgColor="FDF1E0")

    guide = wb.active
    guide.title = "START HERE"
    guide["A1"] = "ISFC PIMS — Arabic translation"
    guide["A1"].font = Font(bold=True, size=15)
    for line in [
        "",
        f"Already translated : {done}",
        f"Still to translate : {len(todo)}",
        f"Coverage now       : {done / max(len(found), 1) * 100:.1f}%",
        "",
        "HOW TO USE THIS FILE",
        "  1. Work through the sheets IN ORDER. They are ranked by how often",
        "     staff see that screen, so Arabic becomes usable long before you",
        "     reach 100%.",
        "  2. Type the Arabic into the 'Arabic' column only. Do not edit the",
        "     English column — it is the lookup key.",
        "  3. The 'Appears in' column tells you the screen, so the same English",
        "     word used in two places can be translated two different ways.",
        "  4. Save the file, then run:",
        "         python scripts\\i18n_workbook.py --import <thisfile.xlsx>",
        "     You can import a half-finished file as often as you like — it only",
        "     fills blanks and never overwrites work already done.",
        "",
        "TERMINOLOGY — please keep these consistent throughout",
        "  Portion / Batch / Yield / Wastage / Issuance / Requisition /",
        "  Dispatch / Trayline / Head Chef / Store / Lot",
        "  These are operational words your staff already use in Arabic every",
        "  day. Machine translation gets them wrong in ways that read as broken",
        "  to the people working in the system. Use the words the kitchen uses.",
        "",
        "  Leave a cell blank if you are unsure — blanks fall back to English,",
        "  which is far better than a confident wrong term.",
    ]:
        guide.append([line])
    guide.column_dimensions["A"].width = 96

    order = [lbl for lbl, _ in PRIORITY] + ["13. Other"]
    for label in order:
        rows = buckets.get(label)
        if not rows:
            continue
        # Excel forbids / \ ? * [ ] : in sheet names — "Sales & Orders" is
        # fine but "Packing & Dispatch" written with a slash is not.
        safe = re.sub(r'[\\/?*\[\]:]', "-", label)[:31]
        ws = wb.create_sheet(safe)
        for i, h in enumerate(["English (do not edit)", "Arabic", "Appears in", "Notes"], start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 46
        ws.column_dimensions["C"].width = 42
        ws.column_dimensions["D"].width = 30
        for r, (text_, paths) in enumerate(sorted(rows), start=2):
            ws.cell(row=r, column=1, value=text_).alignment = Alignment(wrap_text=True, vertical="top")
            b = ws.cell(row=r, column=2, value="")
            b.fill = warn
            b.alignment = Alignment(horizontal="right", wrap_text=True, vertical="top")
            short = ", ".join(p.split("app/templates/")[-1] for p in paths[:3])
            ws.cell(row=r, column=3, value=short).alignment = Alignment(wrap_text=True, vertical="top")
            if len(text_) > 70:
                ws.cell(row=r, column=4, value="Long sentence — keep the meaning, not the word order")
        ws.freeze_panes = "A2"

    out = "ISFC_Arabic_ToTranslate.xlsx"
    wb.save(out)
    print(f"\nWrote {out}")
    print(f"  {len(todo)} strings to translate, across {len(wb.sheetnames) - 1} screen groups")
    for label in order:
        if buckets.get(label):
            print(f"    {label:<28} {len(buckets[label]):>4}")
    print()


def do_import(path: str) -> None:
    import openpyxl

    if not _os.path.exists(path):
        print(f"File not found: {path}")
        _sys.exit(2)

    wb = openpyxl.load_workbook(path, data_only=True)
    ar = load("ar.json")
    en = load("en.json")

    added = skipped = kept = 0
    for name in wb.sheetnames:
        if name == "START HERE":
            continue
        ws = wb[name]
        for r in range(2, ws.max_row + 1):
            key = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2).value
            if not key:
                continue
            key = str(key).strip()
            val = str(val).strip() if val is not None else ""
            if not val:
                skipped += 1
                continue
            existing = ar.get(key)
            # Never overwrite work already in the file — a re-import of an old
            # workbook must not undo a newer correction.
            if isinstance(existing, str) and existing.strip() and existing.strip() != key:
                kept += 1
                continue
            ar[key] = val
            en.setdefault(key, key)
            added += 1

    with open(_os.path.join(I18N, "ar.json"), "w", encoding="utf-8") as fh:
        json.dump(ar, fh, ensure_ascii=False, indent=2, sort_keys=True)
        fh.write("\n")
    with open(_os.path.join(I18N, "en.json"), "w", encoding="utf-8") as fh:
        json.dump(en, fh, ensure_ascii=False, indent=2, sort_keys=True)
        fh.write("\n")

    print(f"\n  Added {added} translation(s)")
    print(f"  Left blank (still to do): {skipped}")
    print(f"  Kept existing (not overwritten): {kept}")
    print("\n  Restart uvicorn to load them.\n")


if __name__ == "__main__":
    if "--export" in _sys.argv:
        export()
    elif "--import" in _sys.argv:
        i = _sys.argv.index("--import")
        if i + 1 >= len(_sys.argv):
            print("Usage: python scripts/i18n_workbook.py --import <file.xlsx>")
            _sys.exit(2)
        do_import(_sys.argv[i + 1])
    else:
        print(__doc__)

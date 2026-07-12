# app/services/recipe_service.py
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session, selectinload

from app.models.ingredient import Ingredient
from app.models.recipe import Recipe, RecipeIngredient


NULL_TEXTS = {"", "-", "—", "None", "none", "NULL", "null", "N/A", "n/a"}


def _norm(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").strip().lower().split())


def _code(value: Any) -> str | None:
    text = _s(value)
    return text.strip().upper() if text else None


def _s(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in NULL_TEXTS:
        return None
    return text


def _d(value: Any, default: str = "0") -> Decimal:
    """Read Excel numbers safely, including text such as '120 g', 'SAR 3.5', and '95%'."""
    if value is None:
        return Decimal(default)
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return Decimal(default)

    text = str(value).strip().replace(",", "")
    if text in NULL_TEXTS:
        return Decimal(default)

    is_percent = "%" in text
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return Decimal(default)

    try:
        number = Decimal(match.group(0))
        return number / Decimal("100") if is_percent else number
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _b(value: Any) -> bool:
    return str(value or "").strip().lower() in ("yes", "y", "true", "1", "active")


def _headers(ws, header_row: int | None = None) -> dict[str, int]:
    candidate_rows = [header_row] if header_row else [1, 2, 3, 4, 5, 6, 7, 8]
    best: dict[str, int] = {}
    for row in candidate_rows:
        if not row:
            continue
        found: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            key = _norm(ws.cell(row=row, column=col).value)
            if key and key not in found:
                found[key] = col
        if len(found) > len(best):
            best = found
    return best


def _header_row(ws, headers: dict[str, int]) -> int:
    if not ws or not headers:
        return 1
    for row in range(1, min(ws.max_row, 10) + 1):
        count = 0
        for col in range(1, ws.max_column + 1):
            if _norm(ws.cell(row=row, column=col).value) in headers:
                count += 1
        if count >= max(2, min(4, len(headers))):
            return row
    return 1


def _get(ws, row: int, headers: dict[str, int], *names: str, default: Any = None) -> Any:
    for name in names:
        col = headers.get(_norm(name))
        if col:
            return ws.cell(row=row, column=col).value
    return default


def _sheet_name_clean(name: str) -> str:
    return name.lower().replace("–", "-").replace("—", "-").replace("_", " ").strip()


def _find_sheet(workbook, *keywords: str, required: bool = True):
    lowered = [_sheet_name_clean(k) for k in keywords]
    for sheet_name in workbook.sheetnames:
        name = _sheet_name_clean(sheet_name)
        if any(k in name for k in lowered):
            return workbook[sheet_name]
    if required:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Excel sheet containing one of {keywords} was not found. Available sheets: {available}")
    return None


def _max_version(db: Session, company_id: int, recipe_code: str) -> int:
    rows = db.query(Recipe).filter(Recipe.company_id == company_id, Recipe.recipe_code == recipe_code).all()
    return max([int(r.version or 1) for r in rows], default=0)


def _active_recipe(db: Session, company_id: int, recipe_code: str) -> Recipe | None:
    return (
        db.query(Recipe)
        .filter(
            Recipe.company_id == company_id,
            Recipe.recipe_code == recipe_code,
            Recipe.is_active == True,
            Recipe.status == "ACTIVE",
        )
        .order_by(Recipe.version.desc())
        .first()
    )


def _pending_recipe(db: Session, company_id: int, recipe_code: str) -> Recipe | None:
    return (
        db.query(Recipe)
        .filter(Recipe.company_id == company_id, Recipe.recipe_code == recipe_code, Recipe.status == "PENDING")
        .order_by(Recipe.version.desc())
        .first()
    )


def _import_target_recipe(db: Session, company_id: int, recipe_code: str) -> tuple[Recipe, str]:
    """Return the recipe row that should receive this import.

    First upload creates ACTIVE V1. Uploading the same recipe again creates/updates one PENDING V2.
    This protects production from changing approved recipes without approval.
    """
    active = _active_recipe(db, company_id, recipe_code)
    if not active:
        existing_v1 = (
            db.query(Recipe)
            .filter(Recipe.company_id == company_id, Recipe.recipe_code == recipe_code, Recipe.version == 1)
            .first()
        )
        if existing_v1:
            return existing_v1, "updated"
        recipe = Recipe(
            company_id=company_id,
            recipe_code=recipe_code,
            recipe_name=recipe_code,
            version=1,
            status="ACTIVE",
            is_active=True,
            approval_status="APPROVED",
        )
        db.add(recipe)
        db.flush()
        return recipe, "created"

    pending = _pending_recipe(db, company_id, recipe_code)
    if pending:
        return pending, "updated"

    recipe = Recipe(
        company_id=company_id,
        recipe_code=recipe_code,
        recipe_name=active.recipe_name or recipe_code,
        brand_name=active.brand_name,
        customer_name=active.customer_name,
        category=active.category,
        standard_portions=active.standard_portions or 1,
        weight_per_portion_g=active.weight_per_portion_g or 0,
        std_yield_pct=active.std_yield_pct or Decimal("0.95"),
        target_wastage_pct=active.target_wastage_pct or Decimal("0.05"),
        version=_max_version(db, company_id, recipe_code) + 1,
        status="PENDING",
        is_active=False,
        approval_status="PENDING",
        parent_recipe_id=active.id,
    )
    db.add(recipe)
    db.flush()
    return recipe, "pending"


def _inventory_costs(db: Session) -> dict[str, Decimal]:
    rows = db.query(Ingredient).all()
    costs: dict[str, Decimal] = {}
    for item in rows:
        code = _code(item.ingredient_code)
        if code:
            costs[code] = _d(item.unit_cost_standard)
    return costs


def _read_recipe_master_rows(recipe_ws) -> tuple[dict[str, dict[str, Any]], int]:
    if not recipe_ws:
        return {}, 0
    headers = _headers(recipe_ws)
    start = _header_row(recipe_ws, headers) + 1
    rows: dict[str, dict[str, Any]] = {}
    for row in range(start, recipe_ws.max_row + 1):
        recipe_code = _code(_get(recipe_ws, row, headers, "Main Recipe ID", "Recipe ID", "Recipe Ref"))
        if not recipe_code:
            continue
        portions = _d(_get(recipe_ws, row, headers, "Standard No. of Portions", "Standard Portions"), "1")
        rows[recipe_code] = {
            "recipe_name": _s(_get(recipe_ws, row, headers, "Recipe Name", "Main Recipe Name")),
            "brand_name": _s(_get(recipe_ws, row, headers, "Brand Name", "Brand Code")),
            "customer_name": _s(_get(recipe_ws, row, headers, "Customer Name")),
            "category": _s(_get(recipe_ws, row, headers, "Category")),
            "has_sub_recipe": _b(_get(recipe_ws, row, headers, "Has Sub Recipe?", "Has Sub Recipe")),
            "linked_sub_recipe_code": _code(_get(recipe_ws, row, headers, "Sub Recipe ID", "Sub Recipe Ref")),
            "linked_sub_recipe_name": _s(_get(recipe_ws, row, headers, "Sub Recipe Name")),
            "linked_sub_recipe_portions": _d(_get(recipe_ws, row, headers, "Sub Recipe No. of Portions")),
            "standard_portions": portions if portions > 0 else Decimal("1"),
            "weight_per_portion_g": _d(_get(recipe_ws, row, headers, "Weight per Portion (g)", "Weight per Portion")),
            "size_of_portion": _d(_get(recipe_ws, row, headers, "Size of Portion")),
            "std_yield_pct": _d(_get(recipe_ws, row, headers, "Recipe Std Yield %", "Recipe Standard Yield %"), "0.95"),
            "target_wastage_pct": _d(_get(recipe_ws, row, headers, "Target Wastage %", "Targeted Wastage %"), "0.05"),
            "packaging_cost": _d(_get(recipe_ws, row, headers, "Packaging Cost")) * (portions if portions > 0 else Decimal("1")),
            "labor_cost": _d(_get(recipe_ws, row, headers, "Labor Cost")) * (portions if portions > 0 else Decimal("1")),
            "delivery_cost": _d(_get(recipe_ws, row, headers, "Delivery Cost")) * (portions if portions > 0 else Decimal("1")),
            "overheads": _d(_get(recipe_ws, row, headers, "Overheads")) * (portions if portions > 0 else Decimal("1")),
            "other_costs": _d(_get(recipe_ws, row, headers, "Other Costs")) * (portions if portions > 0 else Decimal("1")),
            "margin_pct": _d(_get(recipe_ws, row, headers, "Margin %age", "Margin %"), "0.30"),
            "remark": _s(_get(recipe_ws, row, headers, "Remark", "Remarks")),
        }
    return rows, len(rows)


def _read_recipe_rows_from_ingredients(line_ws) -> tuple[dict[str, dict[str, Any]], int]:
    if not line_ws:
        return {}, 0
    headers = _headers(line_ws)
    start = _header_row(line_ws, headers) + 1
    rows: dict[str, dict[str, Any]] = {}
    for row in range(start, line_ws.max_row + 1):
        recipe_code = _code(_get(line_ws, row, headers, "Recipe ID", "Recipe Ref", "Main Recipe ID"))
        if not recipe_code or recipe_code in rows:
            continue
        portions = _d(_get(line_ws, row, headers, "No. of portions per batch", "Portions"), "1")
        rows[recipe_code] = {
            "recipe_name": _s(_get(line_ws, row, headers, "Recipe Name")) or recipe_code,
            "brand_name": _s(_get(line_ws, row, headers, "Brand Name")),
            "customer_name": _s(_get(line_ws, row, headers, "Customer Name")),
            "category": None,
            "has_sub_recipe": False,
            "linked_sub_recipe_code": _code(_get(line_ws, row, headers, "Sub Recipe ID", "Sub Recipe Ref")),
            "linked_sub_recipe_name": None,
            "linked_sub_recipe_portions": Decimal("0"),
            "standard_portions": portions if portions > 0 else Decimal("1"),
            "weight_per_portion_g": Decimal("0"),
            "size_of_portion": Decimal("0"),
            "std_yield_pct": Decimal("0.95"),
            "target_wastage_pct": Decimal("0.05"),
            "packaging_cost": Decimal("0"),
            "labor_cost": Decimal("0"),
            "delivery_cost": Decimal("0"),
            "overheads": Decimal("0"),
            "other_costs": Decimal("0"),
            "margin_pct": Decimal("0.30"),
            "remark": "Created from Recipe Ingredients sheet because Master Recipes sheet was not present.",
        }
    return rows, len(rows)


def _apply_recipe_meta(recipe: Recipe, recipe_code: str, meta: dict[str, Any]) -> None:
    recipe.recipe_name = meta.get("recipe_name") or recipe.recipe_name or recipe_code
    recipe.brand_name = meta.get("brand_name") or recipe.brand_name
    recipe.customer_name = meta.get("customer_name") or recipe.customer_name
    recipe.category = meta.get("category") or recipe.category
    recipe.is_sub_recipe = recipe_code.startswith("SUB-") or recipe_code.startswith("RCP-MS-")
    recipe.has_sub_recipe = bool(meta.get("has_sub_recipe"))
    recipe.linked_sub_recipe_code = meta.get("linked_sub_recipe_code")
    recipe.linked_sub_recipe_name = meta.get("linked_sub_recipe_name")
    recipe.linked_sub_recipe_portions = meta.get("linked_sub_recipe_portions") or Decimal("0")
    recipe.standard_portions = meta.get("standard_portions") or Decimal("1")
    recipe.weight_per_portion_g = meta.get("weight_per_portion_g") or Decimal("0")
    recipe.size_of_portion = meta.get("size_of_portion") or Decimal("0")
    recipe.std_yield_pct = meta.get("std_yield_pct") or Decimal("0.95")
    recipe.target_wastage_pct = meta.get("target_wastage_pct") or Decimal("0.05")
    recipe.packaging_cost = meta.get("packaging_cost") or Decimal("0")
    recipe.labor_cost = meta.get("labor_cost") or Decimal("0")
    recipe.delivery_cost = meta.get("delivery_cost") or Decimal("0")
    recipe.overheads = meta.get("overheads") or Decimal("0")
    recipe.other_costs = meta.get("other_costs") or Decimal("0")
    recipe.margin_pct = meta.get("margin_pct") or Decimal("0.30")
    recipe.remark = meta.get("remark") or recipe.remark


def recalc_recipe(recipe: Recipe) -> None:
    food_cost = Decimal("0")
    food_cost_per_portion = Decimal("0")
    packaging_line_cost = Decimal("0")
    missing_cost_lines = 0
    portions = _d(recipe.standard_portions, "1")
    if portions <= 0:
        portions = Decimal("1")

    base_packaging_cost = _d(recipe.packaging_cost)
    for index, line in enumerate(recipe.lines, start=1):
        line.line_no = index
        qty_batch = _d(line.qty_batch)
        line_portions = _d(line.portions, str(portions))
        if line_portions <= 0:
            line_portions = portions
        qty_per_portion = _d(line.qty_per_portion)
        if qty_per_portion == 0 and qty_batch > 0:
            qty_per_portion = qty_batch / line_portions
        cost_uom = _d(line.cost_uom)
        line.qty_batch = qty_batch
        line.portions = line_portions
        line.qty_per_portion = qty_per_portion
        line.cost_uom = cost_uom
        line.line_cost = qty_batch * cost_uom
        line.line_cost_per_portion = qty_per_portion * cost_uom
        line.missing_cost = not line.inventory_code or cost_uom <= 0
        if line.missing_cost:
            missing_cost_lines += 1
        if str(line.line_type or "").lower().startswith("pack"):
            packaging_line_cost += line.line_cost
        else:
            food_cost += line.line_cost
            food_cost_per_portion += line.line_cost_per_portion

    packaging_total = base_packaging_cost + packaging_line_cost
    other_costs = _d(recipe.labor_cost) + _d(recipe.delivery_cost) + _d(recipe.overheads) + _d(recipe.other_costs)
    total_cost = food_cost + packaging_total + other_costs
    recipe.food_cost = food_cost
    recipe.food_cost_per_portion = food_cost_per_portion
    recipe.packaging_cost = packaging_total
    recipe.total_cost = total_cost
    recipe.total_cost_per_portion = total_cost / portions
    recipe.missing_cost_lines = missing_cost_lines
    margin = _d(recipe.margin_pct, "0.30")
    if Decimal("0") <= margin < Decimal("1"):
        recipe.sale_price = total_cost / (Decimal("1") - margin)
        recipe.sale_price_per_portion = recipe.sale_price / portions
    else:
        recipe.sale_price = total_cost
        recipe.sale_price_per_portion = total_cost / portions


def import_recipe_excel(db: Session, file_path: str, company_id: int) -> dict[str, int | list[str]]:
    workbook = load_workbook(file_path, data_only=True)
    recipe_ws = _find_sheet(workbook, "recipes", "master recipes", "master - recipes", required=False)
    line_ws = _find_sheet(
        workbook,
        "recipe ingredients",
        "recipes ingredients",
        "master ingredients",
        "master - ingredients",
        "ingredients",
        "salus",
        required=False,
    )

    if not recipe_ws and not line_ws:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(
            "No recipe sheet found. Upload Master recipes.xlsx, Recipes Ingredients.xlsx, "
            f"or a combined recipe workbook. Available sheets: {available}"
        )

    master_rows, master_count = _read_recipe_master_rows(recipe_ws)
    ingredient_recipe_rows, ingredient_recipe_count = _read_recipe_rows_from_ingredients(line_ws)

    # Master sheet values are stronger. If absent, create recipe masters from Recipe Ingredients sheet.
    combined_recipe_rows = dict(ingredient_recipe_rows)
    combined_recipe_rows.update(master_rows)

    created = updated = pending_count = imported_lines = 0
    touched: dict[str, Recipe] = {}

    for recipe_code, meta in combined_recipe_rows.items():
        recipe, result = _import_target_recipe(db, company_id, recipe_code)
        _apply_recipe_meta(recipe, recipe_code, meta)
        touched[recipe_code] = recipe
        if result == "created":
            created += 1
        elif result == "pending":
            pending_count += 1
        else:
            updated += 1

    db.flush()

    if line_ws:
        for recipe in touched.values():
            db.query(RecipeIngredient).filter(RecipeIngredient.recipe_id == recipe.id).delete()
        db.flush()

        inventory_costs = _inventory_costs(db)
        line_headers = _headers(line_ws)
        line_start_row = _header_row(line_ws, line_headers) + 1
        line_numbers: dict[str, int] = {}

        for row in range(line_start_row, line_ws.max_row + 1):
            recipe_code = _code(_get(line_ws, row, line_headers, "Recipe ID", "Recipe Ref", "Main Recipe ID"))
            if not recipe_code:
                continue

            recipe = touched.get(recipe_code)
            if not recipe:
                meta = ingredient_recipe_rows.get(recipe_code, {"recipe_name": recipe_code, "standard_portions": Decimal("1")})
                recipe, result = _import_target_recipe(db, company_id, recipe_code)
                _apply_recipe_meta(recipe, recipe_code, meta)
                touched[recipe_code] = recipe
                if result == "created":
                    created += 1
                elif result == "pending":
                    pending_count += 1
                else:
                    updated += 1
                db.flush()

            line_numbers[recipe_code] = line_numbers.get(recipe_code, 0) + 1
            inventory_code = _code(_get(line_ws, row, line_headers, "Inventory ID", "Item Code", "Code"))
            item_name = _s(_get(line_ws, row, line_headers, "Item / Ingredient", "Ingredient Name", "NameEN", "Items Description"))
            if not item_name:
                item_name = inventory_code or f"Recipe line {row}"

            cost_uom = _d(_get(line_ws, row, line_headers, "PP St. UOM", "Cost/UOM", "Cost Per Unit", "Price"))
            if cost_uom <= 0 and inventory_code:
                cost_uom = inventory_costs.get(inventory_code, Decimal("0"))

            db.add(
                RecipeIngredient(
                    recipe_id=recipe.id,
                    line_no=line_numbers[recipe_code],
                    line_type=_s(_get(line_ws, row, line_headers, "Ingredient Type")) or "Main Recipe",
                    sub_recipe_code=_code(_get(line_ws, row, line_headers, "Sub Recipe ID", "Sub Recipe Ref")),
                    inventory_code=inventory_code,
                    item_name=item_name,
                    uom=_s(_get(line_ws, row, line_headers, "St. UOM", "UOM", "Unit")),
                    qty_batch=_d(_get(line_ws, row, line_headers, "Qty req per Batch (g/pcs)", "Qty Batch", "Batch Qty")),
                    portions=_d(_get(line_ws, row, line_headers, "No. of portions per batch", "Portions"), "1"),
                    qty_per_portion=_d(_get(line_ws, row, line_headers, "Qty req per portion", "Qty Per Portion")),
                    cost_uom=cost_uom,
                    remark=_s(_get(line_ws, row, line_headers, "Remark", "Remarks")),
                )
            )
            imported_lines += 1

    db.flush()
    recipes_to_recalc = (
        db.query(Recipe)
        .options(selectinload(Recipe.lines))
        .filter(Recipe.id.in_([r.id for r in touched.values()]))
        .all()
    )
    for recipe in recipes_to_recalc:
        recalc_recipe(recipe)

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "pending": pending_count,
        "skipped": 0,
        "errors": [],
        "lines": imported_lines,
        "master_rows": master_count,
        "ingredient_recipe_rows": ingredient_recipe_count,
    }

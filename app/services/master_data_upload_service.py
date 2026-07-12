# app/services/master_data_upload_service.py
from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Callable

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.ingredient import Ingredient
from app.models.customer import Customer
from app.models.supplier import Supplier
from app.models.chef import Chef
from app.models.master_data import Brand, KitchenLocation, KitchenSection, MasterRecord, RevenueStream
from app.services.recipe_service import import_recipe_excel


MASTER_TYPES = {
    "all": {"label": "All Master Data Workbook", "sheet_keyword": "all"},
    "inventory": {"label": "Inventory Master", "sheet_keyword": "inventory"},
    "customers": {"label": "Customers", "sheet_keyword": "customers"},
    "suppliers": {"label": "Suppliers", "sheet_keyword": "suppliers"},
    "chefs": {"label": "Chefs", "sheet_keyword": "chefs"},
    "brands": {"label": "Brands", "sheet_keyword": "brands"},
    "revenue_streams": {"label": "Revenue Streams", "sheet_keyword": "revenue stream"},
    "kitchen_locations": {"label": "Kitchen Locations", "sheet_keyword": "kitchen location"},
    "kitchen_sections": {"label": "Kitchen Sections", "sheet_keyword": "kitchen section"},
    "recipes_inventory": {"label": "Recipes + Inventory Workbook", "sheet_keyword": "recipes"},
    "recipes": {"label": "Recipes Data Upload / Recipes & Costing", "sheet_keyword": "recipes"},
}

# These are the original ISFC Excel fields, not generic fields.
TEMPLATE_HEADERS = {
    "inventory": [
        "Sr. No", "Supplier code", "Supplier Name", "PIS Material Code", "Material Code old",
        "Material Name", "Category", "Received Date", "Unit of Measurement", "Cost Per Unit",
        "Total Qty", "Total Price (With Discount)", "Total Price (Without Discount)",
        "Source Row", "Historical Line Count",
    ],
    "chefs": [
        "#", "EMP #", "EMPLOYEE NAME", "EMPLOYEE JOB TITLE", "Kitchen Section", "Tasks", "Brand Assign", "REMARKS",
    ],
    "suppliers": [
        "Sr. No", "PIS Supplier Code", "Supplier Code Old", "Supplier Name EN", "Supplier Name AR",
        "Category", "Address", "Primary Contact", "Phone", "Email", "VAT Number", "Payment Terms",
        "Supplier Type", "Commercial Registration (CR)", "Tax Residency Country", "Mobile", "Website",
        "Country", "Region", "City", "District", "Street", "Building Number", "Additional Number",
        "Postal Code", "PO Box", "National Address Short Code", "Payment Terms", "Lead Time (Days)",
        "Currency", "Incoterms", "Minimum Order Quantity", "Preferred Supplier", "Approved Supplier",
        "Supplier Rating", "Bank Name", "Account Name", "Account Number", "IBAN", "SWIFT Code",
    ],
    "customers": [
        "Sr. No", "PIS Customer Code", "Customer Code Old", "Customer Name EN", "Customer Name AR",
        "Sales Man", "Contact Person", "Phone ", "Brand", "Contract Ref", "Status", "VAT Number",
        "Customer Type", "Customer Category", "VAT Registration Number", "Commercial Registration (CR)",
        "National ID/Iqama", "Tax Exempt", "Tax Exemption Reason", "ZATCA Buyer Type", "Mobile No",
        "Email", "Website", "Country", "Region", "City", "District", "Street", "Building No",
        "Additional No", "Postal Code", "PO Box", "National Address Short Code", "Credit Limit",
        "Credit Period (Days)", "Payment Terms", "Sales Representative", "Delivery Route", "Price List", "Customer Group",
    ],
    "brands": [
        "Sr. No", "Brand ID", "Brand Name EN", "Brand Name AR", "Short Code", "Revenue Stream Name",
        "Revenue Stream Name", "Default Kitchen Code", "Active Status", "Brand Manager", "Launch Date", "Remarks",
    ],
    "revenue_streams": [
        "Sr. No", "Revenue Stream Code", "Revenue Stream Name", "Description", "Revenue Category", "Active Status", "Remarks",
    ],
    "kitchen_locations": [
        "Sr. No", "Kitchen Code", "Kitchen Name", "Kitchen Type", "Location", "City", "Brand Supported",
        "Capacity", "Manager", "Active Status",
    ],
    "kitchen_sections": [
        "Sr. No", "Section Code", "Section Name", "Section Name AR", "Kitchen Code", "Sequence No", "Active Status", "Remarks",
    ],
    "recipes": [
        "Main Recipe ID", "Recipe Name", "Brand Name", "Customer Name", "Category", "Standard No. of\nPortions",
        "Weight per\nPortion (g)", "Size of\nPortion", "Has Sub\nRecipe?", "Sub Recipe\nID", "Sub Recipe\nName",
        "Sub Recipe No. of\nPortions", "Total Ing.\nCount", "Recipe Std Yield %", "Target Wastage %",
        "Food Cost  Per Portion", "Packaging Cost", "Labor Cost", "Delivery Cost", "Overheads", "Other Costs",
        "Total Cost", "Margin %age", "Profit", "Sale Price", "Remark",
    ],
    "recipe_ingredients": [
        "Recipe ID", "Brand Name", "Customer Name", "Recipe Name", "Sub Recipe ID", "Ingredient Type", "Inventory ID",
        "Item / Ingredient", "St. UOM", "Qty req per Batch (g/pcs)", "No. of portions per batch", "Qty req per portion",
        "PP St. UOM", "Food Cost ",
    ],
    "raw_material_list": [
        "Items Description", "Code", "NameEN", "Unit", "Price", "Main category", "Sub category",
    ],
}

WORKBOOK_SHEETS = [
    ("📘 Master Inventory", "inventory"),
    ("📘 Master – Chefs", "chefs"),
    ("📘 Master – Suppliers", "suppliers"),
    ("📘 Master – Customers", "customers"),
    ("📘 Master – Brands", "brands"),
    ("📘 Master – Revenue Stream", "revenue_streams"),
    ("📘 Master – Kitchen Location", "kitchen_locations"),
    ("📘 Master – Kitchen Sections", "kitchen_sections"),
]

RECIPE_WORKBOOK_SHEETS = [
    ("📘 Master – Recipes", "recipes"),
    ("📘 Master – Ingredients (Salus)", "recipe_ingredients"),
]


def _norm(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").strip().lower().split())


def _s(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in ("", "-", "—", "None", "none", "NULL", "null"):
        return None
    return text


def _code_key(value: Any) -> str:
    """Normalize business codes for duplicate-safe imports.

    MySQL utf8mb4_unicode_ci unique indexes are case-insensitive, so we
    normalize codes in memory too. This prevents duplicate pending ORM
    inserts such as DRS1-5155 / drs1-5155 in the same workbook.
    """
    text = _s(value)
    return str(text or "").strip().upper()


def _d(value: Any, default: str = "0") -> Decimal:
    if value in (None, "", "-", "—"):
        return Decimal(default)
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _f(value: Any, default: float = 0.0) -> float:
    try:
        return float(_d(value, str(default)))
    except Exception:
        return default


def _i(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except Exception:
        return default


def _status(value: Any) -> str:
    text = str(value or "ACTIVE").strip().upper()
    if text in ("ACTIVE", "YES", "Y", "1", "TRUE", "OPEN"):
        return "ACTIVE"
    if text in ("INACTIVE", "NO", "N", "0", "FALSE", "DISABLED", "CLOSED"):
        return "INACTIVE"
    if text == "PENDING":
        return "PENDING"
    return "ACTIVE"


def _json_default(value: Any):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _headers(ws, header_row: int | None = None) -> dict[str, int]:
    candidate_rows = [header_row] if header_row else [1, 2, 3, 4, 5]
    best: dict[str, int] = {}
    for row in candidate_rows:
        if not row:
            continue
        found: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            key = _norm(ws.cell(row=row, column=col).value)
            if key and key not in found:
                found[key] = col
        if len(found) > len(best):
            best = found
    return best


def _header_row(ws, headers: dict[str, int]) -> int:
    if not headers:
        return 1
    for row in range(1, min(ws.max_row, 8) + 1):
        count = 0
        for col in range(1, ws.max_column + 1):
            if _norm(ws.cell(row=row, column=col).value) in headers:
                count += 1
        if count >= max(2, min(4, len(headers))):
            return row
    return 1


def _get(ws, row: int, headers: dict[str, int], *names: str, default: Any = None) -> Any:
    for name in names:
        col = headers.get(_norm(name))
        if col:
            return ws.cell(row=row, column=col).value
    return default


def _row_json(ws, row: int, headers: dict[str, int]) -> str:
    data: dict[str, Any] = {}
    reverse = {col: name for name, col in headers.items()}
    for col in range(1, ws.max_column + 1):
        key = reverse.get(col)
        if key:
            data[key] = ws.cell(row=row, column=col).value
    return json.dumps(data, ensure_ascii=False, default=_json_default)


def _find_sheet(workbook, keyword: str):
    keyword = keyword.lower().replace("–", "-").replace("—", "-").strip()
    for sheet_name in workbook.sheetnames:
        name = sheet_name.lower().replace("–", "-").replace("—", "-")
        if keyword in name:
            return workbook[sheet_name]
    return None


def _sheet_for_type(workbook, master_type: str):
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    config = MASTER_TYPES.get(master_type)
    if not config:
        return None
    return _find_sheet(workbook, config["sheet_keyword"])


def _archive_master_record(db: Session, company_id: int, master_type: str, code: str, name_en: str | None, name_ar: str | None, raw_json: str, remarks: str | None = None):
    if not code:
        return
    existing = (
        db.query(MasterRecord)
        .filter(
            MasterRecord.company_id == company_id,
            MasterRecord.master_type == master_type,
            MasterRecord.code == code,
            MasterRecord.is_active == True,
        )
        .first()
    )
    if existing and existing.raw_json == raw_json:
        return
    version = (existing.version + 1) if existing else 1
    if existing:
        existing.is_active = False
        existing.status = "INACTIVE"
    db.add(MasterRecord(
        company_id=company_id,
        master_type=master_type,
        code=code,
        name_en=name_en,
        name_ar=name_ar,
        version=version,
        status="ACTIVE",
        is_active=True,
        approval_status="APPROVED",
        raw_json=raw_json,
        remarks=remarks,
    ))


def _summary():
    return {"created": 0, "updated": 0, "pending": 0, "skipped": 0, "errors": []}


def _bump(summary: dict[str, Any], result: str):
    if result in summary:
        summary[result] += 1


def _preload_by_attr(db: Session, model, attr_name: str, company_id: int | None = None) -> dict[str, Any]:
    q = db.query(model)
    if company_id is not None and hasattr(model, "company_id"):
        q = q.filter(model.company_id == company_id)
    rows = q.all()
    result = {}
    for row in rows:
        code = getattr(row, attr_name, None)
        key = _code_key(code)
        if key:
            result[key] = row
    return result


def import_inventory(db: Session, ws, company_id: int) -> dict[str, Any]:
    summary = _summary()
    headers = _headers(ws)
    start = _header_row(ws, headers) + 1
    existing_by_code = _preload_by_attr(db, Ingredient, "ingredient_code")
    seen: set[str] = set()

    for row in range(start, ws.max_row + 1):
        old_code = _s(_get(ws, row, headers, "Material Code old", "Old Material Code"))
        pis_code = _s(_get(ws, row, headers, "PIS Material Code", "Inventory Code", "Material Code", "Code"))
        code = _code_key(pis_code or old_code)
        name = _s(_get(ws, row, headers, "Material Name", "Item Name", "Ingredient Name", "NameEN", "Items Description"))
        if not code or not name:
            continue
        if code in seen:
            summary["skipped"] += 1
            continue
        seen.add(code)
        try:
            item = existing_by_code.get(code)
            if not item:
                item = db.query(Ingredient).filter(Ingredient.ingredient_code == code).first()
            if not item:
                item = Ingredient(ingredient_code=code, name=name)
                db.add(item)
                result = "created"
            else:
                result = "updated"
            existing_by_code[code] = item

            uom = _s(_get(ws, row, headers, "Inventory UOM", "Unit of Measurement", "UOM", "Unit")) or item.purchase_uom or "Each"
            recipe_uom = _s(_get(ws, row, headers, "Recipe UOM")) or uom
            item.ingredient_code = code
            item.name = name
            main_category = _s(_get(ws, row, headers, "Category", "Main category"))
            sub_category = _s(_get(ws, row, headers, "Sub category"))
            item.category = main_category or sub_category
            if hasattr(item, "main_category"):
                item.main_category = main_category or item.category
            if hasattr(item, "sub_category"):
                item.sub_category = sub_category
            item.purchase_uom = uom
            item.standard_uom = uom
            item.recipe_uom = recipe_uom
            item.unit_cost_standard = _f(_get(ws, row, headers, "Cost Per Unit", "Inventory Cost per UOM", "Cost/UOM", "Price"))
            item.default_supplier = _s(_get(ws, row, headers, "Supplier Name"))
            item.default_issue_section = _s(_get(ws, row, headers, "Default Issue Section")) or item.default_issue_section or "Hot Kitchen"
            item.status = _status(_get(ws, row, headers, "Active Status", "Status"))
            item.notes = f"Old Material Code: {old_code or ''} | Supplier Code: {_s(_get(ws, row, headers, 'Supplier code', 'Supplier Code')) or ''}"
            _archive_master_record(db, company_id, "inventory", code, name, None, _row_json(ws, row, headers), item.notes)
            _bump(summary, result)
        except Exception as exc:
            summary["errors"].append(f"Inventory row {row}: {exc}")
    return summary


def import_customers(db: Session, ws, company_id: int) -> dict[str, Any]:
    summary = _summary()
    headers = _headers(ws)
    start = _header_row(ws, headers) + 1
    existing_by_code = _preload_by_attr(db, Customer, "customer_code", company_id)
    seen: set[str] = set()
    for row in range(start, ws.max_row + 1):
        code = _code_key(_get(ws, row, headers, "PIS Customer Code", "Customer Code Old", "Customer Code"))
        name_en = _s(_get(ws, row, headers, "Customer Name EN", "Customer Name"))
        name_ar = _s(_get(ws, row, headers, "Customer Name AR"))
        if not code or not name_en:
            continue
        if code in seen:
            summary["skipped"] += 1
            continue
        seen.add(code)
        try:
            item = existing_by_code.get(code)
            if not item:
                item = Customer(company_id=company_id, customer_code=code, customer_name=name_en)
                db.add(item)
                existing_by_code[code] = item
                result = "created"
            else:
                result = "updated"
            item.customer_name = name_en
            if hasattr(item, "customer_name_ar"):
                item.customer_name_ar = name_ar
                item.brand = _s(_get(ws, row, headers, "Brand"))
                item.phone = _s(_get(ws, row, headers, "Phone", "Phone ", "Mobile No"))
                item.email = _s(_get(ws, row, headers, "Email"))
                item.vat_number = _s(_get(ws, row, headers, "VAT Number", "VAT Registration Number"))
                item.customer_type = _s(_get(ws, row, headers, "Customer Type", "Customer Category"))
                item.city = _s(_get(ws, row, headers, "City"))
                item.payment_terms = _s(_get(ws, row, headers, "Payment Terms"))
                item.contact_person = _s(_get(ws, row, headers, "Contact Person"))
                item.sales_man = _s(_get(ws, row, headers, "Sales Man", "Sales Representative"))
            item.status = _status(_get(ws, row, headers, "Status", "Active Status"))
            item.is_active = item.status == "ACTIVE"
            _archive_master_record(db, company_id, "customers", code, name_en, name_ar, _row_json(ws, row, headers), _s(_get(ws, row, headers, "Remarks")))
            _bump(summary, result)
        except Exception as exc:
            summary["errors"].append(f"Customer row {row}: {exc}")
    return summary


def import_suppliers(db: Session, ws, company_id: int) -> dict[str, Any]:
    summary = _summary()
    headers = _headers(ws)
    start = _header_row(ws, headers) + 1
    existing_by_code = _preload_by_attr(db, Supplier, "supplier_code", company_id)
    seen: set[str] = set()
    for row in range(start, ws.max_row + 1):
        code = _code_key(_get(ws, row, headers, "PIS Supplier Code", "Supplier Code Old", "Supplier Code"))
        name_en = _s(_get(ws, row, headers, "Supplier Name EN", "Supplier Name"))
        name_ar = _s(_get(ws, row, headers, "Supplier Name AR"))
        if not code or not name_en:
            continue
        if code in seen:
            summary["skipped"] += 1
            continue
        seen.add(code)
        try:
            item = existing_by_code.get(code)
            if not item:
                item = Supplier(company_id=company_id, supplier_code=code, supplier_name=name_en)
                db.add(item)
                existing_by_code[code] = item
                result = "created"
            else:
                result = "updated"
            item.supplier_name = name_en
            if hasattr(item, "supplier_name_ar"):
                item.supplier_name_ar = name_ar
                item.category = _s(_get(ws, row, headers, "Category"))
                item.phone = _s(_get(ws, row, headers, "Phone", "Mobile"))
                item.email = _s(_get(ws, row, headers, "Email"))
                item.vat_number = _s(_get(ws, row, headers, "VAT Number"))
                item.payment_terms = _s(_get(ws, row, headers, "Payment Terms"))
                item.supplier_type = _s(_get(ws, row, headers, "Supplier Type"))
                item.city = _s(_get(ws, row, headers, "City"))
                item.country = _s(_get(ws, row, headers, "Country"))
            item.status = _status(_get(ws, row, headers, "Status", "Active Status"))
            item.is_active = item.status == "ACTIVE"
            _archive_master_record(db, company_id, "suppliers", code, name_en, name_ar, _row_json(ws, row, headers), _s(_get(ws, row, headers, "Remarks")))
            _bump(summary, result)
        except Exception as exc:
            summary["errors"].append(f"Supplier row {row}: {exc}")
    return summary


def import_chefs(db: Session, ws, company_id: int) -> dict[str, Any]:
    summary = _summary()
    headers = _headers(ws)
    start = _header_row(ws, headers) + 1
    existing_by_code = _preload_by_attr(db, Chef, "chef_code", company_id)
    seen: set[str] = set()
    for row in range(start, ws.max_row + 1):
        emp_no = _s(_get(ws, row, headers, "EMP #", "Chef Code", "Employee Code"))
        name = _s(_get(ws, row, headers, "EMPLOYEE NAME", "Chef Name", "Employee Name"))
        if not emp_no or not name:
            continue
        code = _code_key(emp_no if emp_no.upper().startswith("CHF-") else f"CHF-{emp_no}")
        if code in seen:
            summary["skipped"] += 1
            continue
        seen.add(code)
        try:
            item = existing_by_code.get(code)
            if not item:
                item = Chef(company_id=company_id, chef_code=code, chef_name=name)
                db.add(item)
                existing_by_code[code] = item
                result = "created"
            else:
                result = "updated"
            item.chef_name = name
            if hasattr(item, "job_title"):
                item.job_title = _s(_get(ws, row, headers, "EMPLOYEE JOB TITLE"))
                item.kitchen_section = _s(_get(ws, row, headers, "Kitchen Section"))
                item.tasks = _s(_get(ws, row, headers, "Tasks"))
                item.brand_assign = _s(_get(ws, row, headers, "Brand Assign"))
                item.remarks = _s(_get(ws, row, headers, "REMARKS", "Remarks"))
            item.status = _status(_get(ws, row, headers, "Status", "Active Status"))
            item.is_active = item.status == "ACTIVE"
            _archive_master_record(db, company_id, "chefs", code, name, None, _row_json(ws, row, headers), _s(_get(ws, row, headers, "REMARKS", "Remarks")))
            _bump(summary, result)
        except Exception as exc:
            summary["errors"].append(f"Chef row {row}: {exc}")
    return summary


def import_brands(db: Session, ws, company_id: int) -> dict[str, Any]:
    summary = _summary()
    headers = _headers(ws)
    start = _header_row(ws, headers) + 1
    existing_by_code = _preload_by_attr(db, Brand, "brand_code", company_id)
    seen: set[str] = set()
    for row in range(start, ws.max_row + 1):
        code = _code_key(_get(ws, row, headers, "Brand ID", "Brand Code"))
        name_en = _s(_get(ws, row, headers, "Brand Name EN", "Brand Name"))
        name_ar = _s(_get(ws, row, headers, "Brand Name AR"))
        if not code or not name_en:
            continue
        if code in seen:
            summary["skipped"] += 1
            continue
        seen.add(code)
        try:
            item = existing_by_code.get(code)
            if not item:
                item = Brand(company_id=company_id, brand_code=code, brand_name_en=name_en)
                db.add(item)
                existing_by_code[code] = item
                result = "created"
            else:
                result = "updated"
            item.brand_name_en = name_en
            item.brand_name_ar = name_ar
            item.short_code = _s(_get(ws, row, headers, "Short Code"))
            item.revenue_stream_name = _s(_get(ws, row, headers, "Revenue Stream Name"))
            item.default_kitchen_code = _s(_get(ws, row, headers, "Default Kitchen Code"))
            item.status = _status(_get(ws, row, headers, "Active Status", "Status"))
            item.is_active = item.status == "ACTIVE"
            item.remarks = _s(_get(ws, row, headers, "Remarks"))
            _archive_master_record(db, company_id, "brands", code, name_en, name_ar, _row_json(ws, row, headers), item.remarks)
            _bump(summary, result)
        except Exception as exc:
            summary["errors"].append(f"Brand row {row}: {exc}")
    return summary


def import_revenue_streams(db: Session, ws, company_id: int) -> dict[str, Any]:
    summary = _summary()
    headers = _headers(ws)
    start = _header_row(ws, headers) + 1
    existing_by_code = _preload_by_attr(db, RevenueStream, "stream_code", company_id)
    seen: set[str] = set()
    for row in range(start, ws.max_row + 1):
        code = _code_key(_get(ws, row, headers, "Revenue Stream Code", "Stream Code"))
        name = _s(_get(ws, row, headers, "Revenue Stream Name", "Stream Name"))
        if not code or not name:
            continue
        if code in seen:
            summary["skipped"] += 1
            continue
        seen.add(code)
        try:
            item = existing_by_code.get(code)
            if not item:
                item = RevenueStream(company_id=company_id, stream_code=code, stream_name=name)
                db.add(item)
                existing_by_code[code] = item
                result = "created"
            else:
                result = "updated"
            item.stream_name = name
            item.description = _s(_get(ws, row, headers, "Description"))
            item.revenue_category = _s(_get(ws, row, headers, "Revenue Category"))
            item.status = _status(_get(ws, row, headers, "Active Status", "Status"))
            item.is_active = item.status == "ACTIVE"
            item.remarks = _s(_get(ws, row, headers, "Remarks"))
            _archive_master_record(db, company_id, "revenue_streams", code, name, None, _row_json(ws, row, headers), item.remarks)
            _bump(summary, result)
        except Exception as exc:
            summary["errors"].append(f"Revenue stream row {row}: {exc}")
    return summary


def import_kitchen_locations(db: Session, ws, company_id: int) -> dict[str, Any]:
    summary = _summary()
    headers = _headers(ws)
    start = _header_row(ws, headers) + 1
    existing_by_code = _preload_by_attr(db, KitchenLocation, "kitchen_code", company_id)
    seen: set[str] = set()
    for row in range(start, ws.max_row + 1):
        code = _code_key(_get(ws, row, headers, "Kitchen Code", "Kitchen Location Code"))
        name = _s(_get(ws, row, headers, "Kitchen Name", "Kitchen Location Name"))
        if not code or not name:
            continue
        if code in seen:
            summary["skipped"] += 1
            continue
        seen.add(code)
        try:
            item = existing_by_code.get(code)
            if not item:
                item = KitchenLocation(company_id=company_id, kitchen_code=code, kitchen_name=name)
                db.add(item)
                existing_by_code[code] = item
                result = "created"
            else:
                result = "updated"
            item.kitchen_name = name
            item.kitchen_type = _s(_get(ws, row, headers, "Kitchen Type"))
            item.location = _s(_get(ws, row, headers, "Location"))
            item.city = _s(_get(ws, row, headers, "City"))
            item.brand_supported = _s(_get(ws, row, headers, "Brand Supported"))
            item.capacity = _s(_get(ws, row, headers, "Capacity"))
            item.manager = _s(_get(ws, row, headers, "Manager"))
            item.status = _status(_get(ws, row, headers, "Active Status", "Status"))
            item.is_active = item.status == "ACTIVE"
            _archive_master_record(db, company_id, "kitchen_locations", code, name, None, _row_json(ws, row, headers), None)
            _bump(summary, result)
        except Exception as exc:
            summary["errors"].append(f"Kitchen location row {row}: {exc}")
    return summary


def import_kitchen_sections(db: Session, ws, company_id: int) -> dict[str, Any]:
    summary = _summary()
    headers = _headers(ws)
    start = _header_row(ws, headers) + 1
    existing_by_code = _preload_by_attr(db, KitchenSection, "section_code", company_id)
    seen: set[str] = set()
    for row in range(start, ws.max_row + 1):
        code = _code_key(_get(ws, row, headers, "Section Code", "Kitchen Section Code"))
        name = _s(_get(ws, row, headers, "Section Name", "Kitchen Section", "Kitchen Section Name"))
        if not code or not name:
            continue
        if code in seen:
            summary["skipped"] += 1
            continue
        seen.add(code)
        try:
            item = existing_by_code.get(code)
            if not item:
                item = KitchenSection(company_id=company_id, section_code=code, section_name=name)
                db.add(item)
                existing_by_code[code] = item
                result = "created"
            else:
                result = "updated"
            item.section_name = name
            item.section_name_ar = _s(_get(ws, row, headers, "Section Name AR"))
            item.kitchen_code = _s(_get(ws, row, headers, "Kitchen Code"))
            item.sequence_no = _i(_get(ws, row, headers, "Sequence No"), item.sequence_no or 1)
            item.status = _status(_get(ws, row, headers, "Active Status", "Status"))
            item.is_active = item.status == "ACTIVE"
            item.remarks = _s(_get(ws, row, headers, "Remarks"))
            _archive_master_record(db, company_id, "kitchen_sections", code, name, item.section_name_ar, _row_json(ws, row, headers), item.remarks)
            _bump(summary, result)
        except Exception as exc:
            summary["errors"].append(f"Kitchen section row {row}: {exc}")
    return summary


IMPORTERS: dict[str, Callable[[Session, Any, int], dict[str, Any]]] = {
    "inventory": import_inventory,
    "customers": import_customers,
    "suppliers": import_suppliers,
    "chefs": import_chefs,
    "brands": import_brands,
    "revenue_streams": import_revenue_streams,
    "kitchen_locations": import_kitchen_locations,
    "kitchen_sections": import_kitchen_sections,
}


def import_master_data_excel(db: Session, file_path: str, company_id: int = 1, master_type: str = "all") -> dict[str, Any]:
    workbook = load_workbook(file_path, data_only=True)
    results: dict[str, Any] = {}

    if master_type == "recipes":
        recipe_result = import_recipe_excel(db, file_path, company_id=company_id)
        return {"recipes": recipe_result}

    if master_type == "recipes_inventory":
        raw_ws = (
            _find_sheet(workbook, "raw material")
            or _find_sheet(workbook, "material list")
            or _find_sheet(workbook, "inventory")
        )

        if raw_ws is None:
            results["inventory"] = {
                "created": 0,
                "updated": 0,
                "pending": 0,
                "skipped": 0,
                "errors": ["Raw material / inventory sheet not found"],
            }
        else:
            try:
                results["inventory"] = import_inventory(db, raw_ws, company_id)
                db.commit()
            except Exception as exc:
                db.rollback()
                results["inventory"] = {
                    "created": 0,
                    "updated": 0,
                    "pending": 0,
                    "skipped": 0,
                    "errors": [str(exc)],
                }

        try:
            results["recipes"] = import_recipe_excel(db, file_path, company_id=company_id)
        except Exception as exc:
            db.rollback()
            results["recipes"] = {
                "created": 0,
                "updated": 0,
                "pending": 0,
                "skipped": 0,
                "errors": [str(exc)],
            }

        return results

    if master_type != "all":
        if master_type not in IMPORTERS:
            return {master_type: {"created": 0, "updated": 0, "pending": 0, "skipped": 0, "errors": ["Unknown master type"]}}
        ws = _sheet_for_type(workbook, master_type)
        if ws is None:
            return {master_type: {"created": 0, "updated": 0, "pending": 0, "skipped": 0, "errors": [f"Sheet not found for {master_type}"]}}
        try:
            results[master_type] = IMPORTERS[master_type](db, ws, company_id)
            db.commit()
        except Exception as exc:
            db.rollback()
            results[master_type] = {"created": 0, "updated": 0, "pending": 0, "skipped": 0, "errors": [str(exc)]}
        return results

    for result_key, config in MASTER_TYPES.items():
        if result_key in ("all", "recipes", "recipes_inventory"):
            continue
        ws = _find_sheet(workbook, config["sheet_keyword"])
        if ws is None:
            # Not every full workbook contains every optional master sheet.
            # Treat missing optional sheets as skipped, not as upload failure.
            results[result_key] = {"created": 0, "updated": 0, "pending": 0, "skipped": 1, "errors": []}
            continue
        try:
            results[result_key] = IMPORTERS[result_key](db, ws, company_id)
            db.flush()
        except Exception as exc:
            db.rollback()
            results[result_key] = {"created": 0, "updated": 0, "pending": 0, "skipped": 0, "errors": [str(exc)]}
            # Continue to next sheet after rollback; previous successful sheets before rollback are not committed.
            # This is intentional to avoid partial duplicate crashes. User can upload single sheets after resolving errors.
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        results["commit"] = {"created": 0, "updated": 0, "pending": 0, "skipped": 0, "errors": [str(exc)]}
    return results

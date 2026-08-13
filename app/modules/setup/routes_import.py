# app/modules/setup/routes_import.py
# =============================================================================
# Batch 109 — GO-LIVE DATA IMPORT: Chart of Accounts + Opening Stock
# -----------------------------------------------------------------------------
# Two importers that look similar and are fundamentally different, which is
# exactly why they are separated here rather than bolted onto Master Upload.
#
# 1. CHART OF ACCOUNTS — a ONE-TIME structural load.
#    The CoA defines what the ledger can post to. Re-importing it after
#    postings exist is dangerous: change an account's type from Asset to
#    Expense and every historical statement silently reinterprets itself.
#    So this importer REFUSES to replace accounts once journals exist against
#    them, and says which ones.
#
# 2. OPENING STOCK — a TRANSACTIONAL load, not master data.
#    Your "Current Inventory Report" is 1,258 lot-level balances across 884
#    items. That is not an item master (you already have 1,429 ingredients) —
#    it is the opening balance of each lot on a given date.
#
#    Loading it as "master data" would be wrong twice over: it would create
#    duplicate item records, and the quantities would never reach the stock
#    ledger, so every availability check would still read zero.
#
#    The correct ERP treatment, and what this does: match each row to the
#    existing ingredient master, then post an OPENING_STOCK movement to
#    inventory_transactions carrying the barcode as the lot number. Stock then
#    behaves exactly as if it had been received — valuation, shortage checks,
#    reorder and FIFO all work with no special cases.
# =============================================================================
from __future__ import annotations

import os
import tempfile
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, Request, UploadFile
from fastapi.responses import RedirectResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from urllib.parse import quote

from app.core.rbac import require_area, require_action
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/setup/import", tags=["Setup"])


def _cid(request: Request) -> int:
    return int(request.session.get("company_id") or 1)


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _num(v) -> float:
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


# Map the workbook's Class column to the account_type the GL expects. The
# statements (P&L, Balance Sheet) group on this, so an unmapped class would
# quietly drop an account out of every report.
CLASS_TO_TYPE = {
    "ASSETS": "ASSET", "ASSET": "ASSET",
    "LIABILITIES": "LIABILITY", "LIABILITY": "LIABILITY",
    "EQUITY": "EQUITY", "CAPITAL": "EQUITY",
    "REVENUE": "REVENUE", "INCOME": "REVENUE", "SALES": "REVENUE",
    "EXPENSES": "EXPENSE", "EXPENSE": "EXPENSE", "COST OF SALES": "EXPENSE",
}


def ensure_coa_hierarchy(db: Session) -> None:
    """Batch 110 — store the FULL account hierarchy, not just code/name/type.

    Batch 109 imported only account_code, account_name and account_type and
    threw the rest away. That flattened a four-level structure
    (Class -> Group -> Subgroup -> Account) into a flat list, which loses the
    two things the hierarchy is actually for: grouped financial statements,
    and knowing where a new account belongs.

    information_schema pre-check per column — ADD COLUMN IF NOT EXISTS is not
    supported on the target MySQL version.
    """
    cols = {
        "class_code": "VARCHAR(10) NULL",
        "class_name": "VARCHAR(120) NULL",
        "group_code": "VARCHAR(10) NULL",
        "group_name": "VARCHAR(120) NULL",
        "subgroup_code": "VARCHAR(10) NULL",
        "subgroup_name": "VARCHAR(120) NULL",
        "account_seq": "VARCHAR(10) NULL",
        "is_demo": "TINYINT(1) NOT NULL DEFAULT 0",
    }
    for col, ddl in cols.items():
        try:
            has = db.execute(text("""
                SELECT COUNT(*) FROM information_schema.columns
                WHERE table_schema = DATABASE() AND table_name = 'gl_accounts'
                  AND column_name = :c
            """), {"c": col}).scalar()
            if not has:
                db.execute(text(f"ALTER TABLE gl_accounts ADD COLUMN {col} {ddl}"))
                db.commit()
        except Exception:
            db.rollback()


# ---------------------------------------------------------------------------
# Landing page
# ---------------------------------------------------------------------------
@router.get("")
def import_home(request: Request, db: Session = Depends(get_db)):
    require_area(request, "settings")
    cid = _cid(request)
    ensure_coa_hierarchy(db)

    def _count(sql, params=None):
        try:
            return db.execute(text(sql), params or {}).scalar() or 0
        except Exception:
            return 0

    return render(request, "setup/import.html", {
        "stats": {
            "accounts": _count("SELECT COUNT(*) FROM gl_accounts WHERE company_id = :c OR company_id IS NULL", {"c": cid}),
            "journals": _count("SELECT COUNT(*) FROM gl_journal_lines"),
            "ingredients": _count("SELECT COUNT(*) FROM ingredients"),
            "opening_rows": _count("SELECT COUNT(*) FROM inventory_transactions WHERE movement_type = 'OPENING_STOCK'"),
        },
        "page_title": "Go-Live Data Import",
    })


# ---------------------------------------------------------------------------
# Chart of Accounts
# ---------------------------------------------------------------------------
def _parse_coa(path: str) -> tuple[list[dict], list[str]]:
    """Read the CoA workbook. Returns (rows, problems)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next((wb[n] for n in wb.sheetnames if "account" in n.lower()), wb[wb.sheetnames[0]])

    # Locate the header row rather than assuming row 1 — exported workbooks
    # routinely carry a title row above the headers.
    header_row, headers = None, {}
    for r in range(1, min(ws.max_row, 15) + 1):
        vals = {}
        for c in range(1, min(ws.max_column, 30) + 1):
            v = _norm(ws.cell(row=r, column=c).value).lower()
            if v:
                vals[v] = c
        # Batch 110: your export uses "GL Account Code" / "Class Name"; an
        # earlier one used "GL Code" / "cls". Accept both — the detector was
        # only looking for the old names, so the new file parsed to zero rows
        # and reported "no header found" rather than importing.
        has_gl = any(k in vals for k in ("gl account code", "gl code", "gl_code"))
        has_ctx = any(k in vals for k in ("account name", "account", "class name", "class"))
        if has_gl and has_ctx:
            header_row, headers = r, vals
            break
    if header_row is None:
        return [], ["Could not find a header row containing 'GL Account Code' or 'GL Code'. "
                    "Check the sheet still has its export headers."]

    def col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    # Batch 110: the export uses full names ("Class Code", "GL Account Code")
    # while an earlier version used short ones ("cls", "GL Code"). Both are
    # accepted so a file from either export still imports.
    c_code = col("gl account code", "gl code", "gl_code", "code")
    c_name = col("gl account name", "account name", "account")
    c_class = col("class name", "class")
    c_class_c = col("class code", "cls")
    c_group = col("group name", "group")
    c_group_c = col("group code", "grp")
    c_sub = col("subgroup name", "sub group", "sub grp")
    c_sub_c = col("subgroup code", "sub grp code")
    c_acct_c = col("account code", "acc")

    rows, problems, seen = [], [], set()
    for r in range(header_row + 1, ws.max_row + 1):
        code = _norm(ws.cell(row=r, column=c_code).value) if c_code else ""
        if not code:
            continue
        # Excel turns numeric codes into floats: 11001.0 -> "11001"
        if code.endswith(".0"):
            code = code[:-2]
        name = _norm(ws.cell(row=r, column=c_name).value) if c_name else ""
        klass = _norm(ws.cell(row=r, column=c_class).value) if c_class else ""
        group = _norm(ws.cell(row=r, column=c_group).value) if c_group else ""
        sub = _norm(ws.cell(row=r, column=c_sub).value) if c_sub else ""

        acct_type = CLASS_TO_TYPE.get(klass.upper(), "")
        if not acct_type:
            problems.append(f"Row {r}: class '{klass}' is not recognised — account {code} would not "
                            "appear on any statement.")
            continue
        if code in seen:
            problems.append(f"Row {r}: GL code {code} appears more than once — the later row wins.")
        seen.add(code)

        def cell(c):
            return _norm(ws.cell(row=r, column=c).value) if c else ""

        def clean_code(v):
            # Excel turns "01" into 1 and 11001 into 11001.0
            if v.endswith(".0"):
                v = v[:-2]
            return v

        rows.append({
            "row": r, "account_code": code, "account_name": name or code,
            "account_type": acct_type, "klass": klass, "group": group, "sub_group": sub,
            "class_code": clean_code(cell(c_class_c)),
            "group_code": clean_code(cell(c_group_c)),
            "subgroup_code": clean_code(cell(c_sub_c)),
            "account_seq": clean_code(cell(c_acct_c)),
        })
    return rows, problems


@router.post("/coa/preview")
async def coa_preview(request: Request, file: UploadFile = File(...),
                      db: Session = Depends(get_db)):
    require_area(request, "settings")
    cid = _cid(request)
    name = file.filename or ""
    if not name.lower().endswith((".xlsx", ".xlsm")):
        return RedirectResponse(
            f"/setup/import?toast=danger&title={quote('Wrong file type')}"
            f"&msg={quote('Upload the .xlsx export of the Chart of Accounts.')}", status_code=303)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        path = tmp.name
    try:
        rows, problems = _parse_coa(path)
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass

    existing = {}
    try:
        for r in db.execute(text("""
            SELECT account_code, account_name, account_type FROM gl_accounts
            WHERE company_id = :c OR company_id IS NULL
        """), {"c": cid}).mappings().all():
            existing[r["account_code"]] = dict(r)
    except Exception:
        existing = {}

    # Which existing accounts have been posted to? Those cannot be safely
    # removed or retyped, and the user has to know before choosing "replace".
    posted = set()
    try:
        for r in db.execute(text("SELECT DISTINCT account_code FROM gl_journal_lines")).all():
            posted.add(r[0])
    except Exception:
        pass

    incoming = {r["account_code"] for r in rows}
    for r in rows:
        cur = existing.get(r["account_code"])
        r["action"] = "update" if cur else "insert"
        r["retype"] = bool(cur and cur["account_type"] != r["account_type"])
    obsolete = [c for c in existing if c not in incoming]
    blocked = [c for c in obsolete if c in posted]

    by_type: dict[str, int] = {}
    for r in rows:
        by_type[r["account_type"]] = by_type.get(r["account_type"], 0) + 1

    return render(request, "setup/coa_preview.html", {
        "rows": rows[:400], "row_count": len(rows), "problems": problems,
        "by_type": sorted(by_type.items()),
        "inserts": sum(1 for r in rows if r["action"] == "insert"),
        "updates": sum(1 for r in rows if r["action"] == "update"),
        "retypes": [r for r in rows if r.get("retype")],
        "obsolete": obsolete, "blocked": blocked,
        "posted_count": len(posted),
        "filename": name,
        "page_title": "Chart of Accounts — Preview",
    })


@router.post("/coa/commit")
async def coa_commit(request: Request, file: UploadFile = File(...),
                     db: Session = Depends(get_db)):
    """Import for real. `replace=1` deactivates accounts not in the file —
    but never ones that already carry postings."""
    require_action(request, "settings", "edit")
    cid = _cid(request)
    form = await request.form()
    replace = bool(form.get("replace"))

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        path = tmp.name
    try:
        rows, problems = _parse_coa(path)
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass

    if not rows:
        return RedirectResponse(
            f"/setup/import?toast=danger&title={quote('Nothing imported')}"
            f"&msg={quote('No usable rows found. ' + (problems[0] if problems else ''))}",
            status_code=303)

    from app.modules.finance.routes import _ensure_finance_schema
    try:
        _ensure_finance_schema(db)
    except Exception:
        pass
    ensure_coa_hierarchy(db)

    inserted = updated = 0
    for r in rows:
        try:
            res = db.execute(text("""
                INSERT INTO gl_accounts
                    (company_id, account_code, account_name, account_type, is_active,
                     class_code, class_name, group_code, group_name,
                     subgroup_code, subgroup_name, account_seq, is_demo)
                VALUES (:cid, :code, :name, :type, 1,
                        :ccode, :cname, :gcode, :gname, :scode, :sname, :aseq, 0)
                ON DUPLICATE KEY UPDATE
                    account_name = VALUES(account_name),
                    account_type = VALUES(account_type),
                    class_code = VALUES(class_code), class_name = VALUES(class_name),
                    group_code = VALUES(group_code), group_name = VALUES(group_name),
                    subgroup_code = VALUES(subgroup_code), subgroup_name = VALUES(subgroup_name),
                    account_seq = VALUES(account_seq),
                    is_demo = 0,
                    is_active = 1
            """), {"cid": cid, "code": r["account_code"], "name": r["account_name"][:255],
                   "type": r["account_type"],
                   "ccode": r.get("class_code") or None, "cname": r.get("klass") or None,
                   "gcode": r.get("group_code") or None, "gname": r.get("group") or None,
                   "scode": r.get("subgroup_code") or None, "sname": r.get("sub_group") or None,
                   "aseq": r.get("account_seq") or None})
            # MySQL returns 1 for insert, 2 for an update that changed something.
            if getattr(res, "rowcount", 1) == 1:
                inserted += 1
            else:
                updated += 1
        except Exception:
            db.rollback()

    deactivated = 0
    if replace:
        incoming = [r["account_code"] for r in rows]
        ph = ",".join(f":c{i}" for i in range(len(incoming)))
        params = {f"c{i}": v for i, v in enumerate(incoming)}
        params["cid"] = cid
        try:
            # Deactivate, never DELETE, and never touch anything with postings.
            # A deleted account orphans its journal lines and breaks every
            # historical statement that referenced it.
            res = db.execute(text(f"""
                UPDATE gl_accounts
                SET is_active = 0
                WHERE (company_id = :cid OR company_id IS NULL)
                  AND account_code NOT IN ({ph})
                  AND account_code NOT IN (SELECT DISTINCT account_code FROM gl_journal_lines)
            """), params)
            deactivated = getattr(res, "rowcount", 0) or 0
        except Exception:
            db.rollback()

    db.commit()
    msg = f"{inserted} account(s) added, {updated} updated"
    if replace:
        msg += f", {deactivated} unused demo account(s) deactivated"
    if problems:
        msg += f". {len(problems)} row(s) skipped — see the preview for detail"
    return RedirectResponse(
        f"/finance/coa?toast=success&title={quote('Chart of Accounts Imported')}&msg={quote(msg + '.')}",
        status_code=303)


# ---------------------------------------------------------------------------
# Opening stock
# ---------------------------------------------------------------------------
def _parse_stock(path: str) -> tuple[list[dict], int, list[str]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row, headers = None, {}
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = {}
        for c in range(1, min(ws.max_column, 30) + 1):
            v = _norm(ws.cell(row=r, column=c).value).lower()
            if v:
                vals[v] = c
        if any(k in vals for k in ("material code", "item code", "ingredient code")):
            header_row, headers = r, vals
            break
    if header_row is None:
        return [], 0, ["Could not find a header row containing 'Material Code'."]

    def col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    c_code = col("material code", "item code", "ingredient code")
    c_name = col("material name", "item name", "description")
    c_qty = col("total qty", "qty", "quantity")
    c_val = col("total price", "total value", "value")
    c_unit = col("cost unit", "unit cost", "rate")
    c_uom = col("issue unit name", "uom", "unit")
    c_lot = col("barcode", "lot", "batch")
    c_date = col("received date", "date")

    rows, problems = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        code = _norm(ws.cell(row=r, column=c_code).value) if c_code else ""
        if not code:
            continue
        qty = _num(ws.cell(row=r, column=c_qty).value) if c_qty else 0.0
        if qty <= 0:
            continue
        val = _num(ws.cell(row=r, column=c_val).value) if c_val else 0.0
        unit = _num(ws.cell(row=r, column=c_unit).value) if c_unit else 0.0
        if unit <= 0 and qty:
            unit = val / qty
        lot = _norm(ws.cell(row=r, column=c_lot).value) if c_lot else ""
        # Barcodes are 16-digit numbers, and Excel stores anything past 15
        # significant digits as a float — so the cell reads "1.61020291119274e+16"
        # instead of "16102029111927400". Left as-is the lot number is unusable
        # for traceability, which is the entire point of capturing it.
        if lot:
            if "e+" in lot.lower():
                try:
                    lot = f"{int(float(lot)):d}"
                except (ValueError, OverflowError):
                    pass
            elif lot.endswith(".0"):
                lot = lot[:-2]
        raw_date = ws.cell(row=r, column=c_date).value if c_date else None
        rdate = raw_date.date() if isinstance(raw_date, datetime) else None

        rows.append({
            "row": r, "code": code,
            "name": _norm(ws.cell(row=r, column=c_name).value) if c_name else "",
            "uom": _norm(ws.cell(row=r, column=c_uom).value) if c_uom else "",
            "qty": qty, "unit_cost": unit, "value": val or qty * unit,
            "lot": lot, "received": rdate,
        })
    return rows, header_row, problems


@router.post("/stock/preview")
async def stock_preview(request: Request, file: UploadFile = File(...),
                        db: Session = Depends(get_db)):
    require_area(request, "settings")
    name = file.filename or ""
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        path = tmp.name
    try:
        rows, header_row, problems = _parse_stock(path)
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass

    known = set()
    try:
        for r in db.execute(text("SELECT ingredient_code FROM ingredients")).all():
            known.add(r[0])
    except Exception:
        pass

    matched = [r for r in rows if r["code"] in known]
    unmatched = [r for r in rows if r["code"] not in known]
    unmatched_codes = sorted({r["code"] for r in unmatched})

    already = 0
    try:
        already = db.execute(text(
            "SELECT COUNT(*) FROM inventory_transactions WHERE movement_type = 'OPENING_STOCK'"
        )).scalar() or 0
    except Exception:
        pass

    return render(request, "setup/stock_preview.html", {
        "rows": matched[:300], "unmatched": unmatched[:200],
        "unmatched_codes": unmatched_codes[:200],
        "totals": {
            "lines": len(rows),
            "matched": len(matched),
            "unmatched": len(unmatched),
            "distinct_items": len({r["code"] for r in matched}),
            "value": round(sum(r["value"] for r in matched), 2),
            "already_loaded": already,
        },
        "problems": problems, "filename": name,
        "page_title": "Opening Stock — Preview",
    })


@router.post("/stock/commit")
async def stock_commit(request: Request, file: UploadFile = File(...),
                       db: Session = Depends(get_db)):
    """Post the opening balances to the stock ledger.

    Each row becomes one OPENING_STOCK movement carrying the barcode as the
    lot number, so it behaves exactly like received stock: valuation, shortage
    checks, reorder cover and lot traceability all work with no special cases.

    qc_status is 'Passed' deliberately — opening stock is already in the store
    and has been accepted; sending it to QC Hold would make every item
    unavailable on day one.
    """
    require_action(request, "settings", "edit")
    cid = _cid(request)
    form = await request.form()
    as_of = _norm(form.get("as_of")) or date.today().isoformat()
    skip_existing = bool(form.get("skip_existing"))

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        path = tmp.name
    try:
        rows, _hr, problems = _parse_stock(path)
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass

    if not rows:
        return RedirectResponse(
            f"/setup/import?toast=danger&title={quote('Nothing imported')}"
            f"&msg={quote('No usable stock rows found in that file.')}", status_code=303)

    if skip_existing:
        try:
            n = db.execute(text(
                "SELECT COUNT(*) FROM inventory_transactions WHERE movement_type = 'OPENING_STOCK'"
            )).scalar() or 0
            if n:
                return RedirectResponse(
                    f"/setup/import?toast=warning&title={quote('Already loaded')}"
                    f"&msg={quote(f'{n} opening stock movements already exist. Untick the guard to load again, but expect double-counted stock.')}",
                    status_code=303)
        except Exception:
            pass

    known = set()
    try:
        for r in db.execute(text("SELECT ingredient_code FROM ingredients")).all():
            known.add(r[0])
    except Exception:
        pass

    from app.core.stock_ledger import post_stock_movement

    posted = skipped = 0
    for r in rows:
        if r["code"] not in known:
            skipped += 1
            continue
        ok = post_stock_movement(
            db, company_id=cid, inventory_code=r["code"],
            item_name=r["name"] or r["code"], uom=r["uom"] or "",
            qty=r["qty"], movement_type="OPENING_STOCK",
            reference_no=f"OPENING-{as_of}", unit_cost=r["unit_cost"],
            remarks=f"Opening balance as of {as_of}"
                    + (f" · received {r['received']}" if r["received"] else ""),
            created_by=request.session.get("username", "system"),
            lot_no=r["lot"] or "", to_location="Main Store", qc_status="Passed",
        )
        if ok:
            posted += 1
        else:
            skipped += 1
    db.commit()

    return RedirectResponse(
        f"/inventory?toast=success&title={quote('Opening Stock Loaded')}"
        f"&msg={quote(f'{posted} lot balance(s) posted to the ledger. {skipped} row(s) skipped (no matching ingredient).')}",
        status_code=303)

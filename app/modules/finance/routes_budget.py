# app/modules/finance/routes_budget.py
# =============================================================================
# Batch 114 — BUDGET CONTROL
# -----------------------------------------------------------------------------
# WHY THIS MATTERS MORE THAN A SPREADSHEET
#
# A budget kept in Excel tells you what you overspent, last month, after the
# money is gone. A budget inside the ERP can tell you BEFORE you commit — at
# the moment a requisition is raised, while it is still free to say no.
#
# So this is not a reporting feature. It plugs into the approval chain built in
# Batch 111: a requisition that would breach its budget is flagged at approval
# time, with the number, and the approver decides knowingly.
#
# THE THREE NUMBERS, AND WHY "COMMITTED" IS THE ONE THAT MATTERS
#
#   ACTUAL     — posted to the GL. Money already spent.
#   COMMITTED  — approved requisitions and open purchase orders not yet
#                invoiced. Money promised but not yet posted.
#   AVAILABLE  — budget − actual − committed.
#
# Almost every naive budget system tracks only ACTUAL, and it is always wrong
# in the same direction: you appear to have budget left that is in fact already
# spoken for by open POs. By the time those invoice, you are over. Tracking the
# commitment is the entire point of doing this in an ERP rather than a
# spreadsheet.
#
# BUDGETING AT ANY LEVEL OF THE HIERARCHY
#
# Batch 110 loaded 203 accounts classified Class -> Group -> Subgroup ->
# Account. A budget can be set at ANY of those levels:
#
#   * "Expenses" (class)                     — one line, whole-business ceiling
#   * "Food Cost" (group)                    — the usual working level
#   * "Chilled Purchases" (subgroup)         — tighter control where it matters
#   * "11209 Cards Under Process" (account)  — a single line
#
# Spend rolls UP: a posting to account 51004 consumes the budget of its
# subgroup, its group and its class simultaneously. That mirrors how a finance
# team actually thinks — a department head owns a group total and does not want
# to maintain fifty account lines.
# =============================================================================
from __future__ import annotations

import io
from datetime import date

from fastapi import APIRouter, Depends, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from urllib.parse import quote

from app.core.rbac import require_area, require_action
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/finance/budget", tags=["Finance"])

LEVELS = [("class", "Class"), ("group", "Group"),
          ("subgroup", "Subgroup"), ("account", "Account")]

PERIODS = [("YEAR", "Full year"), ("H1", "First half"), ("H2", "Second half"),
           ("Q1", "Quarter 1"), ("Q2", "Quarter 2"),
           ("Q3", "Quarter 3"), ("Q4", "Quarter 4")]

# Month ranges per period, used to bound the actuals query.
PERIOD_MONTHS = {
    "YEAR": (1, 12), "H1": (1, 6), "H2": (7, 12),
    "Q1": (1, 3), "Q2": (4, 6), "Q3": (7, 9), "Q4": (10, 12),
}


def _cid(request: Request) -> int:
    return int(request.session.get("company_id") or 1)


def ensure_schema(db: Session) -> None:
    try:
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS budgets (
                id INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                company_id INT NULL,
                fiscal_year INT NOT NULL,
                period VARCHAR(10) NOT NULL DEFAULT 'YEAR',
                level VARCHAR(20) NOT NULL,
                level_code VARCHAR(40) NOT NULL,
                level_name VARCHAR(160) NULL,
                amount DECIMAL(18,4) NOT NULL DEFAULT 0,
                warn_pct DECIMAL(6,2) NOT NULL DEFAULT 85,
                block_over TINYINT(1) NOT NULL DEFAULT 0,
                notes VARCHAR(255) NULL,
                created_by VARCHAR(120) NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_budget (company_id, fiscal_year, period, level, level_code),
                KEY idx_budget_year (fiscal_year)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """))
        db.commit()
    except Exception:
        db.rollback()


def _accounts_for(db: Session, level: str, code: str, cid: int) -> list[str]:
    """Every GL account that rolls up into this budget line.

    This is what makes hierarchy budgeting work: a budget on a group is
    consumed by postings to any account beneath it.
    """
    col = {"class": "class_code", "group": "group_code",
           "subgroup": "subgroup_code", "account": "account_code"}.get(level)
    if not col:
        return []
    if level == "account":
        return [code]
    # Group and subgroup codes repeat across classes (every class has a
    # group "1"), so the parent codes must be part of the match or a budget
    # on Assets>Current would also capture Expenses>Current.
    parts = code.split("|")
    where = ["(company_id = :cid OR company_id IS NULL)"]
    params: dict = {"cid": cid}
    keys = ["class_code", "group_code", "subgroup_code"]
    for i, p in enumerate(parts):
        if i < len(keys) and p != "":
            where.append(f"COALESCE({keys[i]},'') = :p{i}")
            params[f"p{i}"] = p
    try:
        return [r[0] for r in db.execute(text(f"""
            SELECT account_code FROM gl_accounts WHERE {' AND '.join(where)}
        """), params).all()]
    except Exception:
        return []


def _actual(db: Session, accounts: list[str], year: int, period: str, cid: int) -> float:
    """Posted spend. Expense accounts are debit-positive, so actual = debit − credit."""
    if not accounts:
        return 0.0
    m_from, m_to = PERIOD_MONTHS.get(period, (1, 12))
    ph = ",".join(f":a{i}" for i in range(len(accounts)))
    params = {f"a{i}": v for i, v in enumerate(accounts)}
    params.update({"y": year, "mf": m_from, "mt": m_to, "cid": cid})
    try:
        return float(db.execute(text(f"""
            SELECT COALESCE(SUM(COALESCE(l.debit,0) - COALESCE(l.credit,0)), 0)
            FROM gl_journal_lines l
            JOIN gl_journals j ON j.journal_no = l.journal_no
            WHERE l.account_code IN ({ph})
              AND YEAR(j.journal_date) = :y
              AND MONTH(j.journal_date) BETWEEN :mf AND :mt
              AND (j.company_id = :cid OR j.company_id IS NULL)
        """), params).scalar() or 0)
    except Exception:
        return 0.0


def _committed(db: Session, level: str, code: str, year: int, cid: int) -> float:
    """Approved requisitions + open POs not yet invoiced.

    Requisitions and POs carry no GL account, so this cannot be split by
    account — it is reported at the level the user is budgeting and flagged in
    the UI as an approximation. Showing an approximate commitment is far more
    useful than showing none: the alternative is a budget that looks healthy
    while three open POs are about to land on it.
    """
    try:
        pr = float(db.execute(text("""
            SELECT COALESCE(SUM(estimated_value), 0) FROM purchase_requisitions
            WHERE status = 'Approved' AND YEAR(COALESCE(pr_date, created_at)) = :y
              AND (company_id = :cid OR company_id IS NULL)
        """), {"y": year, "cid": cid}).scalar() or 0)
        po = float(db.execute(text("""
            SELECT COALESCE(SUM(GREATEST(COALESCE(pol.ordered_qty,0) - COALESCE(g.recv,0), 0)
                                * COALESCE(pol.unit_price,0)), 0)
            FROM purchase_order_lines pol
            JOIN purchase_orders po ON po.po_no = pol.po_no
            LEFT JOIN (SELECT po_no, inventory_code, SUM(received_qty) AS recv
                       FROM grn_lines GROUP BY po_no, inventory_code) g
                   ON g.po_no = pol.po_no AND g.inventory_code = pol.inventory_code
            WHERE COALESCE(po.status,'') NOT IN ('Cancelled','Closed')
              AND YEAR(po.po_date) = :y
              AND (po.company_id = :cid OR po.company_id IS NULL)
        """), {"y": year, "cid": cid}).scalar() or 0)
        return round(pr + po, 2)
    except Exception:
        return 0.0


def budget_rows(db: Session, cid: int, year: int, period: str = "") -> list[dict]:
    ensure_schema(db)
    where = ["(company_id = :cid OR company_id IS NULL)", "fiscal_year = :y"]
    params: dict = {"cid": cid, "y": year}
    if period:
        where.append("period = :p")
        params["p"] = period
    try:
        rows = [dict(r) for r in db.execute(text(f"""
            SELECT * FROM budgets WHERE {' AND '.join(where)}
            ORDER BY level, level_code
        """), params).mappings().all()]
    except Exception:
        return []

    total_committed = _committed(db, "", "", year, cid)
    for r in rows:
        accounts = _accounts_for(db, r["level"], r["level_code"], cid)
        r["account_count"] = len(accounts)
        r["actual"] = round(_actual(db, accounts, year, r["period"], cid), 2)
        budget = float(r["amount"] or 0)
        # Commitment is apportioned by budget share — it cannot be attributed
        # to a GL account, so this is an estimate and the UI says so.
        total_budget = sum(float(x["amount"] or 0) for x in rows) or 1
        r["committed"] = round(total_committed * (budget / total_budget), 2)
        r["available"] = round(budget - r["actual"] - r["committed"], 2)
        r["used_pct"] = round(((r["actual"] + r["committed"]) / budget * 100) if budget else 0, 1)
        warn = float(r["warn_pct"] or 85)
        r["state"] = ("over" if r["used_pct"] > 100
                      else ("warn" if r["used_pct"] >= warn else "ok"))
    return rows


def check_budget(db: Session, cid: int, amount: float, year: int | None = None) -> dict:
    """Would this spend breach a budget? Called from the approval screen.

    Returns the tightest binding line, so an approver sees the constraint that
    actually bites rather than a list of everything.
    """
    year = year or date.today().year
    rows = budget_rows(db, cid, year)
    if not rows:
        return {"has_budget": False}

    worst = None
    for r in rows:
        after = r["available"] - amount
        pct_after = round(((r["actual"] + r["committed"] + amount) / float(r["amount"] or 1) * 100), 1)
        cand = {**r, "after": round(after, 2), "pct_after": pct_after,
                "breaches": after < 0, "blocks": bool(r["block_over"]) and after < 0}
        if worst is None or cand["after"] < worst["after"]:
            worst = cand
    return {"has_budget": True, "line": worst}


@router.get("")
def budget_screen(request: Request, db: Session = Depends(get_db)):
    require_area(request, "finance")
    ensure_schema(db)
    cid = _cid(request)
    q = request.query_params
    try:
        year = int(q.get("year") or date.today().year)
    except ValueError:
        year = date.today().year
    period = (q.get("period") or "").strip()

    rows = budget_rows(db, cid, year, period)

    # Hierarchy options, so a budget line is always attached to something real.
    options = []
    try:
        for r in db.execute(text("""
            SELECT DISTINCT COALESCE(class_code,'') c, COALESCE(class_name,'') cn,
                   COALESCE(group_code,'') g, COALESCE(group_name,'') gn,
                   COALESCE(subgroup_code,'') s, COALESCE(subgroup_name,'') sn
            FROM gl_accounts
            WHERE (company_id = :cid OR company_id IS NULL)
              AND COALESCE(class_code,'') <> ''
            ORDER BY c, g, s
        """), {"cid": cid}).mappings().all():
            options.append(dict(r))
    except Exception:
        pass

    classes, groups, subgroups = {}, {}, {}
    for o in options:
        classes[o["c"]] = o["cn"]
        groups[f"{o['c']}|{o['g']}"] = f"{o['cn']} › {o['gn']}"
        subgroups[f"{o['c']}|{o['g']}|{o['s']}"] = f"{o['cn']} › {o['gn']} › {o['sn']}"

    totals = {
        "budget": round(sum(float(r["amount"] or 0) for r in rows), 2),
        "actual": round(sum(r["actual"] for r in rows), 2),
        "committed": round(sum(r["committed"] for r in rows), 2),
        "over": sum(1 for r in rows if r["state"] == "over"),
        "warn": sum(1 for r in rows if r["state"] == "warn"),
    }
    totals["available"] = round(totals["budget"] - totals["actual"] - totals["committed"], 2)

    return render(request, "finance/budget.html", {
        "rows": rows, "totals": totals, "year": year, "period": period,
        "levels": LEVELS, "periods": PERIODS,
        "classes": classes, "groups": groups, "subgroups": subgroups,
        "years": list(range(date.today().year - 2, date.today().year + 3)),
        "page_title": "Budget Control",
    })


@router.post("/save")
async def save_budget(request: Request, db: Session = Depends(get_db)):
    require_action(request, "finance", "edit")
    ensure_schema(db)
    form = await request.form()
    cid = _cid(request)
    try:
        year = int(form.get("fiscal_year") or date.today().year)
        amount = float(form.get("amount") or 0)
    except ValueError:
        return RedirectResponse(
            f"/finance/budget?toast=warning&title={quote('Invalid figures')}"
            f"&msg={quote('Year and amount must be numbers.')}", status_code=303)

    level = (form.get("level") or "group").strip()
    code = (form.get("level_code") or "").strip()
    name = (form.get("level_name") or "").strip()
    if not code or amount <= 0:
        return RedirectResponse(
            f"/finance/budget?toast=warning&title={quote('Nothing saved')}"
            f"&msg={quote('Pick a budget line and enter an amount above zero.')}", status_code=303)

    db.execute(text("""
        INSERT INTO budgets (company_id, fiscal_year, period, level, level_code, level_name,
                             amount, warn_pct, block_over, notes, created_by)
        VALUES (:cid, :y, :p, :l, :c, :n, :a, :w, :b, :nt, :by)
        ON DUPLICATE KEY UPDATE
            amount = VALUES(amount), warn_pct = VALUES(warn_pct),
            block_over = VALUES(block_over), notes = VALUES(notes),
            level_name = VALUES(level_name)
    """), {"cid": cid, "y": year, "p": (form.get("period") or "YEAR"),
           "l": level, "c": code, "n": name[:160] or None, "a": amount,
           "w": float(form.get("warn_pct") or 85),
           "b": 1 if form.get("block_over") else 0,
           "nt": (form.get("notes") or "")[:255] or None,
           "by": request.session.get("username", "system")})
    db.commit()
    return RedirectResponse(
        f"/finance/budget?year={year}&toast=success&title={quote('Budget Saved')}"
        f"&msg={quote(name or code)}", status_code=303)


@router.post("/{budget_id}/delete")
async def delete_budget(request: Request, budget_id: int, db: Session = Depends(get_db)):
    require_action(request, "finance", "delete")
    db.execute(text("DELETE FROM budgets WHERE id = :i AND (company_id = :cid OR company_id IS NULL)"),
               {"i": budget_id, "cid": _cid(request)})
    db.commit()
    return RedirectResponse(
        f"/finance/budget?toast=success&title={quote('Removed')}&msg={quote('Budget line deleted.')}",
        status_code=303)


@router.get("/export")
def export_budget(request: Request, db: Session = Depends(get_db)):
    require_area(request, "finance")
    cid = _cid(request)
    try:
        year = int(request.query_params.get("year") or date.today().year)
    except ValueError:
        year = date.today().year
    rows = budget_rows(db, cid, year)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {year}"
    head = PatternFill("solid", fgColor="132947")
    cols = ["Level", "Budget Line", "Period", "Budget", "Actual",
            "Committed", "Available", "Used %", "State"]
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head
        ws.column_dimensions[c.column_letter].width = max(14, len(h) + 8)
    for r, b in enumerate(rows, start=2):
        for i, v in enumerate([b["level"], b["level_name"] or b["level_code"], b["period"],
                               float(b["amount"] or 0), b["actual"], b["committed"],
                               b["available"], b["used_pct"], b["state"]], start=1):
            ws.cell(row=r, column=i, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ISFC_Budget_{year}.xlsx"'},
    )

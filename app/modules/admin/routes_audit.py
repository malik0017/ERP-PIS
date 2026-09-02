# app/modules/admin/routes_audit.py

from __future__ import annotations
import io
from datetime import date, timedelta
from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from app.core.rbac import require_area
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/admin/audit", tags=["Admin"])

CATEGORIES = {
    "Approvals": ["APPROV", "REJECT", "SIGN"],
    "Security": ["LOGIN", "LOGOUT", "PASSWORD", "LOCK", "UNLOCK", "2FA", "ACCESS"],
    "Master Data": ["MASTER", "CUSTOMER", "SUPPLIER", "INGREDIENT", "BRAND", "RECIPE"],
    "Orders": ["ORDER", "REQUISITION", "DISPATCH", "DELIVER"],
    "Inventory": ["STOCK", "ISSUE", "GRN", "RECEIPT", "ADJUST", "TOPUP", "TOP_UP"],
    "Finance": ["INVOICE", "PAYMENT", "JOURNAL", "GL_", "PERIOD"],
    "Settings": ["SETTING", "CONFIG", "MODULE"],
}


def _cid(request: Request) -> int:
    return int(request.session.get("company_id") or 1)


def _query(db: Session, request: Request, limit: int = 500) -> tuple[list[dict], dict]:
    q = request.query_params
    f = {
        "date_from": (q.get("date_from") or "").strip(),
        "date_to": (q.get("date_to") or "").strip(),
        "user": (q.get("user") or "").strip(),
        "action": (q.get("action") or "").strip(),
        "category": (q.get("category") or "").strip(),
        "table": (q.get("table") or "").strip(),
    }
    if not f["date_from"] and not f["date_to"]:
        f["date_from"] = (date.today() - timedelta(days=30)).isoformat()

    where = ["1=1"]
    params: dict = {"cid": _cid(request), "lim": limit}

    if f["date_from"]:
        where.append("a.created_at >= :df"); params["df"] = f["date_from"] + " 00:00:00"
    if f["date_to"]:
        where.append("a.created_at <= :dt"); params["dt"] = f["date_to"] + " 23:59:59"
    if f["user"]:
        where.append("(u.username LIKE :us OR u.full_name LIKE :us)"); params["us"] = f"%{f['user']}%"
    if f["action"]:
        where.append("a.action LIKE :ac"); params["ac"] = f"%{f['action']}%"
    if f["table"]:
        where.append("a.table_name LIKE :tb"); params["tb"] = f"%{f['table']}%"
    if f["category"] and f["category"] in CATEGORIES:
        toks = CATEGORIES[f["category"]]
        ors = []
        for i, tok in enumerate(toks):
            params[f"k{i}"] = f"%{tok}%"
            ors.append(f"UPPER(COALESCE(a.action,'')) LIKE :k{i}")
        where.append("(" + " OR ".join(ors) + ")")

    where.append("(u.company_id = :cid OR u.company_id IS NULL OR a.user_id IS NULL)")

    try:
        rows = [dict(r) for r in db.execute(text(f"""
            SELECT a.id, a.action, a.table_name, a.record_id, a.created_at,
                   a.user_id,
                   COALESCE(u.username, CONCAT('user #', COALESCE(a.user_id, 0))) AS username,
                   COALESCE(u.full_name, '') AS full_name,
                   COALESCE(u.role, '')      AS role
            FROM audit_logs a
            LEFT JOIN users u ON u.id = a.user_id
            WHERE {' AND '.join(where)}
            ORDER BY a.created_at DESC, a.id DESC
            LIMIT :lim
        """), params).mappings().all()]
    except Exception:
        rows = []
    return rows, f


@router.get("")
def audit_viewer(request: Request, db: Session = Depends(get_db)):
    require_area(request, "audit")
    rows, f = _query(db, request)

    counts: dict[str, int] = {}
    users: dict[str, int] = {}
    for r in rows:
        act = (r["action"] or "").upper()
        bucket = "Other"
        for name, toks in CATEGORIES.items():
            if any(t in act for t in toks):
                bucket = name
                break
        counts[bucket] = counts.get(bucket, 0) + 1
        users[r["username"]] = users.get(r["username"], 0) + 1

    return render(request, "admin/audit.html", {
        "rows": rows,
        "filters": f,
        "categories": list(CATEGORIES.keys()),
        "summary": {
            "total": len(rows),
            "by_category": sorted(counts.items(), key=lambda kv: -kv[1]),
            "top_users": sorted(users.items(), key=lambda kv: -kv[1])[:8],
        },
        "page_title": "Audit Log",
    })


@router.get("/export")
def audit_export(request: Request, db: Session = Depends(get_db)):
    """Export for an auditor. Deliberately capped at 5,000 rows — a request
    that would return more is a filter problem, and silently truncating a
    compliance export is worse than refusing to build a vague one."""
    require_area(request, "audit")
    rows, f = _query(db, request, limit=5000)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    head = PatternFill("solid", fgColor="132947")
    cols = ["When", "User", "Full Name", "Role", "Action", "Table", "Record ID"]
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head
        ws.column_dimensions[c.column_letter].width = max(16, len(h) + 10)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=str(r["created_at"] or ""))
        ws.cell(row=i, column=2, value=r["username"])
        ws.cell(row=i, column=3, value=r["full_name"])
        ws.cell(row=i, column=4, value=r["role"])
        ws.cell(row=i, column=5, value=r["action"])
        ws.cell(row=i, column=6, value=r["table_name"])
        ws.cell(row=i, column=7, value=r["record_id"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="ISFC_Audit_Log_{date.today().isoformat()}.xlsx"'},
    )

# app/modules/recipes/routes_excel.py
# =============================================================================
# Batch 101 — RECIPE EXCEL: template download + validate-before-import.
# -----------------------------------------------------------------------------
# The problem this solves, in the order it bites:
#
#   1. There was no template. Users guessed the column names, and a guessed
#      header simply doesn't match — so the value is silently dropped rather
#      than rejected. That is exactly how 55 Frsh recipes ended up with no
#      category: the data was in the file and nothing ever said it wasn't
#      being read.
#
#   2. There was no dry run. Upload went straight to the database. A bad row
#      was discovered afterwards, in the data, by a human noticing.
#
# So: /recipes/template downloads a correctly-shaped workbook, and
# /recipes/validate-excel runs the SAME parser the importer uses and reports
# what would happen — without writing anything.
#
# The validator deliberately reuses recipe_service's own _headers/_get rather
# than reimplementing the parse. A validator that parses differently from the
# importer is worse than none: it certifies files the importer then mangles.
# =============================================================================
from __future__ import annotations

import os
import tempfile
from collections import Counter

from fastapi import APIRouter, Depends, File, Request, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.rbac import require_action, require_area
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/recipes", tags=["Recipes"])

# The columns the importer actually reads, in the order a human would fill
# them. Keep this in sync with recipe_service._get() calls — that is the
# contract this template exists to communicate.
MASTER_COLUMNS = [
    ("Main Recipe ID", "RCP-FRSH-00001", "Unique recipe code. Reused = update, new = insert."),
    ("Recipe Name", "Avocado Egg Ranch", "Required."),
    ("Brand Code", "FRSH", "Must exist in Master Data > Brands."),
    ("Customer Name", "Frsh", "Must exist in Master Data > Customers."),
    ("Category", "Salad", "Free text. Shown in the Recipes list."),
    ("Day", "Sunday", "Weekly-menu day. Leave blank for non-menu recipes."),
    ("Standard No. of Portions", 10, "Must be greater than 0."),
    ("Weight per Portion (g)", 250, ""),
    ("Size of Portion", 1, ""),
    ("Has Sub Recipe?", "No", "Yes / No."),
    ("Sub Recipe ID", "", "Only when Has Sub Recipe = Yes."),
    ("Sub Recipe Name", "", ""),
    ("Sub Recipe No. of Portions", "", ""),
    ("Recipe Standard Yield %", 0.95, "Decimal, e.g. 0.95 for 95%."),
    ("Targeted Wastage %", 0.05, "Decimal, e.g. 0.05 for 5%."),
    ("Packaging Cost", 0, ""),
    ("Labor Cost", 0, ""),
    ("Delivery Cost", 0, ""),
    ("Overheads", 0, ""),
    ("Other Costs", 0, ""),
    ("Margin %age", 0.30, "Decimal, e.g. 0.30 for 30%."),
    ("Remark", "", ""),
]

LINE_COLUMNS = [
    ("Recipe Ref", "RCP-FRSH-00001", "Must match a Main Recipe ID above."),
    ("Recipe Name", "Avocado Egg Ranch", ""),
    ("Customer Name", "Frsh", ""),
    ("Category", "Salad", ""),
    ("Day", "Sunday", ""),
    ("Item Code", "DRY2-1160", "Must exist in Master Data > Ingredients."),
    ("Item / Ingredient", "Yogurt Plain big 10 KG", ""),
    ("Qty Req per Batch (g/pcs)", 325.5, ""),
    ("Qty Req per Portion", 32.55, ""),
    ("UOM", "Gram", ""),
    ("St. UOM", "Gram", ""),
    ("No. of portions per batch", 10, ""),
    ("Sub Recipe Description", "", ""),
]


def _fill(ws, columns, title):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Row 3 is the header row — do not move or rename it. Data starts on row 4."
    ws["A2"].font = Font(italic=True, size=9)

    head = PatternFill("solid", fgColor="132947")
    for i, (name, _sample, _note) in enumerate(columns, start=1):
        c = ws.cell(row=3, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[c.column_letter].width = max(14, min(30, len(name) + 6))

    for i, (_name, sample, _note) in enumerate(columns, start=1):
        ws.cell(row=4, column=i, value=sample)

    # Notes go on their own sheet rather than as cell comments — comments are
    # invisible until hovered and get stripped by most "Save As CSV" round trips.
    return ws


@router.get("/template")
def download_template(request: Request):
    """Download a correctly-shaped recipe workbook."""
    require_area(request, "recipe_list")
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    _fill(wb.active, MASTER_COLUMNS, "Master Recipes")
    wb.active.title = "Master Recipes"
    _fill(wb.create_sheet("Recipe Ingredients"), LINE_COLUMNS, "Recipe Ingredients")

    notes = wb.create_sheet("How to use")
    notes["A1"] = "ISFC PIMS — Recipe import"
    notes["A1"].font = Font(bold=True, size=14)
    guide = [
        "",
        "1. Fill 'Master Recipes' — one row per recipe.",
        "2. Fill 'Recipe Ingredients' — one row per ingredient, per recipe.",
        "3. Keep the header on row 3. Data starts on row 4.",
        "4. Upload via Recipes > Upload Excel. Validation runs FIRST and shows",
        "   you what will happen before anything is written.",
        "",
        "Column reference — Master Recipes",
    ]
    for line in guide:
        notes.append([line])
    for name, _s, note in MASTER_COLUMNS:
        notes.append([name, note])
    notes.append([])
    notes.append(["Column reference — Recipe Ingredients"])
    for name, _s, note in LINE_COLUMNS:
        notes.append([name, note])
    notes.column_dimensions["A"].width = 34
    notes.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ISFC_Recipe_Template.xlsx"'},
    )


@router.post("/validate-excel")
async def validate_excel(request: Request, file: UploadFile = File(...),
                         db: Session = Depends(get_db)):
    """Dry run. Parses the file with the importer's own helpers and reports
    what WOULD happen. Writes nothing."""
    require_action(request, "recipe_list", "add")

    import openpyxl
    from sqlalchemy import text

    from app.services.recipe_service import _headers, _header_row, _get, _norm, _s

    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        return render(request, "recipes/validate.html", {
            "filename": filename, "fatal": "Not an Excel file (.xlsx or .xlsm required).",
            "page_title": "Validate Recipe File"})

    with tempfile.NamedTemporaryFile(delete=False,
                                     suffix=os.path.splitext(filename)[1]) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    report = {"filename": filename, "sheets": [], "rows": [], "fatal": None,
              "counts": {"ok": 0, "warning": 0, "error": 0},
              "will_insert": 0, "will_update": 0}

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        report["sheets"] = wb.sheetnames

        master = next((wb[n] for n in wb.sheetnames if "master" in n.lower()), None)
        lines = next((wb[n] for n in wb.sheetnames
                      if "ingredient" in n.lower() or "line" in n.lower()), None)
        sheet = master or lines
        if sheet is None:
            report["fatal"] = ("No 'Master Recipes' or 'Recipe Ingredients' sheet found. "
                               "Download the template and use its sheet names.")
            raise RuntimeError("no sheet")

        report["using_sheet"] = sheet.title
        report["fallback_mode"] = master is None
        headers = _headers(sheet)
        hr = _header_row(sheet, headers)
        report["header_row"] = hr
        report["headers_found"] = sorted(headers.keys())

        # Which columns the importer expects but cannot see. This is the check
        # that would have caught the missing-category problem on day one.
        expected = ["recipe name", "category", "customer name"]
        report["missing_headers"] = [h for h in expected if h not in headers]

        # Existing codes -> insert vs update, without writing anything.
        try:
            existing = {r[0] for r in db.execute(text(
                "SELECT recipe_code FROM recipes")).all()}
        except Exception:
            existing = set()

        seen = Counter()
        code_key = "main recipe id" if "main recipe id" in headers else "recipe ref"

        for row in range(hr + 1, min(sheet.max_row, hr + 2001) + 1):
            code = _s(_get(sheet, row, headers, "Main Recipe ID", "Recipe Ref"))
            name = _s(_get(sheet, row, headers, "Recipe Name"))
            if not code and not name:
                continue

            issues = []
            level = "ok"

            if not code:
                issues.append("No recipe code — row will be skipped entirely.")
                level = "error"
            else:
                seen[code] += 1
                if seen[code] > 1:
                    issues.append(f"Duplicate code in this file (occurrence {seen[code]}). "
                                  "Later rows overwrite earlier ones.")
                    level = "warning"

            if not name:
                issues.append("No recipe name — the code will be used as the name.")
                level = "warning" if level == "ok" else level

            if not _s(_get(sheet, row, headers, "Category")):
                issues.append("No category — it will show as '—' in the Recipes list.")
                level = "warning" if level == "ok" else level

            portions = _get(sheet, row, headers, "Standard No. of Portions",
                            "No. of portions per batch", "Portions")
            try:
                if portions is not None and float(portions) <= 0:
                    issues.append("Portions is zero or negative — costing per portion "
                                  "cannot be calculated.")
                    level = "error"
            except (TypeError, ValueError):
                if portions not in (None, ""):
                    issues.append(f"Portions is not a number: {portions!r}")
                    level = "error"

            if code:
                if code in existing:
                    report["will_update"] += 1
                else:
                    report["will_insert"] += 1

            report["counts"][level] += 1
            if issues or level != "ok":
                report["rows"].append({
                    "row": row, "code": code or "—", "name": name or "—",
                    "level": level, "issues": issues,
                })

        report["rows"] = report["rows"][:200]

    except Exception as exc:
        if not report["fatal"]:
            report["fatal"] = f"Could not read the file: {type(exc).__name__}: {exc}"
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    return render(request, "recipes/validate.html",
                  {**report, "page_title": "Validate Recipe File"})

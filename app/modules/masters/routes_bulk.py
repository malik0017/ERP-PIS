# app/modules/masters/routes_bulk.py
# =============================================================================
# Batch 104 — BULK MASTER DATA ACTIONS + MISSING DATA EXPORT
# -----------------------------------------------------------------------------
# Same problem the recipes had, same shape of answer: master records arrive by
# the hundred from an Excel upload, and the only tools were per-row buttons.
#
# Works across every master type through one set of endpoints rather than one
# per entity, because the tables share the same shape (code / name / status /
# is_active / version). A new master type gets bulk actions for free by being
# added to TABLES.
#
# DELETE IS DELIBERATELY NOT OFFERED HERE.
# A supplier is referenced by purchase orders, an ingredient by recipes and by
# every stock movement ever made, a customer by orders and AR invoices.
# Deleting one orphans that history silently. Deactivate hides it from every
# picker while keeping the history intact, which is what "remove it" actually
# means in an ERP.
# =============================================================================
from __future__ import annotations

import io

from fastapi import APIRouter, Depends, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from urllib.parse import quote

from app.core.rbac import require_action, require_area
from app.database.session import get_db

router = APIRouter(prefix="/masters", tags=["Master Data"])

# master key -> (table, primary key, code column, name column, RBAC area)
TABLES = {
    "customers":   ("customers", "id", "customer_code", "customer_name", "master_data"),
    "suppliers":   ("suppliers", "id", "supplier_code", "supplier_name", "master_data"),
    "ingredients": ("ingredients", "id", "ingredient_code", "name", "master_data"),
    "brands":      ("brands", "id", "brand_code", "brand_name", "master_data"),
    "chefs":       ("chefs", "id", "chef_code", "chef_name", "master_data"),
    "kitchen_sections": ("kitchen_sections", "id", "section_code", "section_name", "master_data"),
    "revenue_streams":  ("revenue_streams", "id", "channel_code", "channel_name", "master_data"),
}


def _ids(form) -> list[int]:
    out = []
    for raw in form.getlist("record_ids"):
        try:
            out.append(int(raw))
        except (TypeError, ValueError):
            continue
    return out


def _back(form, msg: str, variant: str = "success") -> RedirectResponse:
    target = (form.get("return_to") or "/masters").strip()
    sep = "&" if "?" in target else "?"
    return RedirectResponse(
        f"{target}{sep}toast={variant}&title={quote('Master Data')}&msg={quote(msg)}",
        status_code=303)


def _has_col(db: Session, table: str, col: str) -> bool:
    try:
        return bool(db.execute(text("""
            SELECT COUNT(*) FROM information_schema.columns
            WHERE table_schema = DATABASE() AND table_name = :t AND column_name = :c
        """), {"t": table, "c": col}).scalar())
    except Exception:
        return False


async def _apply(request: Request, db: Session, master: str, sets: dict, verb: str):
    """Shared body for the bulk state changes."""
    cfg = TABLES.get(master)
    if not cfg:
        form = await request.form()
        return _back(form, f"Unknown master type '{master}'.", "danger")
    table, pk, _code, _name, area = cfg
    require_action(request, area, "edit")

    form = await request.form()
    ids = _ids(form)
    if not ids:
        return _back(form, "Nothing was selected.", "warning")

    # Only set columns this table actually has — the master tables are not
    # perfectly uniform, and an UPDATE naming a missing column would 500.
    usable = {k: v for k, v in sets.items() if _has_col(db, table, k)}
    if not usable:
        return _back(form, f"{table} has no status columns to change.", "warning")

    ph = ",".join(f":i{n}" for n in range(len(ids)))
    params = {f"i{n}": v for n, v in enumerate(ids)}
    assigns = []
    for i, (col, val) in enumerate(usable.items()):
        params[f"v{i}"] = val
        assigns.append(f"{col} = :v{i}")

    db.execute(text(f"UPDATE {table} SET {', '.join(assigns)} WHERE {pk} IN ({ph})"), params)
    db.commit()
    return _back(form, f"{len(ids)} record(s) {verb}.")


@router.post("/{master}/bulk/activate")
async def bulk_activate(master: str, request: Request, db: Session = Depends(get_db)):
    return await _apply(request, db, master,
                        {"is_active": 1, "status": "ACTIVE"}, "activated")


@router.post("/{master}/bulk/deactivate")
async def bulk_deactivate(master: str, request: Request, db: Session = Depends(get_db)):
    return await _apply(request, db, master,
                        {"is_active": 0, "status": "INACTIVE"}, "deactivated")


@router.post("/{master}/bulk/approve")
async def bulk_approve(master: str, request: Request, db: Session = Depends(get_db)):
    return await _apply(request, db, master,
                        {"approval_status": "Approved", "is_active": 1, "status": "ACTIVE"},
                        "approved")


@router.post("/{master}/bulk/reject")
async def bulk_reject(master: str, request: Request, db: Session = Depends(get_db)):
    return await _apply(request, db, master,
                        {"approval_status": "Rejected", "is_active": 0, "status": "INACTIVE"},
                        "rejected")


# ---------------------------------------------------------------------------
# Missing data export
# ---------------------------------------------------------------------------
@router.get("/recipe-missing-data/export")
def export_missing(request: Request, db: Session = Depends(get_db)):
    """Download the Missing Data report as a fixable workbook.

    Not just a dump of what's broken. The sheet is laid out so it can be
    filled in and handed back:

      * Suggested Action tells you WHICH of the two fixes each row needs
      * Item Code is left blank for you to paste the real code into
      * A second sheet lists the distinct unmapped names, because the same
        "Ranch sauce" appears on many recipes and fixing it once fixes them all

    The two fixes, and how to tell them apart:
      1. It is BOUGHT  -> add it to Master Data > Ingredients, then put the
                          code in the Item Code column.
      2. It is MADE    -> it is a sub-recipe. Create it as its own recipe with
                          its own ingredients, then reference it.
    Everything with a "Sub Recipe Description" in the source file is almost
    certainly case 2 — that column is exactly where your kitchen recorded that
    the component is produced in-house.
    """
    require_area(request, "recipe_list")

    rows = []
    try:
        rows = [dict(r) for r in db.execute(text("""
            SELECT r.recipe_code, r.recipe_name,
                   COALESCE(r.customer_name, '') AS customer_name,
                   COALESCE(r.day_of_week, '')   AS day_of_week,
                   ri.line_no,
                   COALESCE(ri.inventory_code, '') AS inventory_code,
                   COALESCE(ri.item_name, '')      AS item_name,
                   COALESCE(ri.uom, '')            AS uom,
                   COALESCE(ri.qty_batch, 0)       AS qty_batch,
                   COALESCE(ri.qty_per_portion, 0) AS qty_per_portion,
                   -- Batch 104: the model has sub_recipe_code / line_type /
                   -- remark, NOT sub_recipe_description. Verified against the
                   -- actual RecipeIngredient columns before writing this —
                   -- guessing here would have produced an "Unknown column"
                   -- 500 on the one screen meant to help fix bad data.
                   COALESCE(NULLIF(ri.sub_recipe_code, ''),
                            CASE WHEN UPPER(COALESCE(ri.line_type,'')) LIKE '%SUB%'
                                 THEN 'sub-recipe line' ELSE '' END,
                            COALESCE(ri.remark, '')) AS sub_desc
            FROM recipe_ingredients ri
            JOIN recipes r ON r.id = ri.recipe_id
            WHERE COALESCE(ri.inventory_code, '') = ''
               OR ri.inventory_code NOT IN (SELECT ingredient_code FROM ingredients)
            ORDER BY ri.item_name, r.recipe_code, ri.line_no
            LIMIT 5000
        """)).mappings().all()]
    except Exception:
        rows = []

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Data"
    head = PatternFill("solid", fgColor="132947")
    headers = ["Recipe Code", "Recipe Name", "Customer", "Day", "Line",
               "Item Name (unmapped)", "UOM", "Qty/Batch", "Qty/Portion",
               "Sub Recipe / Note", "Suggested Action", "Item Code (fill this in)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head
        ws.column_dimensions[c.column_letter].width = max(16, min(34, len(h) + 6))

    names: dict[str, dict] = {}
    for i, r in enumerate(rows, start=2):
        made = bool(r["sub_desc"])
        action = ("SUB-RECIPE — it is made in-house. Create it as a recipe."
                  if made else
                  "PURCHASED? — add to Master Data > Ingredients, then put the code here.")
        ws.cell(row=i, column=1, value=r["recipe_code"])
        ws.cell(row=i, column=2, value=r["recipe_name"])
        ws.cell(row=i, column=3, value=r["customer_name"])
        ws.cell(row=i, column=4, value=r["day_of_week"])
        ws.cell(row=i, column=5, value=r["line_no"])
        ws.cell(row=i, column=6, value=r["item_name"])
        ws.cell(row=i, column=7, value=r["uom"])
        ws.cell(row=i, column=8, value=float(r["qty_batch"] or 0))
        ws.cell(row=i, column=9, value=float(r["qty_per_portion"] or 0))
        ws.cell(row=i, column=10, value=r["sub_desc"])
        ws.cell(row=i, column=11, value=action)

        key = (r["item_name"] or "").strip().lower()
        if key:
            e = names.setdefault(key, {"name": r["item_name"], "n": 0,
                                       "made": made, "recipes": []})
            e["n"] += 1
            e["made"] = e["made"] or made
            if len(e["recipes"]) < 8:
                e["recipes"].append(r["recipe_code"])

    ws2 = wb.create_sheet("Fix Once — Distinct Items")
    h2 = ["Item Name", "Times Used", "Looks Like", "Used In (sample)", "Item Code (fill this in)"]
    for i, h in enumerate(h2, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head
        ws2.column_dimensions[c.column_letter].width = max(18, min(46, len(h) + 10))
    for i, e in enumerate(sorted(names.values(), key=lambda x: -x["n"]), start=2):
        ws2.cell(row=i, column=1, value=e["name"])
        ws2.cell(row=i, column=2, value=e["n"])
        ws2.cell(row=i, column=3,
                 value="Sub-recipe (made in-house)" if e["made"] else "Purchased ingredient?")
        ws2.cell(row=i, column=4, value=", ".join(e["recipes"]))

    guide = wb.create_sheet("How to fix")
    guide["A1"] = "Recipe Missing Data — how to fix it"
    guide["A1"].font = Font(bold=True, size=14)
    for line in [
        "",
        "Every row below has an ingredient line with no valid item code. Those lines",
        "cost ZERO and never appear in a shortage check, so the recipe's food cost is",
        "understated and production can start without the material being ordered.",
        "",
        "There are only two fixes:",
        "",
        "1. THE ITEM IS BOUGHT",
        "   Add it in Master Data > Ingredients (code, UOM, cost, supplier), then put",
        "   that code in the 'Item Code' column and re-upload.",
        "",
        "2. THE ITEM IS MADE IN-HOUSE  (a sub-recipe)",
        "   Things like 'Ranch sauce', 'Basil mayo', 'Caramelized onion' are produced,",
        "   not purchased. Create each as its own recipe with its own ingredients, then",
        "   reference it from the parent recipe.",
        "   Any row with a value in 'Sub Recipe / Note' is almost certainly this case —",
        "   that column is where your kitchen already recorded that it is made.",
        "",
        "Start with the 'Fix Once' sheet: the same item appears on many recipes, so",
        "fixing it once there fixes every recipe that uses it.",
    ]:
        guide.append([line])
    guide.column_dimensions["A"].width = 88

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ISFC_Recipe_Missing_Data.xlsx"'},
    )

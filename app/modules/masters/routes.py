# app/modules/masters/routes.py
import json
import os
import tempfile
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.core.auth import get_current_user
from app.core.templates import render
from app.core.rbac import require_area, can_access, normalized_role
from app.core.company import get_current_company_id, scope
from app.models.chef import Chef
from app.models.customer import Customer
from app.models.ingredient import Ingredient
from app.models.supplier import Supplier
from app.models.master_data import Brand, KitchenLocation, KitchenSection, MasterRecord, RevenueStream
from app.services.master_data_upload_service import (
    MASTER_TYPES,
    RECIPE_WORKBOOK_SHEETS,
    TEMPLATE_HEADERS,
    WORKBOOK_SHEETS,
    import_master_data_excel,
)


router = APIRouter(tags=["Masters"])


def _company_id(user) -> int:
    return getattr(user, "company_id", None) or 1

def _role_allowed_master_types(request: Request) -> dict:
    role = normalized_role(request)
    if role in {"SUPER_ADMIN", "ADMIN", "ADMINISTRATOR"}:
        return MASTER_TYPES
    mapping = {
        "HEAD_CHEF": ["chefs", "recipes", "recipes_inventory", "inventory", "kitchen_sections"],
        "SECTION_CHEF": ["chefs", "inventory", "kitchen_sections"],
        "STORE_KEEPER": ["inventory", "suppliers"],
        "STORE_MANAGER": ["inventory", "suppliers", "kitchen_sections"],
        "CUSTOMER": ["customers"],
        "QC_MANAGER": ["recipes", "inventory"],
        "PACKING_MANAGER": ["customers", "brands"],
        "DISPATCH_MANAGER": ["customers", "brands"],
    }
    allowed = mapping.get(role, [])
    return {k: v for k, v in MASTER_TYPES.items() if k in allowed}


def _is_master_upload_admin(request: Request) -> bool:
    return normalized_role(request) in {"SUPER_ADMIN", "ADMIN", "ADMINISTRATOR"}


def _assert_master_template_allowed(request: Request, master_type: str):
    if _is_master_upload_admin(request):
        return
    if master_type not in _role_allowed_master_types(request):
        raise HTTPException(status_code=403, detail="You are not allowed to download this master template")


def _status_filter(q, model, status: str | None):
    if status and status != "ALL" and hasattr(model, "status"):
        q = q.filter(getattr(model, "status") == status)
    return q


def _search_filter(q, fields, search: str | None):
    if not search:
        return q
    like = f"%{search}%"
    clauses = [field.like(like) for field in fields]
    return q.filter(or_(*clauses))


MASTER_LIST_CONFIG = {
    "inventory": {
        "title": "Inventory Master",
        "model": Ingredient,
        "route": "/inventory",
        "code_attr": "ingredient_code",
        "name_attr": "name",
        "name_ar_attr": None,
        "search_fields": [Ingredient.ingredient_code, Ingredient.name, Ingredient.category],
        "show_version": False,
        "extra_cols": [("Category", "category"), ("Inventory UOM", "purchase_uom"), ("Recipe UOM", "recipe_uom"), ("Cost/UOM", "unit_cost_standard"), ("Default Supplier", "default_supplier")],
    },
    "customers": {
        "title": "Customers",
        "model": Customer,
        "route": "/customers",
        "code_attr": "customer_code",
        "name_attr": "customer_name",
        "name_ar_attr": "customer_name_ar" if hasattr(Customer, "customer_name_ar") else None,
        "search_fields": [Customer.customer_code, Customer.customer_name],
        "show_version": True,
        "extra_cols": [("Brand", "brand"), ("Phone", "phone"), ("City", "city")],
    },
    "suppliers": {
        "title": "Suppliers",
        "model": Supplier,
        "route": "/suppliers",
        "code_attr": "supplier_code",
        "name_attr": "supplier_name",
        "name_ar_attr": "supplier_name_ar" if hasattr(Supplier, "supplier_name_ar") else None,
        "search_fields": [Supplier.supplier_code, Supplier.supplier_name],
        "show_version": True,
        "extra_cols": [("Category", "category"), ("Phone", "phone"), ("City", "city")],
    },
    "chefs": {
        "title": "Chefs",
        "model": Chef,
        "route": "/chefs",
        "code_attr": "chef_code",
        "name_attr": "chef_name",
        "name_ar_attr": None,
        "search_fields": [Chef.chef_code, Chef.chef_name],
        "show_version": False,
        "extra_cols": [("Job Title", "job_title"), ("Kitchen Section", "kitchen_section"), ("Brand Assign", "brand_assign")],
    },
    "brands": {
        "title": "Brands",
        "model": Brand,
        "route": "/brands",
        "code_attr": "brand_code",
        "name_attr": "brand_name_en",
        "name_ar_attr": "brand_name_ar",
        "search_fields": [Brand.brand_code, Brand.brand_name_en, Brand.brand_name_ar],
        "show_version": False,
        "extra_cols": [("Short Code", "short_code"), ("Revenue Stream", "revenue_stream_name"), ("Default Kitchen", "default_kitchen_code")],
    },
    "revenue_streams": {
        "title": "Revenue Streams",
        "model": RevenueStream,
        "route": "/revenue-streams",
        "code_attr": "stream_code",
        "name_attr": "stream_name",
        "name_ar_attr": None,
        "search_fields": [RevenueStream.stream_code, RevenueStream.stream_name],
        "show_version": False,
        "extra_cols": [("Category", "revenue_category")],
    },
    "kitchen_locations": {
        "title": "Kitchen Locations",
        "model": KitchenLocation,
        "route": "/kitchen-locations",
        "code_attr": "kitchen_code",
        "name_attr": "kitchen_name",
        "name_ar_attr": None,
        "search_fields": [KitchenLocation.kitchen_code, KitchenLocation.kitchen_name, KitchenLocation.city],
        "show_version": False,
        "extra_cols": [("Type", "kitchen_type"), ("City", "city"), ("Brand Supported", "brand_supported"), ("Manager", "manager")],
    },
    "kitchen_sections": {
        "title": "Kitchen Sections",
        "model": KitchenSection,
        "route": "/kitchen-sections",
        "code_attr": "section_code",
        "name_attr": "section_name",
        "name_ar_attr": "section_name_ar",
        "search_fields": [KitchenSection.section_code, KitchenSection.section_name, KitchenSection.section_name_ar],
        "show_version": False,
        "extra_cols": [("Kitchen Code", "kitchen_code"), ("Sequence", "sequence_no")],
    },
}


def _add_template_sheet(wb: Workbook, title: str, headers: list[str], header_row: int = 1):
    ws = wb.create_sheet(title[:31])
    if header_row > 1:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D7F91")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(14, min(38, len(str(header).replace("\n", " ")) + 4))
    ws.freeze_panes = f"A{header_row + 1}"
    return ws


def _workbook_response(wb: Workbook, filename: str):
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/masters")
def masters_home():
    """Batch 19: /masters had no index route (404 from search/pages links)."""
    return RedirectResponse("/module/masters/dashboard", status_code=303)


@router.get("/masters/upload", response_class=HTMLResponse)
def master_upload_page(request: Request, selected: str = "customers"):
    # Full master workbook upload must stay admin/super-admin only.
    if not _is_master_upload_admin(request):
        raise HTTPException(status_code=403, detail="Full Upload Master Data is admin only. Use the scoped upload button from your allowed master page.")
    return render(
        request,
        "masters/upload.html",
        {
            "page_title": "Upload Master Data",
            "master_types": MASTER_TYPES,
            "selected_master_type": selected,
            "is_admin_master_upload": True,
        },
    )


@router.get("/masters/upload/{master_type}", response_class=HTMLResponse)
def scoped_master_upload_page(master_type: str, request: Request):
    allowed_types = _role_allowed_master_types(request)
    if master_type not in allowed_types:
        raise HTTPException(status_code=403, detail="You are not allowed to upload this master type")
    return render(
        request,
        "masters/upload.html",
        {
            "page_title": f"Upload {allowed_types[master_type].get('label', master_type)}",
            "master_types": allowed_types,
            "selected_master_type": master_type,
            "selected_master_label": allowed_types[master_type].get("label", master_type),
            "scoped_master_upload": True,
            "is_admin_master_upload": False,
        },
    )

@router.get("/masters/template/{master_type}")
def download_master_template(master_type: str, request: Request):
    _assert_master_template_allowed(request, master_type)
    wb = Workbook()

    if master_type == "all":
        for sheet_title, key in WORKBOOK_SHEETS:
            _add_template_sheet(wb, sheet_title, TEMPLATE_HEADERS[key], header_row=1)
        return _workbook_response(wb, "master_data_full_template.xlsx")

    if master_type == "recipes":
        _add_template_sheet(wb, "📘 Master – Recipes", TEMPLATE_HEADERS["recipes"], header_row=3)
        _add_template_sheet(wb, "📘 Master – Ingredients (Salus)", TEMPLATE_HEADERS["recipe_ingredients"], header_row=3)
        return _workbook_response(wb, "recipes_template.xlsx")

    if master_type == "recipes_inventory":
        _add_template_sheet(wb, "📘 Master – Recipes", TEMPLATE_HEADERS["recipes"], header_row=3)
        _add_template_sheet(wb, "Recipe Ingredients", TEMPLATE_HEADERS["recipe_ingredients"], header_row=4)
        _add_template_sheet(wb, "Raw material list", TEMPLATE_HEADERS["raw_material_list"], header_row=2)
        return _workbook_response(wb, "recipes_inventory_workbook_template.xlsx")

    if master_type not in TEMPLATE_HEADERS:
        raise HTTPException(status_code=404, detail="Unknown master template")

    label = MASTER_TYPES.get(master_type, {}).get("label", master_type)
    _add_template_sheet(wb, label, TEMPLATE_HEADERS[master_type], header_row=1)
    return _workbook_response(wb, f"{master_type}_template.xlsx")


@router.post("/masters/upload-master-excel", response_class=HTMLResponse)
async def upload_master_excel(
    request: Request,
    master_type: str = Form("all"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    filename = file.filename or ""
    allowed_types = MASTER_TYPES if _is_master_upload_admin(request) else _role_allowed_master_types(request)
    if master_type not in allowed_types:
        raise HTTPException(status_code=403, detail="You are not allowed to upload this master type.")
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel .xlsx file.")

    suffix = os.path.splitext(filename)[1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        try:
            results = import_master_data_excel(db, tmp_path, company_id=_company_id(current_user), master_type=master_type)
            has_errors = any(bool(v.get("errors")) for v in results.values() if isinstance(v, dict))
            toast_type = "error" if has_errors else "success"
            toast_msg = "Master data upload completed with errors" if has_errors else "Master data upload completed"
        except Exception as exc:
            db.rollback()
            results = {master_type: {"created": 0, "updated": 0, "pending": 0, "skipped": 0, "errors": [str(exc)]}}
            toast_type = "error"
            toast_msg = "Master data upload failed"
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    return render(
        request,
        "masters/import_result.html",
        {
            "page_title": "Master Data Upload Result",
            "filename": filename,
            "master_type": master_type,
            "results": results,
            "toast_type": toast_type,
            "toast_message": toast_msg,
        },
    )


@router.get("/masters/archive", response_class=HTMLResponse)
def master_archive(
    request: Request,
    archive_type: str | None = None,
    status: str | None = "ALL",
    search: str | None = None,
    db: Session = Depends(get_db),
):
    # When opened from a master list, show archived/inactive records from that same live master table.
    if archive_type in MASTER_LIST_CONFIG:
        config = MASTER_LIST_CONFIG[archive_type]
        model = config["model"]
        q_live = db.query(model)
        # Batch 77 fix: this used to hardcode company_id == 1, so any
        # company other than #1 saw either nothing or another company's
        # data here. scope() reads the real active company from the session.
        if hasattr(model, "company_id"):
            q_live = scope(q_live, model, get_current_company_id(request))
        if hasattr(model, "status"):
            if status and status != "ALL":
                q_live = q_live.filter(getattr(model, "status") == status)
            else:
                q_live = q_live.filter(getattr(model, "status").in_(["INACTIVE", "Inactive", "ARCHIVED", "Archived"]))
        elif hasattr(model, "is_active"):
            q_live = q_live.filter(getattr(model, "is_active") == False)
        q_live = _search_filter(q_live, config["search_fields"], search)
        rows = q_live.order_by(model.id.desc()).limit(1000).all()
        return render(
            request,
            "masters/list.html",
            {
                "page_title": f"{config['title']} Archive",
                "master_type": archive_type,
                "rows": rows,
                "search": search or "",
                "status": status or "ALL",
                "archive_type": archive_type,
                "master_types": MASTER_TYPES,
                "config": config,
                "can_upload_this_master": False,
                "is_admin_master_upload": False,
            },
        )

    q = db.query(MasterRecord)
    if archive_type and archive_type != "ALL":
        q = q.filter(MasterRecord.master_type == archive_type)
    if status and status != "ALL":
        q = q.filter(MasterRecord.status == status)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(MasterRecord.code.like(like), MasterRecord.name_en.like(like), MasterRecord.name_ar.like(like), MasterRecord.raw_json.like(like)))
    rows = q.order_by(MasterRecord.master_type, MasterRecord.code, MasterRecord.version.desc()).limit(1000).all()
    return render(
        request,
        "masters/list.html",
        {
            "page_title": "Master Archive",
            "master_type": "archive",
            "rows": rows,
            "search": search or "",
            "status": status or "ALL",
            "archive_type": archive_type or "ALL",
            "master_types": MASTER_TYPES,
            "config": {"show_version": True, "extra_cols": [("Master Type", "master_type"), ("Approval", "approval_status")], "code_attr": "code", "name_attr": "name_en", "name_ar_attr": "name_ar"},
            "can_upload_this_master": _is_master_upload_admin(request),
            "is_admin_master_upload": _is_master_upload_admin(request),
        },
    )


def _render_master_list(master_type: str, request: Request, db: Session, search: str | None, status: str | None):
    config = MASTER_LIST_CONFIG[master_type]
    model = config["model"]
    q = db.query(model)
    # Batch 77 fix: same hardcode as master_archive() above — now uses the
    # logged-in user's actual active company instead of a fixed "1".
    if hasattr(model, "company_id"):
        q = scope(q, model, get_current_company_id(request))
    q = _status_filter(q, model, status)
    q = _search_filter(q, config["search_fields"], search)
    rows = q.order_by(model.id.desc()).limit(1000).all()
    return render(
        request,
        "masters/list.html",
        {
            "page_title": config["title"],
            "master_type": master_type,
            "rows": rows,
            "search": search or "",
            "status": status or "ALL",
            "master_types": MASTER_TYPES,
            "config": config,
            "can_upload_this_master": _is_master_upload_admin(request) or master_type in _role_allowed_master_types(request),
            "is_admin_master_upload": _is_master_upload_admin(request),
        },
    )


@router.get("/inventory", response_class=HTMLResponse)
def inventory_list(request: Request, search: str | None = None, status: str | None = "ALL", db: Session = Depends(get_db)):
    return _render_master_list("inventory", request, db, search, status)


@router.get("/customers", response_class=HTMLResponse)
def customers_list(request: Request, search: str | None = None, status: str | None = "ALL", db: Session = Depends(get_db)):
    return _render_master_list("customers", request, db, search, status)


@router.get("/suppliers", response_class=HTMLResponse)
def suppliers_list(request: Request, search: str | None = None, status: str | None = "ALL", db: Session = Depends(get_db)):
    return _render_master_list("suppliers", request, db, search, status)


@router.get("/chefs", response_class=HTMLResponse)
def chefs_list(request: Request, search: str | None = None, status: str | None = "ALL", db: Session = Depends(get_db)):
    return _render_master_list("chefs", request, db, search, status)


@router.get("/brands", response_class=HTMLResponse)
def brands_list(request: Request, search: str | None = None, status: str | None = "ALL", db: Session = Depends(get_db)):
    return _render_master_list("brands", request, db, search, status)


@router.get("/revenue-streams", response_class=HTMLResponse)
def revenue_streams_list(request: Request, search: str | None = None, status: str | None = "ALL", db: Session = Depends(get_db)):
    return _render_master_list("revenue_streams", request, db, search, status)


@router.get("/kitchen-locations", response_class=HTMLResponse)
def kitchen_locations_list(request: Request, search: str | None = None, status: str | None = "ALL", db: Session = Depends(get_db)):
    return _render_master_list("kitchen_locations", request, db, search, status)


@router.get("/kitchen-sections", response_class=HTMLResponse)
def kitchen_sections_list(request: Request, search: str | None = None, status: str | None = "ALL", db: Session = Depends(get_db)):
    return _render_master_list("kitchen_sections", request, db, search, status)


@router.get("/masters/{master_type}/{row_id}", response_class=HTMLResponse)
def master_detail(master_type: str, row_id: int, request: Request, db: Session = Depends(get_db)):
    if master_type == "archive":
        row = db.query(MasterRecord).filter(MasterRecord.id == row_id).first()
        config = {"code_attr": "code", "name_attr": "name_en", "name_ar_attr": "name_ar", "extra_cols": []}
        raw_json = row.raw_json if row else None
        raw_data = None
        if raw_json:
            try:
                raw_data = json.loads(raw_json)
            except Exception:
                raw_data = None
    else:
        config = MASTER_LIST_CONFIG.get(master_type)
        if not config:
            raise HTTPException(status_code=404, detail="Unknown master type")
        model = config["model"]
        q = db.query(model).filter(model.id == row_id)
        # Batch 77 fix: this had NO company filter at all — any logged-in
        # user could view any other company's master record just by
        # guessing/incrementing the row id in the URL. Scope it like every
        # other master-data query in this file.
        if hasattr(model, "company_id"):
            q = scope(q, model, get_current_company_id(request))
        row = q.first()
        raw_data = None
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")
    return render(request, "masters/detail.html", {"page_title": "Master Detail", "master_type": master_type, "row": row, "raw_data": raw_data, "config": config})


# ===== Phase 2E: generic edit and approval routes =====

def _editable_fields_for(master_type: str, row):
    config = MASTER_LIST_CONFIG.get(master_type)
    if not config:
        return []
    fields = []
    for attr, label in [
        (config.get("code_attr"), "Code"),
        (config.get("name_attr"), "Name EN"),
        (config.get("name_ar_attr"), "Name AR"),
    ]:
        if attr and hasattr(row, attr):
            fields.append({"name": attr, "label": label, "value": getattr(row, attr, None), "type": "text"})
    for label, attr in config.get("extra_cols", []):
        if attr and hasattr(row, attr):
            fields.append({"name": attr, "label": label, "value": getattr(row, attr, None), "type": "text"})
    if hasattr(row, "status"):
        fields.append({"name": "status", "label": "Status", "value": getattr(row, "status", "ACTIVE"), "type": "select"})
    return fields


@router.get("/masters/pending", response_class=HTMLResponse)
def pending_masters(request: Request, db: Session = Depends(get_db)):
    rows = (
        db.query(MasterRecord)
        .filter(MasterRecord.approval_status == "PENDING")
        .order_by(MasterRecord.master_type, MasterRecord.code, MasterRecord.version.desc())
        .all()
    )
    return render(request, "masters/pending.html", {"page_title": "Pending Master Approvals", "rows": rows})


@router.post("/masters/archive/{record_id}/approve")
def approve_master_record(record_id: int, db: Session = Depends(get_db)):
    row = db.query(MasterRecord).filter(MasterRecord.id == record_id).first()
    if row:
        row.approval_status = "APPROVED"
        row.status = "ACTIVE"
        row.is_active = True
        db.commit()
    return RedirectResponse(url="/masters/pending?toast=success&msg=Master record approved", status_code=303)


@router.post("/masters/archive/{record_id}/reject")
def reject_master_record(record_id: int, db: Session = Depends(get_db)):
    row = db.query(MasterRecord).filter(MasterRecord.id == record_id).first()
    if row:
        row.approval_status = "REJECTED"
        row.status = "INACTIVE"
        row.is_active = False
        db.commit()
    return RedirectResponse(url="/masters/pending?toast=warning&msg=Master record rejected", status_code=303)


@router.get("/masters/{master_type}/{row_id}/edit", response_class=HTMLResponse)
def edit_master_record(master_type: str, row_id: int, request: Request, db: Session = Depends(get_db)):
    config = MASTER_LIST_CONFIG.get(master_type)
    if not config:
        raise HTTPException(status_code=404, detail="Unknown master type")
    row = db.query(config["model"]).filter(config["model"].id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")
    return render(request, "masters/form.html", {
        "page_title": config["title"],
        "master_type": master_type,
        "row": row,
        "editable_fields": _editable_fields_for(master_type, row),
    })


@router.post("/masters/{master_type}/{row_id}/edit")
async def save_master_record(master_type: str, row_id: int, request: Request, db: Session = Depends(get_db)):
    config = MASTER_LIST_CONFIG.get(master_type)
    if not config:
        raise HTTPException(status_code=404, detail="Unknown master type")
    row = db.query(config["model"]).filter(config["model"].id == row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")
    form = await request.form()
    for field in _editable_fields_for(master_type, row):
        name = field["name"]
        if hasattr(row, name):
            setattr(row, name, str(form.get(name) or "").strip() or None)
    if hasattr(row, "is_active") and hasattr(row, "status"):
        row.is_active = getattr(row, "status", "ACTIVE") in ("ACTIVE", "Active")
    db.commit()
    route = MASTER_LIST_CONFIG[master_type]["route"]
    return RedirectResponse(url=f"{route}?toast=success&msg=Master record updated", status_code=303)

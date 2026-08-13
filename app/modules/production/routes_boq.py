# app/modules/production/routes_boq.py
# =============================================================================
# Batch 105 — BILL OF QUANTITY export for the kitchen
# -----------------------------------------------------------------------------
# What the Head Chef screen was missing: a way to get the consolidated material
# requirement OUT of the system and onto the pass, filtered the way the kitchen
# actually works.
#
# Two views, because they answer two different questions:
#
#   ORDER-WISE     "what does THIS order need"      -> one sheet per order
#   CONSOLIDATED   "what do I need to pull from the
#                   store this morning, in total"    -> one line per ingredient
#                                                      across every order in
#                                                      the filter
#
# The consolidated view is the one that matters at 5am: a store keeper does not
# want twelve separate lists that each ask for flour, they want one line saying
# 43 kg. This is the same explosion the BOM screen does — deliberately reusing
# bom_lines rather than re-imploding the recipes, so the printed sheet and the
# on-screen BOM can never disagree.
#
# SAP B1 calls this the Pick and Pack / Production Order component list; Odoo
# calls it the MO component report. Same document, same purpose.
# =============================================================================
from __future__ import annotations

import io
from datetime import date

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.rbac import require_area
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/production/boq", tags=["Production"])


def _cid(request: Request) -> int:
    return int(request.session.get("company_id") or 1)


def _filters(request: Request) -> dict:
    q = request.query_params
    return {
        "date_from": (q.get("date_from") or "").strip(),
        "date_to": (q.get("date_to") or "").strip(),
        "customer": (q.get("customer") or "").strip(),
        "order_no": (q.get("order_no") or "").strip(),
        "brand": (q.get("brand") or "").strip(),
        "kitchen": (q.get("kitchen") or "").strip(),
    }


def _where(f: dict, cid: int) -> tuple[str, dict]:
    clauses = ["(o.company_id = :cid OR o.company_id IS NULL)"]
    params: dict = {"cid": cid}
    if f["date_from"]:
        clauses.append("o.required_delivery_date >= :df"); params["df"] = f["date_from"]
    if f["date_to"]:
        clauses.append("o.required_delivery_date <= :dt"); params["dt"] = f["date_to"]
    if f["customer"]:
        clauses.append("o.customer_name LIKE :cu"); params["cu"] = f"%{f['customer']}%"
    if f["order_no"]:
        clauses.append("o.order_no LIKE :on"); params["on"] = f"%{f['order_no']}%"
    if f["brand"]:
        clauses.append("COALESCE(o.brand,'') LIKE :br"); params["br"] = f"%{f['brand']}%"
    if f["kitchen"]:
        clauses.append("COALESCE(o.kitchen,'') LIKE :kt"); params["kt"] = f"%{f['kitchen']}%"
    # Cancelled and rejected orders must never reach a picking list.
    clauses.append("COALESCE(o.status,'') NOT IN ('Cancelled','Rejected')")
    return " AND ".join(clauses), params


def consolidated(db: Session, f: dict, cid: int) -> list[dict]:
    where, params = _where(f, cid)
    try:
        return [dict(r) for r in db.execute(text(f"""
            SELECT b.ingredient_code,
                   MAX(COALESCE(b.ingredient_name, b.ingredient_code)) AS item_name,
                   MAX(COALESCE(b.standard_uom, ''))            AS uom,
                   MAX(COALESCE(i.storage_type, ''))            AS storage_type,
                   MAX(COALESCE(NULLIF(b.default_issue_section, ''), i.default_issue_section, '')) AS section,
                   -- Batch 105: total_required_with_waste_standard, NOT a
                   -- "total_qty_standard" (which does not exist). Waste is
                   -- included deliberately: a pick list that ignores expected
                   -- wastage sends the kitchen short every single time.
                   SUM(COALESCE(b.total_required_with_waste_standard,
                                b.required_qty_standard, 0))  AS required_qty,
                   COUNT(DISTINCT b.order_no)                   AS order_count,
                   MAX(COALESCE(b.unit_cost_standard, i.unit_cost_standard, 0)) AS unit_cost
            FROM bom_lines b
            JOIN customer_orders o ON o.order_no = b.order_no
            LEFT JOIN ingredients i ON i.ingredient_code = b.ingredient_code
            WHERE {where}
            GROUP BY b.ingredient_code
            ORDER BY section, item_name
            LIMIT 3000
        """), params).mappings().all()]
    except Exception:
        return []


def order_wise(db: Session, f: dict, cid: int) -> list[dict]:
    where, params = _where(f, cid)
    try:
        return [dict(r) for r in db.execute(text(f"""
            SELECT b.order_no,
                   COALESCE(o.customer_name, '') AS customer_name,
                   COALESCE(o.brand, '')         AS brand,
                   o.required_delivery_date      AS delivery_date,
                   COALESCE(b.recipe_no, '')     AS recipe_no,
                   b.ingredient_code,
                   COALESCE(b.ingredient_name, b.ingredient_code) AS item_name,
                   COALESCE(b.standard_uom, '')  AS uom,
                   COALESCE(b.total_required_with_waste_standard,
                            b.required_qty_standard, 0) AS required_qty
            FROM bom_lines b
            JOIN customer_orders o ON o.order_no = b.order_no
            WHERE {where}
            ORDER BY o.required_delivery_date, b.order_no, b.ingredient_name
            LIMIT 8000
        """), params).mappings().all()]
    except Exception:
        return []


@router.get("/export")
def export_boq(request: Request, db: Session = Depends(get_db)):
    """Bill of Quantity workbook: consolidated pick list + per-order detail."""
    require_area(request, "bom")
    cid = _cid(request)
    f = _filters(request)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    head = PatternFill("solid", fgColor="132947")
    sub = PatternFill("solid", fgColor="EAEFF5")

    def header(ws, cols, title, subtitle):
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = subtitle
        ws["A2"].font = Font(italic=True, size=9)
        for i, h in enumerate(cols, start=1):
            c = ws.cell(row=4, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head
            c.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[c.column_letter].width = max(14, min(34, len(h) + 8))

    crit = ", ".join(v for k, v in f.items() if v) or "all open orders"
    stamp = f"Generated {date.today().isoformat()} · Filter: {crit}"

    # --- Sheet 1: consolidated pick list ---
    rows = consolidated(db, f, cid)
    ws = wb.active
    ws.title = "Pick List (Consolidated)"
    header(ws, ["Section", "Storage", "Item Code", "Ingredient", "UOM",
                "Total Required", "Orders", "Est. Value", "Picked ✓"],
           "CONSOLIDATED PICK LIST", stamp)
    r = 5
    for x in rows:
        qty = float(x["required_qty"] or 0)
        ws.cell(row=r, column=1, value=x["section"] or "—")
        ws.cell(row=r, column=2, value=x["storage_type"] or "")
        ws.cell(row=r, column=3, value=x["ingredient_code"])
        ws.cell(row=r, column=4, value=x["item_name"])
        ws.cell(row=r, column=5, value=x["uom"])
        ws.cell(row=r, column=6, value=round(qty, 3))
        ws.cell(row=r, column=7, value=int(x["order_count"] or 0))
        ws.cell(row=r, column=8, value=round(qty * float(x["unit_cost"] or 0), 2))
        r += 1
    if rows:
        t = ws.cell(row=r, column=4, value="TOTAL")
        t.font = Font(bold=True)
        t.fill = sub
        tv = ws.cell(row=r, column=8,
                     value=round(sum(float(x["required_qty"] or 0) * float(x["unit_cost"] or 0)
                                     for x in rows), 2))
        tv.font = Font(bold=True)
        tv.fill = sub

    # --- Sheet 2: per-order detail ---
    ow = order_wise(db, f, cid)
    ws2 = wb.create_sheet("By Order")
    header(ws2, ["Delivery", "Order", "Customer", "Brand", "Recipe",
                 "Item Code", "Ingredient", "UOM", "Required"],
           "BILL OF QUANTITY — BY ORDER", stamp)
    r = 5
    for x in ow:
        ws2.cell(row=r, column=1, value=str(x["delivery_date"] or ""))
        ws2.cell(row=r, column=2, value=x["order_no"])
        ws2.cell(row=r, column=3, value=x["customer_name"])
        ws2.cell(row=r, column=4, value=x["brand"])
        ws2.cell(row=r, column=5, value=x["recipe_no"])
        ws2.cell(row=r, column=6, value=x["ingredient_code"])
        ws2.cell(row=r, column=7, value=x["item_name"])
        ws2.cell(row=r, column=8, value=x["uom"])
        ws2.cell(row=r, column=9, value=round(float(x["required_qty"] or 0), 3))
        r += 1

    # --- Sheet 3: by customer ---
    ws3 = wb.create_sheet("By Customer")
    header(ws3, ["Customer", "Orders", "Ingredients", "Total Qty"],
           "BILL OF QUANTITY — BY CUSTOMER", stamp)
    agg: dict[str, dict] = {}
    for x in ow:
        e = agg.setdefault(x["customer_name"] or "—",
                           {"orders": set(), "items": set(), "qty": 0.0})
        e["orders"].add(x["order_no"])
        e["items"].add(x["ingredient_code"])
        e["qty"] += float(x["required_qty"] or 0)
    r = 5
    for name, e in sorted(agg.items(), key=lambda kv: -kv[1]["qty"]):
        ws3.cell(row=r, column=1, value=name)
        ws3.cell(row=r, column=2, value=len(e["orders"]))
        ws3.cell(row=r, column=3, value=len(e["items"]))
        ws3.cell(row=r, column=4, value=round(e["qty"], 3))
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="ISFC_BOQ_{date.today().isoformat()}.xlsx"'},
    )


@router.get("/preview")
def preview_boq(request: Request, db: Session = Depends(get_db)):
    """Batch 106 — see the BOQ on screen BEFORE downloading it.

    Downloading blind meant opening Excel to find out whether the filter was
    right, and repeating that until it was. The preview shows exactly what the
    workbook will contain, with the download one click away once it looks
    right.

    Also supplies the picker lists. The Customer / Order / Brand boxes were
    free text against data the user cannot see — typing "Ma'una" when the
    record says "Ma'una Foundation (FRSH)" silently returned nothing, which
    is indistinguishable from "no orders match".
    """
    require_area(request, "bom")
    cid = _cid(request)
    f = _filters(request)

    rows = consolidated(db, f, cid)
    ow = order_wise(db, f, cid)

    by_customer: dict[str, dict] = {}
    for x in ow:
        e = by_customer.setdefault(x["customer_name"] or "—",
                                   {"orders": set(), "items": set(), "qty": 0.0})
        e["orders"].add(x["order_no"])
        e["items"].add(x["ingredient_code"])
        e["qty"] += float(x["required_qty"] or 0)
    customers_summary = sorted(
        # Batch 108: the key is "item_count", NOT "items" — in Jinja, {{ x.items }}
        # resolves to the dict's built-in .items METHOD before it looks for a key
        # of that name, and prints "<built-in method items of dict object ...>".
        [{"customer": k, "orders": len(v["orders"]), "item_count": len(v["items"]),
          "qty": round(v["qty"], 3)} for k, v in by_customer.items()],
        key=lambda r: -r["qty"])

    # Picker options — drawn from orders that could actually appear in a BOQ,
    # not the whole master, so every option returns at least one row.
    def _opts(sql):
        try:
            return [r[0] for r in db.execute(text(sql), {"cid": cid}).all() if r[0]]
        except Exception:
            return []

    open_clause = ("(company_id = :cid OR company_id IS NULL) "
                   "AND COALESCE(status,'') NOT IN ('Cancelled','Rejected')")
    pick = {
        "customers": _opts(f"SELECT DISTINCT customer_name FROM customer_orders "
                           f"WHERE {open_clause} ORDER BY customer_name LIMIT 500"),
        "orders": _opts(f"SELECT DISTINCT order_no FROM customer_orders "
                        f"WHERE {open_clause} ORDER BY order_no DESC LIMIT 500"),
        "brands": _opts(f"SELECT DISTINCT brand FROM customer_orders "
                        f"WHERE {open_clause} AND COALESCE(brand,'') <> '' "
                        f"ORDER BY brand LIMIT 200"),
    }

    total_value = sum(float(x["required_qty"] or 0) * float(x["unit_cost"] or 0) for x in rows)
    by_section: dict[str, int] = {}
    for x in rows:
        by_section[x["section"] or "—"] = by_section.get(x["section"] or "—", 0) + 1

    return render(request, "production/boq_preview.html", {
        "rows": rows,
        "order_rows": ow[:500],
        "customers_summary": customers_summary,
        "pick": pick,
        "filters": f,
        "totals": {
            "ingredients": len(rows),
            "orders": len({x["order_no"] for x in ow}),
            "value": round(total_value, 2),
            "sections": by_section,
        },
        "page_title": "Bill of Quantity",
    })

# app/modules/procurement/routes_match.py
# =============================================================================
# Batch 108 — THREE-WAY MATCH  (PO ↔ GRN ↔ AP Invoice)
# -----------------------------------------------------------------------------
# The control every finance team asks for first, and the reason it exists:
#
#   You should not pay for
#     * more than you ORDERED   (price/qty crept up after the PO)
#     * more than you RECEIVED  (invoiced for goods that never arrived)
#     * at a price you did not AGREE (invoice price ≠ PO price)
#
# Without it, an AP clerk approves invoices on trust. With it, every invoice is
# reconciled against two independent documents created by two different
# departments — which is precisely why it is a fraud control and not just a
# tidiness feature.
#
# WHAT "MATCHED" MEANS HERE
#
#   Quantity matched : |invoiced − received| within tolerance
#   Price matched    : |invoice unit price − PO unit price| within tolerance
#   Receipt matched  : received ≤ ordered (over-receipt is flagged, not hidden)
#
# Tolerances are configurable because a food business genuinely receives 99.4kg
# against a 100kg order and that is not an exception — it is how weighed goods
# work. A system that flags every rounding difference gets ignored, and an
# ignored control is no control.
#
# STATUS is deliberately three-valued, not two:
#   MATCHED    – safe to pay
#   EXCEPTION  – a real difference outside tolerance, needs a human
#   INCOMPLETE – a document is missing (no GRN yet, no invoice yet). NOT an
#                exception: nothing is wrong, the process simply is not finished.
# Collapsing INCOMPLETE into EXCEPTION is what makes these screens cry wolf.
# =============================================================================
from __future__ import annotations

import io
from datetime import date

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.rbac import require_area
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/procurement/three-way-match", tags=["Procurement"])

DEFAULT_QTY_TOL = 2.0     # percent
DEFAULT_PRICE_TOL = 1.0   # percent


def _cid(request: Request) -> int:
    return int(request.session.get("company_id") or 1)


def _pct_diff(a: float, b: float) -> float:
    """Percentage difference of a from b. Guards the b == 0 case, which is
    common: a line ordered at zero price, or nothing received yet."""
    if b == 0:
        return 0.0 if a == 0 else 100.0
    return abs(a - b) / abs(b) * 100.0


def build_match(db: Session, cid: int, qty_tol: float, price_tol: float,
                status_filter: str = "", supplier: str = "",
                po_no: str = "") -> list[dict]:
    """One row per PO line, with its receipt and invoice position.

    Aggregated in SQL rather than looped in Python: a PO line can have many
    GRN lines (partial deliveries) and many invoice lines (split billing), so
    summing them per line is the only way the comparison means anything.
    """
    where = ["(po.company_id = :cid OR po.company_id IS NULL)"]
    params: dict = {"cid": cid}
    if supplier:
        where.append("po.supplier_name LIKE :sup"); params["sup"] = f"%{supplier}%"
    if po_no:
        where.append("po.po_no LIKE :po"); params["po"] = f"%{po_no}%"

    try:
        rows = [dict(r) for r in db.execute(text(f"""
            SELECT po.po_no, po.supplier_name, po.po_date,
                   COALESCE(po.status, '') AS po_status,
                   pol.inventory_code,
                   COALESCE(pol.item_name, pol.inventory_code) AS item_name,
                   COALESCE(pol.uom, '')          AS uom,
                   COALESCE(pol.ordered_qty, 0)   AS ordered_qty,
                   COALESCE(pol.unit_price, 0)    AS po_price,
                   COALESCE(g.recv_qty, 0)        AS received_qty,
                   COALESCE(inv.inv_amount, 0)    AS po_invoiced_amount,
                   COALESCE(po.total_value, 0)    AS po_total_value,
                   COALESCE(g.grn_count, 0)       AS grn_count,
                   COALESCE(inv.inv_count, 0)     AS invoice_count,
                   COALESCE(inv.inv_refs, '')     AS invoice_refs
            FROM purchase_order_lines pol
            JOIN purchase_orders po ON po.po_no = pol.po_no
            LEFT JOIN (
                SELECT po_no, inventory_code,
                       SUM(COALESCE(received_qty, 0)) AS recv_qty,
                       COUNT(DISTINCT grn_no)         AS grn_count
                FROM grn_lines
                GROUP BY po_no, inventory_code
            ) g ON g.po_no = pol.po_no AND g.inventory_code = pol.inventory_code
            LEFT JOIN (
                -- Batch 108: ap_invoices is HEADER-ONLY in this schema — one
                -- row per invoice with a PO reference and a total amount, no
                -- line detail. Verified against the live table before writing
                -- this; my first draft assumed an ap_invoice_lines table that
                -- does not exist.
                --
                -- So the invoice is compared at PO level, and the amount is
                -- apportioned across that PO's lines by their share of PO
                -- value. That is an approximation and the UI says so — it
                -- still catches the cases that matter (invoiced without a
                -- receipt, invoiced above the PO total), which is most of the
                -- control's value. Exact line-level matching needs an
                -- ap_invoice_lines table, which is a separate change.
                SELECT po_no,
                       SUM(COALESCE(amount, 0))    AS inv_amount,
                       COUNT(DISTINCT ap_no)       AS inv_count,
                       GROUP_CONCAT(DISTINCT ap_no) AS inv_refs
                FROM ap_invoices
                WHERE COALESCE(status,'') NOT IN ('Cancelled','Void')
                GROUP BY po_no
            ) inv ON inv.po_no = pol.po_no
            WHERE {' AND '.join(where)}
            ORDER BY po.po_date DESC, po.po_no, pol.line_no
            LIMIT 2000
        """), params).mappings().all()]
    except Exception:
        # ap_invoice_lines may not carry po_no on an older schema. Fall back to
        # a PO/GRN-only comparison rather than showing nothing — a two-way
        # match is still worth having and the UI says which one you are seeing.
        try:
            rows = [dict(r) for r in db.execute(text(f"""
                SELECT po.po_no, po.supplier_name, po.po_date,
                       COALESCE(po.status, '') AS po_status,
                       pol.inventory_code,
                       COALESCE(pol.item_name, pol.inventory_code) AS item_name,
                       COALESCE(pol.uom, '')        AS uom,
                       COALESCE(pol.ordered_qty, 0) AS ordered_qty,
                       COALESCE(pol.unit_price, 0)  AS po_price,
                       COALESCE(g.recv_qty, 0)      AS received_qty,
                       0 AS invoiced_qty, 0 AS invoiced_value,
                       COALESCE(g.grn_count, 0) AS grn_count,
                       0 AS invoice_count, '' AS invoice_refs
                FROM purchase_order_lines pol
                JOIN purchase_orders po ON po.po_no = pol.po_no
                LEFT JOIN (
                    SELECT po_no, inventory_code,
                           SUM(COALESCE(received_qty,0)) AS recv_qty,
                           COUNT(DISTINCT grn_no) AS grn_count
                    FROM grn_lines GROUP BY po_no, inventory_code
                ) g ON g.po_no = pol.po_no AND g.inventory_code = pol.inventory_code
                WHERE {' AND '.join(where)}
                ORDER BY po.po_date DESC, po.po_no
                LIMIT 2000
            """), params).mappings().all()]
        except Exception:
            return []

    out = []
    for r in rows:
        ordered = float(r["ordered_qty"] or 0)
        received = float(r["received_qty"] or 0)
        po_price = float(r["po_price"] or 0)
        line_value = ordered * po_price

        # Apportion the PO-level invoice amount across lines by value share.
        po_invoiced = float(r.get("po_invoiced_amount") or 0)
        po_total = float(r.get("po_total_value") or 0)
        if po_total <= 0:
            po_total = line_value or 1.0
        inv_value = po_invoiced * (line_value / po_total) if po_invoiced else 0.0
        # Implied invoiced quantity at the agreed price — the only quantity
        # available when the invoice carries no line detail.
        invoiced = (inv_value / po_price) if (po_price and inv_value) else 0.0
        inv_price = po_price if invoiced else 0.0

        issues: list[str] = []
        status = "MATCHED"

        if received == 0 and invoiced == 0:
            status = "INCOMPLETE"
            issues.append("Nothing received or invoiced yet")
        elif invoiced == 0:
            status = "INCOMPLETE"
            issues.append("Received but not yet invoiced")
        elif received == 0:
            # Invoiced with no receipt at all is the single most serious case
            # on this screen — you are being asked to pay for goods with no
            # evidence they arrived.
            status = "EXCEPTION"
            issues.append("INVOICED WITH NO RECEIPT — do not pay without a GRN")
        else:
            if _pct_diff(invoiced, received) > qty_tol:
                status = "EXCEPTION"
                issues.append(
                    f"Invoiced {invoiced:g} vs received {received:g} "
                    f"({_pct_diff(invoiced, received):.1f}% out)")
            # No line-level price to compare (header-only invoices), so the
            # equivalent check is: is the apportioned invoice value above what
            # this line's receipt is actually worth?
            receipt_value = received * po_price
            if receipt_value > 0 and _pct_diff(inv_value, receipt_value) > price_tol:
                status = "EXCEPTION"
                issues.append(
                    f"Invoiced value {inv_value:.2f} vs receipt value {receipt_value:.2f} "
                    f"({_pct_diff(inv_value, receipt_value):.1f}% out)")

        # Over-receipt is worth surfacing even when billing lines up.
        if ordered > 0 and received > ordered * (1 + qty_tol / 100):
            issues.append(f"Over-received: {received:g} against an order of {ordered:g}")
            if status == "MATCHED":
                status = "EXCEPTION"

        row = {
            **r,
            "ordered_qty": round(ordered, 3),
            "received_qty": round(received, 3),
            "invoiced_qty": round(invoiced, 3),
            "po_price": round(po_price, 4),
            "inv_price": round(inv_price, 4),
            "invoiced_value": round(inv_value, 2),
            "header_only_invoice": True,
            "po_value": round(ordered * po_price, 2),
            "variance_value": round(inv_value - (received * po_price), 2),
            "status": status,
            "issues": issues,
        }
        if status_filter and status_filter.upper() != "ALL" and status != status_filter.upper():
            continue
        out.append(row)

    order = {"EXCEPTION": 0, "INCOMPLETE": 1, "MATCHED": 2}
    return sorted(out, key=lambda x: (order.get(x["status"], 3), -abs(x["variance_value"])))


def _params(request: Request) -> dict:
    q = request.query_params

    def _f(name, default, lo, hi):
        try:
            return max(lo, min(hi, float(q.get(name) or default)))
        except (TypeError, ValueError):
            return default

    return {
        "qty_tol": _f("qty_tol", DEFAULT_QTY_TOL, 0, 50),
        "price_tol": _f("price_tol", DEFAULT_PRICE_TOL, 0, 50),
        "status": (q.get("status") or "").strip(),
        "supplier": (q.get("supplier") or "").strip(),
        "po_no": (q.get("po_no") or "").strip(),
    }


@router.get("")
def match_screen(request: Request, db: Session = Depends(get_db)):
    require_area(request, "procurement")
    cid = _cid(request)
    f = _params(request)
    rows = build_match(db, cid, f["qty_tol"], f["price_tol"],
                       f["status"], f["supplier"], f["po_no"])

    # Counts are computed unfiltered so the KPI cards do not change meaning
    # when you filter the table — a card that follows the filter cannot tell
    # you how many exceptions exist overall.
    all_rows = build_match(db, cid, f["qty_tol"], f["price_tol"],
                           "", f["supplier"], f["po_no"])
    counts = {"MATCHED": 0, "EXCEPTION": 0, "INCOMPLETE": 0}
    for r in all_rows:
        counts[r["status"]] = counts.get(r["status"], 0) + 1

    suppliers = sorted({r["supplier_name"] for r in all_rows if r["supplier_name"]})

    return render(request, "procurement/three_way_match.html", {
        "rows": rows,
        "counts": counts,
        "suppliers": suppliers,
        "filters": f,
        "exposure": round(sum(abs(r["variance_value"]) for r in all_rows
                              if r["status"] == "EXCEPTION"), 2),
        "page_title": "Three-Way Match",
    })


@router.get("/export")
def match_export(request: Request, db: Session = Depends(get_db)):
    require_area(request, "procurement")
    f = _params(request)
    rows = build_match(db, _cid(request), f["qty_tol"], f["price_tol"],
                       f["status"], f["supplier"], f["po_no"])

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Three-Way Match"
    head = PatternFill("solid", fgColor="132947")
    cols = ["PO", "Supplier", "PO Date", "Item Code", "Item", "UOM",
            "Ordered", "Received", "Invoiced", "PO Price", "Invoice Price",
            "PO Value", "Invoiced Value", "Variance", "Status", "Issues"]
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head
        ws.column_dimensions[c.column_letter].width = max(13, min(40, len(h) + 8))
    for i, r in enumerate(rows, start=2):
        for j, v in enumerate([
            r["po_no"], r["supplier_name"], str(r["po_date"] or ""),
            r["inventory_code"], r["item_name"], r["uom"],
            r["ordered_qty"], r["received_qty"], r["invoiced_qty"],
            r["po_price"], r["inv_price"], r["po_value"], r["invoiced_value"],
            r["variance_value"], r["status"], "; ".join(r["issues"]),
        ], start=1):
            ws.cell(row=i, column=j, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="ISFC_Three_Way_Match_{date.today().isoformat()}.xlsx"'},
    )

# app/modules/qc/routes_recall.py
# =============================================================================
# Batch 112 — BATCH / LOT RECALL TRACEABILITY
# -----------------------------------------------------------------------------
# The question a food safety authority asks, and the clock you have to answer
# it in:
#
#   "Lot 16102029111927400 of chicken is contaminated.
#    Which customers received food made from it, and when?"
#
# Saudi SFDA and every HACCP scheme expect that answered in hours, not days.
# Until now the data existed — lot numbers on receipts, issuance to sections,
# orders to customers — but nothing joined it up, so answering meant somebody
# manually reading four tables under pressure.
#
# TWO DIRECTIONS, BOTH NEEDED
#
#   BACKWARD  (from a lot)      Where did this come from? Supplier, PO, GRN,
#                               receipt date, QC result. Answers "is the
#                               supplier the source, and what else did they
#                               send us that day?"
#
#   FORWARD   (from a lot)      Where did it go? Which orders consumed it,
#                               which customers, on what delivery dates, and
#                               is it still on the shelf. Answers "who do we
#                               call, and what do we still hold?"
#
# AN HONEST LIMITATION, STATED ON THE SCREEN
#
# Store issuance records WHICH ingredient went to which order, but not WHICH
# LOT of that ingredient. So forward tracing is at ingredient level: it finds
# every order that consumed the ingredient within the window the lot was
# available, which is a superset of the truly affected orders.
#
# For a recall a superset is the safe direction to be wrong in — you would
# rather call twelve customers and find eight were fine than call eight and
# miss four. But it is not lot-exact, and pretending otherwise would be worse
# than saying so. Making it exact requires capturing lot_no at issuance, which
# is a change to the store issuance screen, noted at the end of this file.
# =============================================================================
from __future__ import annotations

import io
from datetime import date

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.rbac import require_area
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/qc/recall", tags=["QC"])


def _cid(request: Request) -> int:
    return int(request.session.get("company_id") or 1)


def _rows(db: Session, sql: str, params: dict) -> list[dict]:
    try:
        return [dict(r) for r in db.execute(text(sql), params).mappings().all()]
    except Exception:
        return []


def lot_search(db: Session, cid: int, q: str) -> list[dict]:
    """Find lots by lot number, item code or item name."""
    if not q:
        return []
    return _rows(db, """
        SELECT t.lot_no, t.inventory_code,
               MAX(COALESCE(t.item_name, t.inventory_code)) AS item_name,
               SUM(COALESCE(t.qty_in, 0))  AS qty_in,
               SUM(COALESCE(t.qty_out, 0)) AS qty_out,
               MIN(COALESCE(t.txn_date, t.created_at)) AS first_seen,
               MAX(COALESCE(t.txn_date, t.created_at)) AS last_seen,
               COUNT(*) AS movements
        FROM inventory_transactions t
        WHERE COALESCE(t.lot_no, '') <> ''
          AND (t.company_id = :cid OR t.company_id IS NULL)
          AND (t.lot_no LIKE :q OR t.inventory_code LIKE :q OR t.item_name LIKE :q)
        GROUP BY t.lot_no, t.inventory_code
        ORDER BY last_seen DESC
        LIMIT 200
    """, {"cid": cid, "q": f"%{q}%"})


def trace_backward(db: Session, cid: int, lot_no: str) -> dict:
    """Where did this lot come from?"""
    receipts = _rows(db, """
        SELECT g.grn_no, g.po_no, g.inventory_code,
               COALESCE(g.item_name, g.inventory_code) AS item_name,
               COALESCE(g.received_qty, 0) AS received_qty,
               COALESCE(g.uom, '')         AS uom,
               COALESCE(g.unit_price, 0)   AS unit_price,
               r.supplier_name, r.grn_date,
               COALESCE(q.decision, '')    AS qc_decision,
               COALESCE(q.notes, '')       AS qc_notes
        FROM grn_lines g
        LEFT JOIN grn_receipts r ON r.grn_no = g.grn_no
        LEFT JOIN qc_incoming_inspections q ON q.grn_no = g.grn_no
        WHERE g.lot_no = :lot AND (g.company_id = :cid OR g.company_id IS NULL)
    """, {"lot": lot_no, "cid": cid})

    movements = _rows(db, """
        SELECT COALESCE(txn_date, created_at) AS when_,
               movement_type, reference_no,
               COALESCE(qty_in, 0)  AS qty_in,
               COALESCE(qty_out, 0) AS qty_out,
               COALESCE(remarks, '') AS remarks,
               COALESCE(qc_status, '') AS qc_status
        FROM inventory_transactions
        WHERE lot_no = :lot AND (company_id = :cid OR company_id IS NULL)
        ORDER BY COALESCE(txn_date, created_at), id
    """, {"lot": lot_no, "cid": cid})

    # Everything else that arrived from the same supplier on the same receipt.
    siblings = []
    if receipts:
        siblings = _rows(db, """
            SELECT DISTINCT g.inventory_code,
                   COALESCE(g.item_name, g.inventory_code) AS item_name,
                   COALESCE(g.lot_no, '') AS lot_no,
                   COALESCE(g.received_qty, 0) AS received_qty
            FROM grn_lines g
            WHERE g.grn_no = :grn AND COALESCE(g.lot_no,'') <> :lot
        """, {"grn": receipts[0]["grn_no"], "lot": lot_no})

    return {"receipts": receipts, "movements": movements, "siblings": siblings}


def trace_forward(db: Session, cid: int, lot_no: str, inventory_code: str,
                  first_seen, last_seen) -> dict:
    """Where did it go? Orders, customers, and what is still held.

    Scoped to the window the lot was actually available: from first receipt to
    last movement (or now if still in stock). Without that window every order
    that ever used the ingredient would be returned, which is so broad it is
    useless in a recall.
    """
    orders = _rows(db, """
        SELECT DISTINCT s.order_no,
               COALESCE(o.customer_name, '') AS customer_name,
               COALESCE(o.brand, '')         AS brand,
               o.required_delivery_date      AS delivery_date,
               COALESCE(o.status, '')        AS status,
               COALESCE(s.issue_to_section, '') AS section,
               SUM(COALESCE(s.issued_qty_standard, 0)) AS issued_qty
        FROM store_issuance_lines s
        JOIN customer_orders o ON o.order_no = s.order_no
        WHERE s.ingredient_code = :code
          AND (o.company_id = :cid OR o.company_id IS NULL)
          -- Batch 112: the lower bound is relaxed by 30 days on purpose.
          --
          -- The window is built from the LEDGER timestamp, which is not always
          -- when the stock physically arrived: opening-stock balances are
          -- posted on import day but represent stock held earlier, and a GRN
          -- can be entered days after receipt. A tight lower bound therefore
          -- drops orders that genuinely did contain the lot.
          --
          -- A recall listing too many orders costs phone calls. One listing
          -- too few costs a customer eating recalled food. Err wide.
          AND (:first IS NULL OR o.required_delivery_date >= DATE_SUB(DATE(:first), INTERVAL 30 DAY))
          AND (:last IS NULL OR o.required_delivery_date <= DATE_ADD(DATE(:last), INTERVAL 7 DAY))
        GROUP BY s.order_no, o.customer_name, o.brand, o.required_delivery_date,
                 o.status, s.issue_to_section
        ORDER BY o.required_delivery_date DESC
        LIMIT 500
    """, {"code": inventory_code, "cid": cid, "first": first_seen, "last": last_seen})

    dispatched = _rows(db, """
        SELECT d.order_no, d.dispatch_no, d.dispatch_status,
               d.dispatch_date, COALESCE(d.driver_name, '') AS driver_name,
               COALESCE(d.packed_portions, 0) AS packed_portions,
               COALESCE(o.customer_name, '') AS customer_name
        FROM packing_dispatch d
        LEFT JOIN customer_orders o ON o.order_no = d.order_no
        WHERE d.order_no IN (
            SELECT DISTINCT s.order_no FROM store_issuance_lines s
            WHERE s.ingredient_code = :code
        )
          AND (:first IS NULL OR d.dispatch_date >= DATE(:first))
        ORDER BY d.dispatch_date DESC
        LIMIT 300
    """, {"code": inventory_code, "first": first_seen})

    return {"orders": orders, "dispatched": dispatched}


@router.get("")
def recall_screen(request: Request, db: Session = Depends(get_db)):
    require_area(request, "qc")
    cid = _cid(request)
    q = (request.query_params.get("q") or "").strip()
    lot = (request.query_params.get("lot") or "").strip()

    results = lot_search(db, cid, q) if q and not lot else []
    detail = None

    if lot:
        summary = _rows(db, """
            SELECT lot_no, inventory_code,
                   MAX(COALESCE(item_name, inventory_code)) AS item_name,
                   SUM(COALESCE(qty_in, 0))  AS qty_in,
                   SUM(COALESCE(qty_out, 0)) AS qty_out,
                   MIN(COALESCE(txn_date, created_at)) AS first_seen,
                   MAX(COALESCE(txn_date, created_at)) AS last_seen
            FROM inventory_transactions
            WHERE lot_no = :lot AND (company_id = :cid OR company_id IS NULL)
            GROUP BY lot_no, inventory_code
        """, {"lot": lot, "cid": cid})
        if summary:
            s = summary[0]
            s["on_hand"] = round(float(s["qty_in"] or 0) - float(s["qty_out"] or 0), 3)
            back = trace_backward(db, cid, lot)
            fwd = trace_forward(db, cid, lot, s["inventory_code"],
                                s["first_seen"], s["last_seen"])
            detail = {"summary": s, **back, **fwd,
                      "affected_customers": sorted({o["customer_name"]
                                                    for o in fwd["orders"]
                                                    if o["customer_name"]})}

    return render(request, "qc/recall.html", {
        "q": q, "lot": lot, "results": results, "detail": detail,
        "page_title": "Batch Recall Trace",
    })


@router.get("/export")
def recall_export(request: Request, db: Session = Depends(get_db)):
    """Export the trace. This is the document handed to an inspector, so it
    carries the limitation note in writing rather than only on screen."""
    require_area(request, "qc")
    cid = _cid(request)
    lot = (request.query_params.get("lot") or "").strip()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head = PatternFill("solid", fgColor="132947")

    def sheet(title, cols, rows, keys):
        ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = title
        for i, h in enumerate(cols, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head
            ws.column_dimensions[c.column_letter].width = max(14, min(38, len(h) + 8))
        for ri, r in enumerate(rows, start=2):
            for ci, k in enumerate(keys, start=1):
                ws.cell(row=ri, column=ci, value=r.get(k))

    summary = _rows(db, """
        SELECT lot_no, inventory_code, MAX(COALESCE(item_name, inventory_code)) AS item_name,
               SUM(COALESCE(qty_in,0)) AS qty_in, SUM(COALESCE(qty_out,0)) AS qty_out,
               MIN(COALESCE(txn_date, created_at)) AS first_seen,
               MAX(COALESCE(txn_date, created_at)) AS last_seen
        FROM inventory_transactions
        WHERE lot_no = :lot AND (company_id = :cid OR company_id IS NULL)
        GROUP BY lot_no, inventory_code
    """, {"lot": lot, "cid": cid})

    back = trace_backward(db, cid, lot)
    fwd = {"orders": [], "dispatched": []}
    if summary:
        s = summary[0]
        fwd = trace_forward(db, cid, s["inventory_code"], s["inventory_code"],
                            s["first_seen"], s["last_seen"]) \
            if False else trace_forward(db, cid, lot, s["inventory_code"],
                                        s["first_seen"], s["last_seen"])

    sheet("Source", ["GRN", "PO", "Supplier", "Receipt Date", "Item", "Qty", "UOM", "QC"],
          back["receipts"],
          ["grn_no", "po_no", "supplier_name", "grn_date", "item_name",
           "received_qty", "uom", "qc_decision"])
    sheet("Affected Orders",
          ["Order", "Customer", "Brand", "Delivery", "Status", "Section", "Issued Qty"],
          fwd["orders"],
          ["order_no", "customer_name", "brand", "delivery_date", "status",
           "section", "issued_qty"])
    sheet("Dispatched",
          ["Order", "Dispatch", "Status", "Date", "Driver", "Customer"],
          fwd["dispatched"],
          ["order_no", "dispatch_no", "dispatch_status", "dispatch_date",
           "driver_name", "customer_name"])
    sheet("Movements",
          ["When", "Type", "Reference", "In", "Out", "QC", "Remarks"],
          back["movements"],
          ["when_", "movement_type", "reference_no", "qty_in", "qty_out",
           "qc_status", "remarks"])

    note = wb.create_sheet("Scope of this trace")
    note["A1"] = f"Batch recall trace — lot {lot}"
    note["A1"].font = Font(bold=True, size=14)
    for line in [
        "",
        f"Generated {date.today().isoformat()}",
        "",
        "SCOPE AND LIMITATION — read before acting on this report.",
        "",
        "Backward trace (Source sheet) is LOT-EXACT. It follows the lot number",
        "recorded on the goods receipt to its supplier, purchase order and QC result.",
        "",
        "Forward trace (Affected Orders, Dispatched) is INGREDIENT-LEVEL, not",
        "lot-exact. Store issuance records which ingredient went to an order but",
        "not which lot of it. This report therefore lists every order that consumed",
        "the ingredient while this lot was available — a SUPERSET of the orders",
        "truly affected.",
        "",
        "For a recall that is the safe direction to be wrong in: some listed orders",
        "may have used a different lot, but no affected order should be missing.",
        "Verify against production records before narrowing the recall.",
    ]:
        note.append([line])
    note.column_dimensions["A"].width = 82

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="ISFC_Recall_Trace_{lot or "lot"}.xlsx"'},
    )

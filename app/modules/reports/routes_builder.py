# app/modules/reports/routes_builder.py
# =============================================================================
# Batch 113 — CUSTOM REPORT BUILDER
# -----------------------------------------------------------------------------
# Every "can you add a column / filter / total to that report" request is a
# development request today. This makes the common ones self-service.
#
# THE SECURITY DECISION, AND WHY IT IS NOT A SQL BOX
#
# The obvious build is a textarea that runs whatever SQL the user types. That
# is a data-exfiltration hole and a DROP TABLE waiting to happen, and it
# bypasses every company-scoping rule in the system.
#
# Instead reports are built from DATASETS defined here in code. A dataset
# declares its base query, its selectable columns and its filterable columns.
# The user picks from those. Every generated query is assembled from
# whitelisted identifiers and bound parameters only — a column name that is
# not in the dataset simply cannot reach SQL.
#
# Company scoping is applied by the dataset, not by the user's filter, so a
# saved report cannot be edited into one that reads another company's rows.
#
# WHAT IT DELIBERATELY DOES NOT DO
#
# No joins the datasets do not already contain, no computed expressions, no
# sub-selects. Those are the features that turn a report builder into an
# unsandboxed query tool. If a report genuinely needs a new join, the right
# answer is a new dataset here — five minutes of work, reviewed, and safe for
# everyone afterwards.
# =============================================================================
from __future__ import annotations

import io
import json
from datetime import date

from fastapi import APIRouter, Depends, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from urllib.parse import quote

from app.core.rbac import require_area, require_action
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/reports/builder", tags=["Reports"])

# ---------------------------------------------------------------------------
# Datasets — the only things a report can be built from.
#   col: (sql expression, label, type)   type drives formatting and filters
# ---------------------------------------------------------------------------
DATASETS: dict[str, dict] = {
    "orders": {
        "label": "Customer Orders",
        "area": "production_orders",
        "base": """FROM customer_orders o
                   LEFT JOIN (SELECT order_no, SUM(COALESCE(required_portions,0)) AS portions,
                                     SUM(COALESCE(required_portions,0)*COALESCE(selling_price_per_portion,0)) AS value
                              FROM order_lines GROUP BY order_no) l ON l.order_no = o.order_no""",
        "scope": "(o.company_id = :cid OR o.company_id IS NULL)",
        "columns": {
            "order_no": ("o.order_no", "Order No", "text"),
            "customer_name": ("o.customer_name", "Customer", "text"),
            "brand": ("COALESCE(o.brand,'')", "Brand", "text"),
            "channel": ("COALESCE(o.channel,'')", "Channel", "text"),
            "status": ("COALESCE(o.status,'')", "Status", "text"),
            "order_date": ("o.order_date", "Order Date", "date"),
            "delivery_date": ("o.required_delivery_date", "Delivery Date", "date"),
            "portions": ("COALESCE(l.portions,0)", "Portions", "number"),
            "value": ("COALESCE(l.value,0)", "Sales Value", "money"),
            "kitchen": ("COALESCE(o.kitchen,'')", "Kitchen", "text"),
        },
    },
    "stock": {
        "label": "Stock on Hand",
        "area": "inventory_valuation",
        "base": """FROM ingredients i
                   LEFT JOIN (SELECT inventory_code,
                                     SUM(CASE WHEN qc_status IN ('Pending','Failed') THEN 0
                                              ELSE COALESCE(qty_in,0) END)
                                   - SUM(COALESCE(qty_out,0)) AS on_hand
                              FROM inventory_transactions GROUP BY inventory_code) t
                          ON t.inventory_code = i.ingredient_code""",
        "scope": "1=1",
        "columns": {
            "code": ("i.ingredient_code", "Item Code", "text"),
            "name": ("i.name", "Item", "text"),
            "category": ("COALESCE(i.category,'')", "Category", "text"),
            "supplier": ("COALESCE(i.default_supplier,'')", "Supplier", "text"),
            "storage": ("COALESCE(i.storage_type,'')", "Storage", "text"),
            "uom": ("COALESCE(i.standard_uom,'')", "UOM", "text"),
            "on_hand": ("COALESCE(t.on_hand,0)", "On Hand", "number"),
            "unit_cost": ("COALESCE(i.unit_cost_standard,0)", "Unit Cost", "money"),
            "value": ("COALESCE(t.on_hand,0)*COALESCE(i.unit_cost_standard,0)", "Stock Value", "money"),
            "min_stock": ("COALESCE(i.min_stock_standard,0)", "Minimum", "number"),
        },
    },
    "purchases": {
        "label": "Purchase Orders",
        "area": "procurement",
        "base": """FROM purchase_orders po
                   LEFT JOIN purchase_order_lines pol ON pol.po_no = po.po_no""",
        "scope": "(po.company_id = :cid OR po.company_id IS NULL)",
        "columns": {
            "po_no": ("po.po_no", "PO No", "text"),
            "supplier": ("COALESCE(po.supplier_name,'')", "Supplier", "text"),
            "po_date": ("po.po_date", "PO Date", "date"),
            "status": ("COALESCE(po.status,'')", "Status", "text"),
            "item": ("COALESCE(pol.item_name,'')", "Item", "text"),
            "item_code": ("COALESCE(pol.inventory_code,'')", "Item Code", "text"),
            "qty": ("COALESCE(pol.ordered_qty,0)", "Ordered Qty", "number"),
            "price": ("COALESCE(pol.unit_price,0)", "Unit Price", "money"),
            "value": ("COALESCE(pol.ordered_qty,0)*COALESCE(pol.unit_price,0)", "Line Value", "money"),
        },
    },
    "recipes": {
        "label": "Recipes & Costing",
        "area": "recipe_list",
        "base": "FROM recipes r",
        "scope": "(r.company_id = :cid OR r.company_id IS NULL)",
        "columns": {
            "code": ("r.recipe_code", "Recipe Code", "text"),
            "name": ("r.recipe_name", "Recipe", "text"),
            "customer": ("COALESCE(r.customer_name,'')", "Customer", "text"),
            "category": ("COALESCE(r.category,'')", "Category", "text"),
            "day": ("COALESCE(r.day_of_week,'')", "Menu Day", "text"),
            "status": ("COALESCE(r.status,'')", "Status", "text"),
            "portions": ("COALESCE(r.standard_portions,0)", "Std Portions", "number"),
            "food_cost": ("COALESCE(r.food_cost_per_portion,0)", "Food Cost", "money"),
            "sale_price": ("COALESCE(r.sale_price_per_portion,0)", "Sale Price", "money"),
            "margin": ("COALESCE(r.sale_price_per_portion,0)-COALESCE(r.food_cost_per_portion,0)",
                       "Margin", "money"),
        },
    },
}

OPS = {
    "eq": ("=", "is"), "ne": ("<>", "is not"),
    "gt": (">", "greater than"), "lt": ("<", "less than"),
    "gte": (">=", "at least"), "lte": ("<=", "at most"),
    "contains": ("LIKE", "contains"),
}
AGGS = {"": "None", "sum": "Sum", "avg": "Average", "count": "Count",
        "min": "Minimum", "max": "Maximum"}


def _cid(request: Request) -> int:
    return int(request.session.get("company_id") or 1)


def ensure_schema(db: Session) -> None:
    try:
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS saved_reports (
                id INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                company_id INT NULL,
                name VARCHAR(160) NOT NULL,
                dataset VARCHAR(40) NOT NULL,
                config TEXT NOT NULL,
                is_shared TINYINT(1) NOT NULL DEFAULT 0,
                created_by VARCHAR(120) NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                KEY idx_rep_company (company_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """))
        db.commit()
    except Exception:
        db.rollback()


def build_query(ds_key: str, cfg: dict, cid: int) -> tuple[str, dict, list[dict]]:
    """Assemble SQL from whitelisted parts only.

    Every identifier comes from the dataset definition; every user value is a
    bound parameter. Nothing the user typed reaches SQL as text.
    """
    ds = DATASETS[ds_key]
    cols = ds["columns"]

    chosen = [c for c in (cfg.get("columns") or []) if c in cols]
    if not chosen:
        chosen = list(cols)[:5]

    group_by = [c for c in (cfg.get("group_by") or []) if c in cols]
    aggs = {k: v for k, v in (cfg.get("aggs") or {}).items()
            if k in cols and v in AGGS and v}

    select_parts, meta = [], []
    for c in chosen:
        expr, label, typ = cols[c]
        agg = aggs.get(c)
        if group_by and agg:
            select_parts.append(f"{agg.upper()}({expr}) AS `{c}`")
            meta.append({"key": c, "label": f"{AGGS[agg]} of {label}", "type": typ})
        elif group_by and c not in group_by:
            # A column that is neither grouped nor aggregated would be an
            # arbitrary row's value. MySQL allows it; it is meaningless. Take
            # MAX so the output is at least deterministic, and say so.
            select_parts.append(f"MAX({expr}) AS `{c}`")
            meta.append({"key": c, "label": f"{label} (sample)", "type": typ})
        else:
            select_parts.append(f"{expr} AS `{c}`")
            meta.append({"key": c, "label": label, "type": typ})

    where = [ds["scope"]]
    params: dict = {"cid": cid}
    for i, f in enumerate(cfg.get("filters") or []):
        col = f.get("column")
        op = f.get("op")
        val = f.get("value")
        if col not in cols or op not in OPS or val in (None, ""):
            continue
        expr = cols[col][0]
        sqlop = OPS[op][0]
        if op == "contains":
            where.append(f"{expr} LIKE :f{i}")
            params[f"f{i}"] = f"%{val}%"
        else:
            where.append(f"{expr} {sqlop} :f{i}")
            params[f"f{i}"] = val

    sql = f"SELECT {', '.join(select_parts)} {ds['base']} WHERE {' AND '.join(where)}"
    if group_by:
        sql += " GROUP BY " + ", ".join(cols[g][0] for g in group_by)

    sort = cfg.get("sort")
    if sort in cols:
        direction = "DESC" if (cfg.get("sort_dir") or "asc").lower() == "desc" else "ASC"
        sql += f" ORDER BY `{sort}` {direction}"

    try:
        limit = max(1, min(int(cfg.get("limit") or 500), 5000))
    except (TypeError, ValueError):
        limit = 500
    sql += f" LIMIT {limit}"
    return sql, params, meta


def run_report(db: Session, ds_key: str, cfg: dict, cid: int) -> tuple[list[dict], list[dict], str]:
    if ds_key not in DATASETS:
        return [], [], "Unknown dataset."
    sql, params, meta = build_query(ds_key, cfg, cid)
    try:
        rows = [dict(r) for r in db.execute(text(sql), params).mappings().all()]
        return rows, meta, ""
    except Exception as exc:
        # Surfaced, not swallowed — a report that silently returns nothing is
        # indistinguishable from a report with no matching data.
        return [], meta, f"{type(exc).__name__}: {str(exc)[:180]}"


def _cfg_from_form(form) -> dict:
    aggs = {}
    for c in form.getlist("agg_col"):
        v = form.get(f"agg_{c}")
        if v:
            aggs[c] = v
    filters = []
    fcols, fops, fvals = form.getlist("f_col"), form.getlist("f_op"), form.getlist("f_val")
    for i, c in enumerate(fcols):
        if c and i < len(fops) and i < len(fvals) and fvals[i] != "":
            filters.append({"column": c, "op": fops[i], "value": fvals[i]})
    return {
        "columns": form.getlist("column"),
        "group_by": form.getlist("group_by"),
        "aggs": aggs,
        "filters": filters,
        "sort": form.get("sort") or "",
        "sort_dir": form.get("sort_dir") or "asc",
        "limit": form.get("limit") or 500,
    }


@router.get("")
def builder_home(request: Request, db: Session = Depends(get_db)):
    require_area(request, "reports")
    ensure_schema(db)
    cid = _cid(request)
    ds_key = (request.query_params.get("dataset") or "orders").strip()
    if ds_key not in DATASETS:
        ds_key = "orders"

    saved = []
    try:
        saved = [dict(r) for r in db.execute(text("""
            SELECT id, name, dataset, created_by, is_shared, updated_at
            FROM saved_reports
            WHERE (company_id = :cid OR company_id IS NULL)
            ORDER BY name LIMIT 200
        """), {"cid": cid}).mappings().all()]
    except Exception:
        pass

    return render(request, "reports/builder.html", {
        "datasets": {k: v["label"] for k, v in DATASETS.items()},
        "ds_key": ds_key,
        "columns": {k: {"label": v[1], "type": v[2]} for k, v in DATASETS[ds_key]["columns"].items()},
        "ops": {k: v[1] for k, v in OPS.items()},
        "aggs": AGGS,
        "saved": saved,
        "rows": [], "meta": [], "error": "", "cfg": {},
        "page_title": "Report Builder",
    })


@router.post("/run")
async def builder_run(request: Request, db: Session = Depends(get_db)):
    require_area(request, "reports")
    ensure_schema(db)
    cid = _cid(request)
    form = await request.form()
    ds_key = (form.get("dataset") or "orders").strip()
    if ds_key not in DATASETS:
        ds_key = "orders"

    # The dataset carries its own RBAC area, so a user cannot build a report
    # over data they are not allowed to open elsewhere.
    require_area(request, DATASETS[ds_key]["area"])

    cfg = _cfg_from_form(form)
    rows, meta, err = run_report(db, ds_key, cfg, cid)

    totals = {}
    for m in meta:
        if m["type"] in ("number", "money"):
            totals[m["key"]] = round(sum(float(r.get(m["key"]) or 0) for r in rows), 2)

    saved = []
    try:
        saved = [dict(r) for r in db.execute(text("""
            SELECT id, name, dataset, created_by, is_shared, updated_at FROM saved_reports
            WHERE (company_id = :cid OR company_id IS NULL) ORDER BY name LIMIT 200
        """), {"cid": cid}).mappings().all()]
    except Exception:
        pass

    return render(request, "reports/builder.html", {
        "datasets": {k: v["label"] for k, v in DATASETS.items()},
        "ds_key": ds_key,
        "columns": {k: {"label": v[1], "type": v[2]} for k, v in DATASETS[ds_key]["columns"].items()},
        "ops": {k: v[1] for k, v in OPS.items()},
        "aggs": AGGS, "saved": saved,
        "rows": rows, "meta": meta, "error": err, "cfg": cfg, "totals": totals,
        "page_title": "Report Builder",
    })


@router.post("/save")
async def builder_save(request: Request, db: Session = Depends(get_db)):
    require_action(request, "reports", "add")
    ensure_schema(db)
    form = await request.form()
    name = (form.get("name") or "").strip()
    if not name:
        return RedirectResponse(
            f"/reports/builder?toast=warning&title={quote('Name required')}"
            f"&msg={quote('Give the report a name so it can be found again.')}", status_code=303)

    ds_key = (form.get("dataset") or "orders").strip()
    cfg = _cfg_from_form(form)
    db.execute(text("""
        INSERT INTO saved_reports (company_id, name, dataset, config, is_shared, created_by)
        VALUES (:cid, :n, :d, :c, :s, :by)
    """), {"cid": _cid(request), "n": name[:160], "d": ds_key,
           "c": json.dumps(cfg), "s": 1 if form.get("is_shared") else 0,
           "by": request.session.get("username", "system")})
    db.commit()
    return RedirectResponse(
        f"/reports/builder?toast=success&title={quote('Report Saved')}&msg={quote(name)}",
        status_code=303)


@router.get("/open/{report_id}")
def builder_open(request: Request, report_id: int, db: Session = Depends(get_db)):
    require_area(request, "reports")
    ensure_schema(db)
    cid = _cid(request)
    row = db.execute(text("""
        SELECT * FROM saved_reports WHERE id = :i AND (company_id = :cid OR company_id IS NULL)
    """), {"i": report_id, "cid": cid}).mappings().first()
    if not row:
        return RedirectResponse(
            f"/reports/builder?toast=danger&title={quote('Not found')}"
            f"&msg={quote('That report does not exist for this company.')}", status_code=303)

    ds_key = row["dataset"] if row["dataset"] in DATASETS else "orders"
    require_area(request, DATASETS[ds_key]["area"])
    try:
        cfg = json.loads(row["config"])
    except Exception:
        cfg = {}
    rows, meta, err = run_report(db, ds_key, cfg, cid)
    totals = {m["key"]: round(sum(float(r.get(m["key"]) or 0) for r in rows), 2)
              for m in meta if m["type"] in ("number", "money")}

    saved = [dict(r) for r in db.execute(text("""
        SELECT id, name, dataset, created_by, is_shared, updated_at FROM saved_reports
        WHERE (company_id = :cid OR company_id IS NULL) ORDER BY name LIMIT 200
    """), {"cid": cid}).mappings().all()]

    return render(request, "reports/builder.html", {
        "datasets": {k: v["label"] for k, v in DATASETS.items()},
        "ds_key": ds_key,
        "columns": {k: {"label": v[1], "type": v[2]} for k, v in DATASETS[ds_key]["columns"].items()},
        "ops": {k: v[1] for k, v in OPS.items()},
        "aggs": AGGS, "saved": saved,
        "rows": rows, "meta": meta, "error": err, "cfg": cfg, "totals": totals,
        "report_name": row["name"], "report_id": row["id"],
        "page_title": f"Report — {row['name']}",
    })


@router.post("/delete/{report_id}")
async def builder_delete(request: Request, report_id: int, db: Session = Depends(get_db)):
    require_action(request, "reports", "delete")
    db.execute(text("DELETE FROM saved_reports WHERE id = :i AND (company_id = :cid OR company_id IS NULL)"),
               {"i": report_id, "cid": _cid(request)})
    db.commit()
    return RedirectResponse(
        f"/reports/builder?toast=success&title={quote('Deleted')}&msg={quote('Report removed.')}",
        status_code=303)


@router.post("/export")
async def builder_export(request: Request, db: Session = Depends(get_db)):
    require_area(request, "reports")
    form = await request.form()
    ds_key = (form.get("dataset") or "orders").strip()
    if ds_key not in DATASETS:
        ds_key = "orders"
    require_area(request, DATASETS[ds_key]["area"])

    cfg = _cfg_from_form(form)
    rows, meta, _err = run_report(db, ds_key, cfg, _cid(request))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    head = PatternFill("solid", fgColor="132947")
    for i, m in enumerate(meta, start=1):
        c = ws.cell(row=1, column=i, value=m["label"])
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head
        ws.column_dimensions[c.column_letter].width = max(14, len(m["label"]) + 6)
    for r, row in enumerate(rows, start=2):
        for i, m in enumerate(meta, start=1):
            ws.cell(row=r, column=i, value=row.get(m["key"]))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="ISFC_Report_{date.today().isoformat()}.xlsx"'},
    )

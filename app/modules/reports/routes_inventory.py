# app/modules/reports/routes_inventory.py
# =============================================================================
# Batch 103 — INVENTORY REPORTING, SUPPLIER-WISE
# -----------------------------------------------------------------------------
# What procurement and finance actually ask for, and none of which existed:
#
#   1. Supplier spend      — what have we bought from each supplier, received
#                            vs ordered, and how reliably.
#   2. Stock valuation     — what is on hand, at what cost, by supplier and by
#                            storage type.
#   3. Slow / dead stock   — what was received and never issued.
#   4. Reorder exposure    — what is below its minimum.
#
# Everything reads the ledger (inventory_transactions), never
# ingredients.current_stock, and everything respects the Batch 93 QC gate:
# quantity sitting in QC Hold is received but NOT available, and the two are
# reported separately rather than being quietly added together.
# =============================================================================
from __future__ import annotations

import io
from datetime import date, timedelta

from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.rbac import require_area
from app.core.templates import render
from app.database.session import get_db

router = APIRouter(prefix="/reports/inventory", tags=["Reports"])


def _cid(request: Request) -> int:
    return int(request.session.get("company_id") or 1)


def _rows(db: Session, sql: str, params: dict) -> list[dict]:
    try:
        return [dict(r) for r in db.execute(text(sql), params).mappings().all()]
    except Exception:
        return []


def supplier_summary(db: Session, cid: int, date_from: str, date_to: str) -> list[dict]:
    """Purchasing and receiving performance per supplier.

    ordered vs received is the number procurement argues about, so both are
    shown side by side with the fill rate rather than only the total spend.
    """
    return _rows(db, """
        SELECT po.supplier_name,
               COUNT(DISTINCT po.po_no)                       AS po_count,
               COALESCE(SUM(pol.ordered_qty), 0)              AS ordered_qty,
               COALESCE(SUM(pol.line_value), 0)               AS ordered_value,
               COALESCE(SUM(g.received_qty), 0)               AS received_qty,
               COUNT(DISTINCT g.grn_no)                       AS grn_count
        FROM purchase_orders po
        LEFT JOIN purchase_order_lines pol ON pol.po_no = po.po_no
        LEFT JOIN (
            SELECT gl.po_no, gl.inventory_code, SUM(gl.received_qty) AS received_qty,
                   MIN(gl.grn_no) AS grn_no
            FROM grn_receipt_lines gl GROUP BY gl.po_no, gl.inventory_code
        ) g ON g.po_no = pol.po_no AND g.inventory_code = pol.inventory_code
        WHERE (po.company_id = :cid OR po.company_id IS NULL)
          AND (:df = '' OR po.po_date >= :df)
          AND (:dt = '' OR po.po_date <= :dt)
          AND COALESCE(po.supplier_name, '') <> ''
        GROUP BY po.supplier_name
        ORDER BY ordered_value DESC
        LIMIT 500
    """, {"cid": cid, "df": date_from, "dt": date_to})


def stock_by_supplier(db: Session, cid: int) -> list[dict]:
    """On-hand value grouped by the ingredient's default supplier.

    Answers "how much of our stock money is tied up with each supplier",
    which is what drives negotiation and consolidation decisions.
    """
    return _rows(db, """
        SELECT COALESCE(NULLIF(i.default_supplier, ''), '(no supplier set)') AS supplier_name,
               COUNT(DISTINCT i.ingredient_code) AS item_count,
               COALESCE(SUM(t.on_hand), 0)       AS on_hand_qty,
               COALESCE(SUM(t.on_hand * COALESCE(i.unit_cost_standard, 0)), 0) AS on_hand_value,
               COALESCE(SUM(t.in_qc), 0)         AS qc_hold_qty
        FROM ingredients i
        LEFT JOIN (
            SELECT inventory_code,
                   SUM(CASE WHEN qc_status IN ('Pending','Failed') THEN 0
                            ELSE COALESCE(qty_in, 0) END) - SUM(COALESCE(qty_out, 0)) AS on_hand,
                   SUM(CASE WHEN qc_status = 'Pending' THEN COALESCE(qty_in, 0) ELSE 0 END) AS in_qc
            FROM inventory_transactions
            WHERE (company_id = :cid OR company_id IS NULL)
            GROUP BY inventory_code
        ) t ON t.inventory_code = i.ingredient_code
        GROUP BY supplier_name
        HAVING on_hand_qty <> 0 OR qc_hold_qty <> 0
        ORDER BY on_hand_value DESC
        LIMIT 500
    """, {"cid": cid})


def slow_moving(db: Session, cid: int, days: int = 60) -> list[dict]:
    """Received but not issued in `days`. Capital sitting still, and in a food
    business also a shelf-life risk, so last movement is shown not just qty."""
    return _rows(db, """
        SELECT i.ingredient_code, i.name AS item_name,
               COALESCE(NULLIF(i.default_supplier, ''), '—') AS supplier_name,
               COALESCE(i.storage_type, '') AS storage_type,
               COALESCE(t.on_hand, 0) AS on_hand_qty,
               COALESCE(t.on_hand, 0) * COALESCE(i.unit_cost_standard, 0) AS on_hand_value,
               t.last_out, t.last_in
        FROM ingredients i
        JOIN (
            SELECT inventory_code,
                   SUM(CASE WHEN qc_status IN ('Pending','Failed') THEN 0
                            ELSE COALESCE(qty_in, 0) END) - SUM(COALESCE(qty_out, 0)) AS on_hand,
                   MAX(CASE WHEN COALESCE(qty_out, 0) > 0 THEN COALESCE(txn_date, created_at) END) AS last_out,
                   MAX(CASE WHEN COALESCE(qty_in, 0) > 0 THEN COALESCE(txn_date, created_at) END) AS last_in
            FROM inventory_transactions
            WHERE (company_id = :cid OR company_id IS NULL)
            GROUP BY inventory_code
        ) t ON t.inventory_code = i.ingredient_code
        WHERE t.on_hand > 0
          AND (t.last_out IS NULL OR t.last_out < DATE_SUB(NOW(), INTERVAL :d DAY))
        ORDER BY on_hand_value DESC
        LIMIT 300
    """, {"cid": cid, "d": days})


def below_minimum(db: Session, cid: int) -> list[dict]:
    """Items under their configured minimum — the reorder shortlist."""
    return _rows(db, """
        SELECT i.ingredient_code, i.name AS item_name,
               COALESCE(NULLIF(i.default_supplier, ''), '—') AS supplier_name,
               COALESCE(i.min_stock_standard, 0) AS min_stock,
               COALESCE(t.on_hand, 0) AS on_hand_qty,
               COALESCE(i.min_stock_standard, 0) - COALESCE(t.on_hand, 0) AS shortfall,
               COALESCE(i.unit_cost_standard, 0) AS unit_cost
        FROM ingredients i
        LEFT JOIN (
            SELECT inventory_code,
                   SUM(CASE WHEN qc_status IN ('Pending','Failed') THEN 0
                            ELSE COALESCE(qty_in, 0) END) - SUM(COALESCE(qty_out, 0)) AS on_hand
            FROM inventory_transactions
            WHERE (company_id = :cid OR company_id IS NULL)
            GROUP BY inventory_code
        ) t ON t.inventory_code = i.ingredient_code
        WHERE COALESCE(i.min_stock_standard, 0) > 0
          AND COALESCE(t.on_hand, 0) < COALESCE(i.min_stock_standard, 0)
        ORDER BY shortfall DESC
        LIMIT 300
    """, {"cid": cid})


@router.get("")
def inventory_reports(request: Request, db: Session = Depends(get_db)):
    require_area(request, "reports")
    cid = _cid(request)
    date_from = (request.query_params.get("date_from") or "").strip()
    date_to = (request.query_params.get("date_to") or "").strip()
    slow_days = int(request.query_params.get("slow_days") or 60)

    if not date_from and not date_to:
        date_from = (date.today() - timedelta(days=90)).isoformat()
        date_to = date.today().isoformat()

    suppliers = supplier_summary(db, cid, date_from, date_to)
    for s in suppliers:
        o, r = float(s.get("ordered_qty") or 0), float(s.get("received_qty") or 0)
        s["fill_rate"] = round((r / o * 100) if o else 0, 1)
        s["open_qty"] = round(o - r, 3)

    stock = stock_by_supplier(db, cid)
    slow = slow_moving(db, cid, slow_days)
    low = below_minimum(db, cid)

    return render(request, "reports/inventory.html", {
        "suppliers": suppliers, "stock": stock, "slow": slow, "low": low,
        "filters": {"date_from": date_from, "date_to": date_to, "slow_days": slow_days},
        "totals": {
            "ordered_value": sum(float(s.get("ordered_value") or 0) for s in suppliers),
            "on_hand_value": sum(float(s.get("on_hand_value") or 0) for s in stock),
            "slow_value": sum(float(s.get("on_hand_value") or 0) for s in slow),
            "supplier_count": len(stock),
        },
        "page_title": "Inventory Reports",
    })


@router.get("/export")
def export_inventory(request: Request, db: Session = Depends(get_db)):
    """One workbook, one sheet per report — the format finance actually wants
    rather than four separate CSV downloads to reconcile by hand."""
    require_area(request, "reports")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    cid = _cid(request)
    date_from = (request.query_params.get("date_from") or "").strip()
    date_to = (request.query_params.get("date_to") or "").strip()
    slow_days = int(request.query_params.get("slow_days") or 60)

    wb = Workbook()
    head = PatternFill("solid", fgColor="132947")

    def sheet(title, headers, rows, keys):
        ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = title
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = head
            ws.column_dimensions[c.column_letter].width = max(14, len(h) + 4)
        for r_i, row in enumerate(rows, start=2):
            for c_i, k in enumerate(keys, start=1):
                ws.cell(row=r_i, column=c_i, value=row.get(k))

    sup = supplier_summary(db, cid, date_from, date_to)
    for s in sup:
        o, r = float(s.get("ordered_qty") or 0), float(s.get("received_qty") or 0)
        s["fill_rate"] = round((r / o * 100) if o else 0, 1)
    sheet("Supplier Spend",
          ["Supplier", "POs", "Ordered Qty", "Ordered Value", "Received Qty", "GRNs", "Fill %"],
          sup, ["supplier_name", "po_count", "ordered_qty", "ordered_value",
                "received_qty", "grn_count", "fill_rate"])

    sheet("Stock by Supplier",
          ["Supplier", "Items", "On Hand Qty", "On Hand Value", "In QC Hold"],
          stock_by_supplier(db, cid),
          ["supplier_name", "item_count", "on_hand_qty", "on_hand_value", "qc_hold_qty"])

    sheet(f"Slow Moving {slow_days}d",
          ["Code", "Item", "Supplier", "Storage", "On Hand", "Value", "Last Issued", "Last Received"],
          slow_moving(db, cid, slow_days),
          ["ingredient_code", "item_name", "supplier_name", "storage_type",
           "on_hand_qty", "on_hand_value", "last_out", "last_in"])

    sheet("Below Minimum",
          ["Code", "Item", "Supplier", "Minimum", "On Hand", "Shortfall", "Unit Cost"],
          below_minimum(db, cid),
          ["ingredient_code", "item_name", "supplier_name", "min_stock",
           "on_hand_qty", "shortfall", "unit_cost"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="ISFC_Inventory_Reports_{date.today().isoformat()}.xlsx"'},
    )

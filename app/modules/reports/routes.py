from __future__ import annotations

import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.templates import render
from app.core.rbac import require_area
from app.database.session import get_db

router = APIRouter(tags=["Reports"])



# ---------------------------------------------------------------------------
# Batch 18: EXPORTS registry restored — it was missing from the file, which
# broke EVERY report download (CSV and Excel) with a NameError 500.
# One SQL per report key used by the Reports Center cards and quick exports.
# ---------------------------------------------------------------------------
EXPORTS: dict[str, str] = {
    "recipe-master": """
        SELECT recipe_code, recipe_name, COALESCE(customer_name,'') AS customer,
               COALESCE(category,'') AS category, COALESCE(version,1) AS version,
               COALESCE(status,'') AS status, COALESCE(approval_status,'') AS approval,
               COALESCE(food_cost,0) AS food_cost, COALESCE(total_cost,0) AS total_cost,
               COALESCE(sale_price,0) AS sale_price, COALESCE(missing_cost_lines,0) AS missing_cost_lines
        FROM recipes ORDER BY recipe_code""",
    "recipe-bom": """
        SELECT r.recipe_code, r.recipe_name, ri.inventory_code,
               COALESCE(ri.item_name,'') AS item_name, COALESCE(ri.uom,'') AS uom,
               COALESCE(ri.qty_batch,0) AS qty_batch, COALESCE(ri.qty_per_portion,0) AS qty_per_portion,
               COALESCE(ri.cost_uom,0) AS cost_uom, COALESCE(ri.line_cost,0) AS line_cost
        FROM recipe_ingredients ri JOIN recipes r ON r.id = ri.recipe_id
        ORDER BY r.recipe_code, ri.id""",
    "order-register": """
        SELECT order_no, order_date, customer_name, COALESCE(brand,'') AS brand,
               COALESCE(channel,'') AS channel, COALESCE(required_delivery_date,'') AS delivery_date,
               COALESCE(required_delivery_time,'') AS delivery_time,
               COALESCE(total_planned_portions,0) AS portions,
               COALESCE(total_estimated_food_cost,0) AS food_cost,
               COALESCE(total_estimated_selling_value,0) AS sale_value,
               COALESCE(total_estimated_margin,0) AS margin, status
        FROM customer_orders ORDER BY id DESC""",
    "bom-lines": """
        SELECT order_no, COALESCE(recipe_no,'') AS recipe_no, COALESCE(recipe_name,'') AS recipe_name,
               COALESCE(inventory_code,'') AS inventory_code,
               COALESCE(item_name, ingredient_name,'') AS item_name,
               COALESCE(issue_section, section,'') AS issue_section,
               COALESCE(required_qty, quantity, 0) AS required_qty, COALESCE(uom,'') AS uom,
               COALESCE(estimated_cost,0) AS estimated_cost
        FROM bom_lines ORDER BY order_no, id""",
    "store-issuance": """
        SELECT order_no, COALESCE(inventory_code,'') AS inventory_code,
               COALESCE(item_name, ingredient_name,'') AS item_name,
               COALESCE(required_qty,0) AS required_qty, COALESCE(issued_qty,0) AS issued_qty,
               COALESCE(required_qty,0) - COALESCE(issued_qty,0) AS short_qty,
               COALESCE(finalized,0) AS finalized
        FROM store_issuance_lines ORDER BY order_no, id""",
    "yield-wastage": """
        SELECT order_no, COALESCE(current_section,'') AS section,
               COALESCE(recipe_no,'') AS recipe_no, COALESCE(recipe_name,'') AS recipe_name,
               COALESCE(ingredient_code,'') AS item_code, COALESCE(ingredient_name,'') AS item_name,
               COALESCE(issued_qty_standard,0) AS input_qty,
               COALESCE(processed_qty_standard,0) AS output_qty,
               COALESCE(waste_qty_standard,0) AS waste_qty,
               COALESCE(returned_qty_standard,0) AS return_qty,
               COALESCE(transferred_qty_standard,0) AS transfer_qty,
               COALESCE(transaction_status,'') AS status,
               COALESCE(waste_reason, section_remarks, '') AS reason
        FROM kitchen_section_transactions ORDER BY updated_at DESC""",
    "qc-checks": """
        SELECT order_no, COALESCE(check_point,'') AS check_point, COALESCE(status,'') AS status,
               COALESCE(score,0) AS score, COALESCE(issue_found,'') AS issue_found,
               COALESCE(corrective_action,'') AS corrective_action
        FROM qc_checks ORDER BY id DESC""",
    "packing": """
        SELECT order_no, COALESCE(packed_portions,0) AS packed_portions,
               COALESCE(rejected_portions,0) AS rejected_portions,
               COALESCE(packing_status, dispatch_status,'') AS status,
               COALESCE(packing_remarks, remarks,'') AS remarks
        FROM packing_dispatch ORDER BY id DESC""",
    "dispatch": """
        SELECT order_no, COALESCE(vehicle_no,'') AS vehicle_no, COALESCE(driver_name,'') AS driver_name,
               COALESCE(delivery_temperature,'') AS delivery_temperature,
               COALESCE(dispatch_status,'') AS dispatch_status
        FROM packing_dispatch ORDER BY id DESC""",
    "bom-section-cost": """
        SELECT COALESCE(issue_section, section, 'Unassigned') AS issue_section,
               ROUND(SUM(COALESCE(estimated_cost,0)),2) AS material_cost,
               COUNT(*) AS bom_lines
        FROM bom_lines GROUP BY COALESCE(issue_section, section, 'Unassigned')
        ORDER BY material_cost DESC""",
    "bom-category-cost": """
        SELECT COALESCE(i.main_category,'Uncategorized') AS main_category,
               ROUND(SUM(COALESCE(b.estimated_cost,0)),2) AS material_cost,
               COUNT(*) AS bom_lines
        FROM bom_lines b LEFT JOIN ingredients i ON i.ingredient_code = b.inventory_code
        GROUP BY COALESCE(i.main_category,'Uncategorized')
        ORDER BY material_cost DESC""",
}


def _xlsx_response(rows: list[dict], filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="315EFB")
            c.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        for col in ws.columns:
            width = min(42, max(12, max(len(str(cell.value or "")) for cell in col) + 2))
            ws.column_dimensions[col[0].column_letter].width = width
        ws.freeze_panes = "A2"
    else:
        ws.append(["message"]); ws.append(["No data found"])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


def _rows(db: Session, sql: str, params: dict | None = None):
    try:
        return [dict(r) for r in db.execute(text(sql), params or {}).mappings().all()]
    except Exception:
        return []


def _one(db: Session, sql: str, params: dict | None = None):
    try:
        return db.execute(text(sql), params or {}).scalar() or 0
    except Exception:
        return 0


def _first(db: Session, sql: str, params: dict | None = None):
    rows = _rows(db, sql, params)
    return rows[0] if rows else None


def _money(v):
    try:
        return round(float(v or 0), 2)
    except Exception:
        return 0


def _pct(a, b):
    try:
        return round((float(a or 0) / float(b or 0)) * 100, 2) if float(b or 0) else 0
    except Exception:
        return 0


def _stage_class(label: str, current_status: str | None, has_data: bool = False) -> str:
    current_status = (current_status or "").upper()
    label_up = (label or "").upper()
    if not current_status and has_data:
        return "done"
    order = [
        "ORDER", "HEAD CHEF", "BOM", "STORE", "KITCHEN", "QC", "PACKING", "DISPATCH",
    ]
    status_map = {
        "SUBMITTED": "ORDER",
        "DRAFT": "ORDER",
        "HEAD CHEF APPROVED": "HEAD CHEF",
        "BOM GENERATED": "BOM",
        "STORE PENDING": "STORE",
        "IN PRODUCTION": "KITCHEN",
        "QC PENDING": "QC",
        "QC PASSED": "QC",
        "PACKING PENDING": "PACKING",
        "PACKED": "PACKING",
        "OUT FOR DELIVERY": "DISPATCH",
        "DELIVERED": "DISPATCH",
        "CLOSED": "DISPATCH",
        "DISPATCHED": "DISPATCH",
    }
    current = status_map.get(current_status, "ORDER")
    try:
        i = order.index(label_up)
        c = order.index(current)
        if i < c:
            return "done"
        if i == c:
            return "active"
        return "pending"
    except ValueError:
        return "pending"


@router.get("/relationship-map")
@router.get("/reports/relationship-map")  # Batch 11: alias — dashboards link here
def relationship_map(
    request: Request,
    order_no: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    require_area(request, "relationship")
    """SAP B1-style document relationship map.

    This page intentionally behaves like SAP B1 Relationship Map: a selected
    transaction document sits in the middle and the user can see the documents
    before/after it, plus the master-data objects feeding the document flow.
    """
    company_id = int(request.session.get("company_id") or 1)

    recent_orders = _rows(db, """
        SELECT order_no, customer_name, brand, channel, status,
               COALESCE(required_delivery_date,'') AS delivery_date,
               COALESCE(required_delivery_time,'') AS delivery_time,
               COALESCE(total_planned_portions,0) AS portions,
               COALESCE(total_estimated_food_cost,0) AS food_cost,
               COALESCE(total_estimated_selling_value,0) AS selling_value,
               COALESCE(total_estimated_margin,0) AS margin
        FROM customer_orders
        ORDER BY id DESC
        LIMIT 25
    """)
    if not order_no and recent_orders:
        order_no = recent_orders[0]["order_no"]

    selected_order = None
    if order_no:
        selected_order = _first(db, """
            SELECT order_no, customer_name, brand, channel, status,
                   COALESCE(required_delivery_date,'') AS delivery_date,
                   COALESCE(required_delivery_time,'') AS delivery_time,
                   COALESCE(cooking_date,'') AS cooking_date,
                   COALESCE(cooking_time,'') AS cooking_time,
                   COALESCE(material_receiving_date,'') AS material_receiving_date,
                   COALESCE(material_receiving_time,'') AS material_receiving_time,
                   COALESCE(total_planned_portions,0) AS portions,
                   COALESCE(total_estimated_food_cost,0) AS food_cost,
                   COALESCE(total_estimated_selling_value,0) AS selling_value,
                   COALESCE(total_estimated_margin,0) AS margin
            FROM customer_orders
            WHERE order_no = :order_no
            LIMIT 1
        """, {"order_no": order_no})

    counts = {
        "customers": _one(db, "SELECT COUNT(*) FROM customers WHERE company_id=:company_id", {"company_id": company_id}),
        "brands": _one(db, "SELECT COUNT(*) FROM brands WHERE company_id=:company_id", {"company_id": company_id}),
        "channels": _one(db, "SELECT COUNT(*) FROM revenue_streams WHERE company_id=:company_id", {"company_id": company_id}),
        "inventory": _one(db, "SELECT COUNT(*) FROM ingredients WHERE company_id=:company_id OR company_id IS NULL", {"company_id": company_id}),
        "recipes": _one(db, """
            SELECT COUNT(*) FROM recipes
            WHERE company_id=:company_id AND UPPER(TRIM(COALESCE(status,'')))='ACTIVE' AND COALESCE(is_active,1)=1
        """, {"company_id": company_id}),
        "recipe_bom": _one(db, """
            SELECT COUNT(*) FROM recipe_ingredients ri
            JOIN recipes r ON r.id = ri.recipe_id
            WHERE r.company_id=:company_id AND UPPER(TRIM(COALESCE(r.status,'')))='ACTIVE'
        """, {"company_id": company_id}),
        "orders": _one(db, "SELECT COUNT(*) FROM customer_orders"),
        "head_chef": _one(db, "SELECT COUNT(*) FROM head_chef_plans"),
        "bom_lines": _one(db, "SELECT COUNT(*) FROM bom_lines"),
        "store_lines": _one(db, "SELECT COUNT(*) FROM store_issuance_lines"),
        "section_txns": _one(db, "SELECT COUNT(*) FROM kitchen_section_transactions"),
        "qc": _one(db, "SELECT COUNT(*) FROM qc_checks"),
        "packing": _one(db, "SELECT COUNT(*) FROM packing_dispatch"),
        "dispatch": _one(db, "SELECT COUNT(*) FROM packing_dispatch WHERE dispatch_status IN ('Out for Delivery','Delivered','Dispatched','Closed')"),
    }

    order_params = {"order_no": order_no or ""}
    doc_counts = {
        "order_lines": _one(db, "SELECT COUNT(*) FROM order_lines WHERE order_no=:order_no", order_params),
        "head_chef": _one(db, "SELECT COUNT(*) FROM head_chef_plans WHERE order_no=:order_no", order_params),
        "bom_lines": _one(db, "SELECT COUNT(*) FROM bom_lines WHERE order_no=:order_no", order_params),
        "store_lines": _one(db, "SELECT COUNT(*) FROM store_issuance_lines WHERE order_no=:order_no", order_params),
        "section_txns": _one(db, "SELECT COUNT(*) FROM kitchen_section_transactions WHERE order_no=:order_no", order_params),
        "qc_checks": _one(db, "SELECT COUNT(*) FROM qc_checks WHERE order_no=:order_no", order_params),
        "packing_docs": _one(db, "SELECT COUNT(*) FROM packing_dispatch WHERE order_no=:order_no", order_params),
    }

    flow_status = selected_order.get("status") if selected_order else ""
    doc_flow = [
        {
            "key": "order", "label": "Order", "doc_type": "Sales / Internal Order",
            "doc_no": order_no or "No Order", "url": f"/production/orders/{order_no}" if order_no else "/production/orders",
            "metric": doc_counts["order_lines"], "metric_label": "recipe lines", "amount": _money(selected_order.get("selling_value") if selected_order else 0),
            "state": _stage_class("Order", flow_status, bool(selected_order)), "icon": "bi-cart-check",
        },
        {
            "key": "headchef", "label": "Head Chef", "doc_type": "Approval / Schedule",
            "doc_no": "Cooking + material timing", "url": f"/production/orders/{order_no}" if order_no else "/production/head-chef",
            "metric": doc_counts["head_chef"], "metric_label": "plans", "amount": "", "state": _stage_class("Head Chef", flow_status, doc_counts["head_chef"] > 0), "icon": "bi-calendar2-check",
        },
        {
            "key": "bom", "label": "BOM", "doc_type": "Production BOM",
            "doc_no": "Generated material demand", "url": f"/production/orders/{order_no}" if order_no else "/production/orders",
            "metric": doc_counts["bom_lines"], "metric_label": "BOM lines", "amount": _money(_one(db, "SELECT SUM(COALESCE(estimated_cost,0)) FROM bom_lines WHERE order_no=:order_no", order_params)), "state": _stage_class("BOM", flow_status, doc_counts["bom_lines"] > 0), "icon": "bi-diagram-3",
        },
        {
            "key": "store", "label": "Store", "doc_type": "Issue for Production",
            "doc_no": "Material issue", "url": "/production/store-issuance",
            "metric": doc_counts["store_lines"], "metric_label": "store lines", "amount": "", "state": _stage_class("Store", flow_status, doc_counts["store_lines"] > 0), "icon": "bi-box-arrow-up",
        },
        {
            "key": "kitchen", "label": "Kitchen", "doc_type": "Receipt / Process / Transfer",
            "doc_no": "Section movement", "url": "/production/section/Bakery-Pastry",
            "metric": doc_counts["section_txns"], "metric_label": "movements", "amount": "", "state": _stage_class("Kitchen", flow_status, doc_counts["section_txns"] > 0), "icon": "bi-shop",
        },
        {
            "key": "qc", "label": "QC", "doc_type": "Quality Certificate",
            "doc_no": "QC pass / hold / reject", "url": "/qc",
            "metric": doc_counts["qc_checks"], "metric_label": "checks", "amount": "", "state": _stage_class("QC", flow_status, doc_counts["qc_checks"] > 0), "icon": "bi-patch-check",
        },
        {
            "key": "packing", "label": "Packing", "doc_type": "Trayline / Packing Slip",
            "doc_no": "Packed portions", "url": "/packing",
            "metric": doc_counts["packing_docs"], "metric_label": "packing docs", "amount": "", "state": _stage_class("Packing", flow_status, doc_counts["packing_docs"] > 0), "icon": "bi-box-seam",
        },
        {
            "key": "dispatch", "label": "Dispatch", "doc_type": "Delivery Note",
            "doc_no": "Driver / vehicle / closure", "url": "/dispatch",
            "metric": _one(db, "SELECT COUNT(*) FROM packing_dispatch WHERE order_no=:order_no AND dispatch_status IN ('Out for Delivery','Delivered','Dispatched','Closed')", order_params),
            "metric_label": "delivery docs", "amount": "", "state": _stage_class("Dispatch", flow_status, False), "icon": "bi-truck",
        },
    ]

    master_nodes = [
        {"title": "Customer Master", "metric": counts["customers"], "url": "/customers", "feeds": "Order header", "icon": "bi-people"},
        {"title": "Brand Master", "metric": counts["brands"], "url": "/brands", "feeds": "Order + recipe", "icon": "bi-tags"},
        {"title": "Sales Channel", "metric": counts["channels"], "url": "/revenue-streams", "feeds": "Order commercial route", "icon": "bi-broadcast"},
        {"title": "Recipe Master", "metric": counts["recipes"], "url": "/recipes?status=ACTIVE", "feeds": "Order recipe lines", "icon": "bi-journal-text"},
        {"title": "Recipe BOM", "metric": counts["recipe_bom"], "url": "/recipes/ingredients?status=ACTIVE", "feeds": "Production BOM", "icon": "bi-list-check"},
        {"title": "Inventory Master", "metric": counts["inventory"], "url": "/inventory", "feeds": "BOM cost, UOM, section", "icon": "bi-boxes"},
    ]

    flows = _rows(db, """
        SELECT COALESCE(NULLIF(status,''),'Submitted') AS status, COUNT(*) AS total
        FROM customer_orders
        GROUP BY COALESCE(NULLIF(status,''),'Submitted')
        ORDER BY total DESC, status ASC
    """)

    report_links = [
        {"title": "Order Register", "url": "/reports/export/order-register", "desc": "All orders with customer, brand, delivery, food cost, sale and margin."},
        {"title": "BOM Cost by Section", "url": "/reports/export/bom-section-cost", "desc": "Material cost split by store issue section."},
        {"title": "BOM Cost by Category", "url": "/reports/export/bom-category-cost", "desc": "Material cost split by main/sub category."},
        {"title": "Yield & Wastage", "url": "/reports/yield-wastage", "desc": "Section input/output/waste and transfer analysis."},
        {"title": "QC & Packing", "url": "/reports", "desc": "Final quality, packing and dispatch readiness."},
    ]

    return render(request, "reports/relationship_map.html", {
        "counts": counts,
        "flows": flows,
        "selected_order": selected_order,
        "recent_orders": recent_orders,
        "doc_flow": doc_flow,
        "master_nodes": master_nodes,
        "report_links": report_links,
        "page_title": "SAP B1-style Relationship Map",
    })


@router.get("/reports")
def reports_center(request: Request, db: Session = Depends(get_db)):
    require_area(request, "reports")
    cards = [
        {"group": "Master Data", "title": "Recipe Master & Version Report", "url": "/recipes?status=ALL", "export": "/reports/export/recipe-master", "metric": _one(db, "SELECT COUNT(*) FROM recipes"), "icon": "bi-journal-text"},
        {"group": "Master Data", "title": "Recipe Ingredients / BOM Master", "url": "/recipes/ingredients?status=ACTIVE", "export": "/reports/export/recipe-bom", "metric": _one(db, "SELECT COUNT(*) FROM recipe_ingredients"), "icon": "bi-diagram-3"},
        {"group": "Order Flow", "title": "Customer Order Register",  "url": "/production/orders", "export": "/reports/export/order-register", "metric": _one(db, "SELECT COUNT(*) FROM customer_orders"), "icon": "bi-cart-check"},
        {"group": "Production", "title": "BOM by Customer / Brand / Category / Section", "url": "/production/orders", "export": "/reports/export/bom-lines", "metric": _one(db, "SELECT COUNT(*) FROM bom_lines"), "icon": "bi-boxes"},
        {"group": "Store", "title": "Store Issuance Variance", "url": "/production/store-issuance", "export": "/reports/export/store-issuance", "metric": _one(db, "SELECT COUNT(*) FROM store_issuance_lines"), "icon": "bi-box-arrow-up"},
        {"group": "Kitchen", "title": "Section Yield & Wastage", "url": "/reports/yield-wastage", "export": "/reports/export/yield-wastage", "metric": _one(db, "SELECT COUNT(*) FROM kitchen_section_transactions"), "icon": "bi-graph-down-arrow"},
        {"group": "Quality", "title": "QC Checklist Report", "url": "/qc", "export": "/reports/export/qc-checks", "metric": _one(db, "SELECT COUNT(*) FROM qc_checks"), "icon": "bi-patch-check"},
        {"group": "Logistics", "title": "Trayline / Packing Report", "url": "/packing", "export": "/reports/export/packing", "metric": _one(db, "SELECT COUNT(*) FROM packing_dispatch"), "icon": "bi-box-seam"},
        {"group": "Logistics", "title": "Dispatch & Delivery Report", "url": "/dispatch", "export": "/reports/export/dispatch", "metric": _one(db, "SELECT COUNT(*) FROM packing_dispatch WHERE dispatch_status IN ('Packed','Out for Delivery','Delivered','Dispatched','Closed')"), "icon": "bi-truck"},
    ]

    kpis = {
        "reports": len(cards),
        "active_recipes": _one(db, "SELECT COUNT(*) FROM recipes WHERE UPPER(TRIM(COALESCE(status,'')))='ACTIVE' AND COALESCE(is_active,1)=1"),
        "orders": _one(db, "SELECT COUNT(*) FROM customer_orders"),
        "bom_lines": _one(db, "SELECT COUNT(*) FROM bom_lines"),
        "store_lines": _one(db, "SELECT COUNT(*) FROM store_issuance_lines"),
        "section_txns": _one(db, "SELECT COUNT(*) FROM kitchen_section_transactions"),
        "qc_checks": _one(db, "SELECT COUNT(*) FROM qc_checks"),
        "packing_docs": _one(db, "SELECT COUNT(*) FROM packing_dispatch"),
    }

    yield_rows = _rows(db, """
        SELECT COALESCE(current_section,'Unassigned') AS section_name,
               COUNT(*) AS total_lines,
               ROUND(SUM(COALESCE(issued_qty_standard,0)),2) AS input_qty,
               ROUND(SUM(COALESCE(processed_qty_standard,0)),2) AS output_qty,
               ROUND(SUM(COALESCE(waste_qty_standard,0)),2) AS waste_qty,
               ROUND(CASE WHEN SUM(COALESCE(issued_qty_standard,0)) > 0 THEN SUM(COALESCE(waste_qty_standard,0)) / SUM(COALESCE(issued_qty_standard,0)) * 100 ELSE 0 END,2) AS waste_pct
        FROM kitchen_section_transactions
        GROUP BY COALESCE(current_section,'Unassigned')
        ORDER BY waste_pct DESC, total_lines DESC
        LIMIT 12
    """)

    bom_section = _rows(db, """
        SELECT x.label, COUNT(*) AS lines, ROUND(SUM(x.cost),2) AS cost
        FROM (
            SELECT COALESCE(NULLIF(bl.default_issue_section,''), NULLIF(si.issue_to_section,''), NULLIF(i.default_issue_section,''), 'Unassigned') AS label,
                   COALESCE(NULLIF(bl.estimated_cost,0),
                            COALESCE(bl.total_required_with_waste_standard, bl.required_qty_standard, si.required_qty_with_waste_standard, si.required_qty_standard, 0)
                            * COALESCE(NULLIF(bl.unit_cost_standard,0), NULLIF(i.unit_cost_standard,0), 0), 0) AS cost
            FROM bom_lines bl
            LEFT JOIN store_issuance_lines si ON si.bom_line_id = bl.id
            LEFT JOIN ingredients i ON LOWER(TRIM(i.ingredient_code)) = LOWER(TRIM(bl.ingredient_code))
        ) x
        GROUP BY x.label
        ORDER BY cost DESC, lines DESC
        LIMIT 8
    """)

    bom_category = _rows(db, """
        SELECT x.label, COUNT(*) AS lines, ROUND(SUM(x.cost),2) AS cost
        FROM (
            SELECT COALESCE(NULLIF(bl.ingredient_main_category,''), NULLIF(si.ingredient_main_category,''), NULLIF(i.main_category,''), NULLIF(bl.ingredient_category,''), NULLIF(i.category,''), 'Unassigned') AS label,
                   COALESCE(NULLIF(bl.estimated_cost,0),
                            COALESCE(bl.total_required_with_waste_standard, bl.required_qty_standard, si.required_qty_with_waste_standard, si.required_qty_standard, 0)
                            * COALESCE(NULLIF(bl.unit_cost_standard,0), NULLIF(i.unit_cost_standard,0), 0), 0) AS cost
            FROM bom_lines bl
            LEFT JOIN store_issuance_lines si ON si.bom_line_id = bl.id
            LEFT JOIN ingredients i ON LOWER(TRIM(i.ingredient_code)) = LOWER(TRIM(bl.ingredient_code))
        ) x
        GROUP BY x.label
        ORDER BY cost DESC, lines DESC
        LIMIT 8
    """)

    order_status = _rows(db, """
        SELECT COALESCE(NULLIF(status,''),'Submitted') AS label, COUNT(*) AS total
        FROM customer_orders
        GROUP BY COALESCE(NULLIF(status,''),'Submitted')
        ORDER BY total DESC, label ASC
    """)

    # Batch 68: live financial-report KPIs so the Reports Center links into the
    # completed Finance engine (statements from Batch 67).
    fin = {
        "revenue": _one(db, "SELECT COALESCE(SUM(amount),0) FROM ar_invoices WHERE COALESCE(status,'') <> 'Cancelled'"),
        "ar_open": _one(db, "SELECT COALESCE(SUM(amount-COALESCE(paid_amount,0)),0) FROM ar_invoices WHERE COALESCE(status,'') NOT IN ('Paid','Cancelled')"),
        "ap_open": _one(db, "SELECT COALESCE(SUM(amount-COALESCE(paid_amount,0)),0) FROM ap_invoices WHERE COALESCE(status,'') NOT IN ('Paid','Cancelled')"),
        "journals": _one(db, "SELECT COUNT(*) FROM gl_journals"),
    }

    return render(request, "reports/index.html", {
        "cards": cards,
        "kpis": kpis,
        "yield_rows": yield_rows,
        "bom_section": bom_section,
        "bom_category": bom_category,
        "order_status": order_status,
        "fin": fin,
        "page_title": "Reports Center",
    })


@router.get("/reports/yield-wastage")
def yield_wastage(request: Request, db: Session = Depends(get_db)):
    require_area(request, "reports")
    """Batch 17: professional filters (order / section / status / date range),
    section summary, and CSV export that respects the active filters."""
    q = request.query_params
    f = {
        "order_no": (q.get("order_no") or "").strip(),
        "section": (q.get("section") or "").strip(),
        "status": (q.get("status") or "").strip(),
        "from_date": (q.get("from_date") or "").strip(),
        "to_date": (q.get("to_date") or "").strip(),
    }
    where, params = ["1=1"], {}
    if f["order_no"]:
        where.append("order_no LIKE :ono"); params["ono"] = f"%{f['order_no']}%"
    if f["section"]:
        where.append("COALESCE(current_section,'') = :sec"); params["sec"] = f["section"]
    if f["status"]:
        where.append("COALESCE(transaction_status,'') = :st"); params["st"] = f["status"]
    if f["from_date"]:
        where.append("DATE(updated_at) >= :fd"); params["fd"] = f["from_date"]
    if f["to_date"]:
        where.append("DATE(updated_at) <= :td"); params["td"] = f["to_date"]
    W = " AND ".join(where)

    rows = _rows(db, f"""
        SELECT order_no, current_section, recipe_no, recipe_name, ingredient_code, ingredient_name,
               standard_uom,
               COALESCE(issued_qty_standard,0) AS input_qty,
               COALESCE(processed_qty_standard,0) AS output_qty,
               COALESCE(waste_qty_standard,0) AS waste_qty,
               COALESCE(returned_qty_standard,0) AS return_qty,
               COALESCE(transferred_qty_standard,0) AS transferred_qty,
               ROUND(CASE WHEN COALESCE(issued_qty_standard,0) > 0 THEN COALESCE(waste_qty_standard,0) / COALESCE(issued_qty_standard,0) * 100 ELSE 0 END,2) AS waste_pct,
               transaction_status, waste_reason, section_remarks
        FROM kitchen_section_transactions
        WHERE {W}
        ORDER BY updated_at DESC, order_no DESC
        LIMIT 1000
    """, params)

    # Batch 21: downloadable SECTION SUMMARY (respects active filters).
    if (q.get("export") or "") == "summary":
        import csv, io
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Section", "Input", "Output", "Waste", "Waste %"])
        srows = _rows(db, f"""
            SELECT COALESCE(current_section,'Unassigned') AS section_name,
                   ROUND(SUM(COALESCE(issued_qty_standard,0)),2) AS input_qty,
                   ROUND(SUM(COALESCE(processed_qty_standard,0)),2) AS output_qty,
                   ROUND(SUM(COALESCE(waste_qty_standard,0)),2) AS waste_qty,
                   ROUND(CASE WHEN SUM(COALESCE(issued_qty_standard,0)) > 0 THEN SUM(COALESCE(waste_qty_standard,0)) / SUM(COALESCE(issued_qty_standard,0)) * 100 ELSE 0 END,2) AS waste_pct
            FROM kitchen_section_transactions
            WHERE {W}
            GROUP BY COALESCE(current_section,'Unassigned')
            ORDER BY waste_pct DESC
        """, params)
        for r in srows:
            w.writerow([r["section_name"], r["input_qty"], r["output_qty"], r["waste_qty"], r["waste_pct"]])
        from fastapi.responses import Response as _Resp
        return _Resp(buf.getvalue(), media_type="text/csv",
                     headers={"Content-Disposition": 'attachment; filename="yield_section_summary.csv"'})

    if (q.get("export") or "") == "csv":
        import csv, io
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Order", "Section", "Recipe No", "Recipe", "Item Code", "Item", "UOM",
                    "Input", "Output", "Waste", "Waste %", "Return", "Transfer", "Status", "Reason"])
        for r in rows:
            w.writerow([r["order_no"], r["current_section"], r["recipe_no"], r["recipe_name"],
                        r["ingredient_code"], r["ingredient_name"], r["standard_uom"],
                        r["input_qty"], r["output_qty"], r["waste_qty"], r["waste_pct"],
                        r["return_qty"], r["transferred_qty"], r["transaction_status"],
                        (r["waste_reason"] or r["section_remarks"] or "")])
        from fastapi.responses import Response as _Resp
        return _Resp(buf.getvalue(), media_type="text/csv",
                     headers={"Content-Disposition": 'attachment; filename="yield_wastage.csv"'})

    summary = _rows(db, f"""
        SELECT COALESCE(current_section,'Unassigned') AS section_name,
               ROUND(SUM(COALESCE(issued_qty_standard,0)),2) AS input_qty,
               ROUND(SUM(COALESCE(processed_qty_standard,0)),2) AS output_qty,
               ROUND(SUM(COALESCE(waste_qty_standard,0)),2) AS waste_qty,
               ROUND(CASE WHEN SUM(COALESCE(issued_qty_standard,0)) > 0 THEN SUM(COALESCE(waste_qty_standard,0)) / SUM(COALESCE(issued_qty_standard,0)) * 100 ELSE 0 END,2) AS waste_pct
        FROM kitchen_section_transactions
        WHERE {W}
        GROUP BY COALESCE(current_section,'Unassigned')
        ORDER BY waste_pct DESC
    """, params)

    sections = [r["s"] for r in _rows(db, "SELECT DISTINCT COALESCE(current_section,'') AS s FROM kitchen_section_transactions ORDER BY 1") if r["s"]]
    statuses = [r["s"] for r in _rows(db, "SELECT DISTINCT COALESCE(transaction_status,'') AS s FROM kitchen_section_transactions ORDER BY 1") if r["s"]]

    return render(request, "reports/yield_wastage.html", {
        "rows": rows, "summary": summary, "filters": f,
        "sections": sections, "statuses": statuses,
        "page_title": "Yield & Wastage Report",
    })


@router.get("/reports/export/{report_key}")
def export_report(request: Request, report_key: str, format: str = Query(default="csv"), db: Session = Depends(get_db)):
    require_area(request, "reports")
    sql = EXPORTS.get(report_key)
    if not sql:
        sql = EXPORTS["order-register"]
        report_key = "order-register"
    data = _rows(db, sql)
    if str(format).lower() in {"xlsx", "excel"}:
        return _xlsx_response(data, f"isfc_{report_key}.xlsx")
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    if data:
        writer.writerow(data[0].keys())
        for row in data:
            writer.writerow([row.get(k, "") for k in data[0].keys()])
    else:
        writer.writerow(["message"])
        writer.writerow(["No data found for this report"])
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=isfc_{report_key}.csv"},
    )


@router.get("/reports/printable/{form_type}/{order_no}", response_class=HTMLResponse)
def printable_form(form_type: str, order_no: str, request: Request, db: Session = Depends(get_db)):
    require_area(request, "reports")
    """Six professional printable documents, one route.

    order-sheet | bom-sheet | store-issue-slip | qc-certificate | packing-slip | delivery-note
    Each form loads its own dataset; the template renders a print-ready page
    with company header, document meta, signatures and auto window.print().
    """
    allowed = {
        "order-sheet": "Order Sheet", "bom-sheet": "BOM Sheet", "store-issue-slip": "Store Issue Slip",
        "qc-certificate": "QC Certificate", "packing-slip": "Packing Slip", "delivery-note": "Delivery Note",
    }
    title = allowed.get(form_type)
    if not title:
        return render(request, "errors/404.html", {}, status_code=404)

    order = _first(db, """
        SELECT order_no, customer_name, COALESCE(customer_no,'') AS customer_no,
               COALESCE(brand,'') AS brand, COALESCE(channel,'') AS channel, status,
               COALESCE(order_date,'') AS order_date,
               COALESCE(required_delivery_date,'') AS required_delivery_date,
               COALESCE(required_delivery_time,'') AS required_delivery_time,
               COALESCE(cooking_date,'') AS cooking_date, COALESCE(cooking_time,'') AS cooking_time,
               COALESCE(total_planned_portions,0) AS total_planned_portions,
               COALESCE(total_estimated_food_cost,0) AS total_estimated_food_cost,
               COALESCE(total_estimated_selling_value,0) AS total_estimated_selling_value,
               COALESCE(total_estimated_margin,0) AS total_estimated_margin
        FROM customer_orders WHERE order_no=:order_no LIMIT 1
    """, {"order_no": order_no}) or {"order_no": order_no}

    lines, extra = [], {}

    if form_type == "order-sheet":
        lines = _rows(db, """
            SELECT recipe_no, recipe_name, portions, planned_batches,
                   COALESCE(food_cost,0) AS food_cost, COALESCE(selling_value,0) AS selling_value,
                   COALESCE(margin,0) AS margin
            FROM order_lines WHERE order_no=:o ORDER BY line_no
        """, {"o": order_no})

    elif form_type == "bom-sheet":
        lines = _rows(db, """
            SELECT COALESCE(ingredient_main_category,'') AS main_category, ingredient_code, ingredient_name,
                   ROUND(COALESCE(total_required_with_waste_standard,0),4) AS qty, standard_uom,
                   ROUND(COALESCE(total_estimated_cost,0),2) AS cost,
                   COALESCE(default_issue_section,'') AS section
            FROM bom_lines WHERE order_no=:o
            ORDER BY default_issue_section, ingredient_main_category, ingredient_name
        """, {"o": order_no})

    elif form_type == "store-issue-slip":
        lines = _rows(db, """
            SELECT ingredient_code, ingredient_name,
                   ROUND(COALESCE(required_qty_standard,0),4) AS required_qty,
                   ROUND(COALESCE(input_material_issued,0),4) AS issued_qty,
                   COALESCE(issued_uom, standard_uom, '') AS uom,
                   COALESCE(issue_to_section,'') AS section,
                   COALESCE(lot_no,'') AS lot_no, COALESCE(supplier_name,'') AS supplier,
                   CASE WHEN COALESCE(finalized,0)=1 THEN 'Issued' ELSE 'Pending' END AS line_status
            FROM store_issuance_lines WHERE order_no=:o
            ORDER BY issue_to_section, ingredient_name
        """, {"o": order_no})

    elif form_type == "qc-certificate":
        lines = _rows(db, """
            SELECT qc_no, COALESCE(check_type,'') AS check_type, COALESCE(qc_status,'') AS status,
                   COALESCE(score,0) AS score, COALESCE(issue_found,'') AS issue_found,
                   COALESCE(corrective_action,'') AS corrective_action,
                   COALESCE(checked_by,'') AS checked_by, checked_at
            FROM qc_checks WHERE order_no=:o ORDER BY id DESC
        """, {"o": order_no})

    elif form_type in ("packing-slip", "delivery-note"):
        extra["pd"] = _first(db, """
            SELECT dispatch_no, COALESCE(packed_portions,0) AS packed_portions,
                   COALESCE(rejected_portions,0) AS rejected_portions,
                   COALESCE(dispatch_status,'') AS dispatch_status,
                   COALESCE(vehicle_no,'') AS vehicle_no, COALESCE(driver_name,'') AS driver_name,
                   COALESCE(delivery_temp,'') AS delivery_temp,
                   COALESCE(dispatch_date,'') AS dispatch_date, COALESCE(remarks,'') AS remarks
            FROM packing_dispatch WHERE order_no=:o ORDER BY id DESC LIMIT 1
        """, {"o": order_no}) or {}
        lines = _rows(db, """
            SELECT recipe_no, recipe_name, portions
            FROM order_lines WHERE order_no=:o ORDER BY line_no
        """, {"o": order_no})

    return render(request, "reports/printable_form.html", {
        "title": title, "order": order, "lines": lines, "extra": extra,
        "form_type": form_type, "printed_by": request.session.get("username", ""),
    })


# ============================================================================
# Batch 12 — SAP-style Relationship TREE (lazy drill-down across all modules)
# ============================================================================
# Page:  /reports/relationship-tree
# API :  /reports/api/relationship-node?type=<node_type>&key=<id>
#
# Every node returns: { label, badge, url?, children_type?, leaf }
# The tree starts at the module roots (Sales, Production, Inventory,
# Procurement, Finance, Masters) and each click lazily loads that node's
# children from the DB, so the whole ERP is navigable as one linked tree.

from fastapi.responses import JSONResponse as _JSON


def _tree_rows(db, sql, params=None):
    try:
        return [dict(r) for r in db.execute(text(sql), params or {}).mappings().all()]
    except Exception:
        return []


def _tree_count(db, sql, params=None):
    try:
        return int(db.execute(text(sql), params or {}).scalar() or 0)
    except Exception:
        return 0


@router.get("/reports/relationship-tree")
def relationship_tree(request: Request, db: Session = Depends(get_db)):
    require_area(request, "relationship")
    """SAP B1-style relationship tree page. Roots show core module coverage;
    every node expands to its linked documents down to line level."""
    roots = [
        {"type": "root_sales",       "label": "Sales & Orders",       "icon": "shopping-cart",
         "badge": _tree_count(db, "SELECT COUNT(*) FROM customer_orders")},
        {"type": "root_production",  "label": "Production",           "icon": "activity",
         "badge": _tree_count(db, "SELECT COUNT(*) FROM bom_lines")},
        {"type": "root_inventory",   "label": "Inventory",            "icon": "box",
         "badge": _tree_count(db, "SELECT COUNT(*) FROM ingredients")},
        {"type": "root_procurement", "label": "Procurement",          "icon": "shopping-bag",
         "badge": _tree_count(db, "SELECT COUNT(*) FROM purchase_orders")},
        {"type": "root_finance",     "label": "Finance (AR)",         "icon": "dollar-sign",
         "badge": _tree_count(db, "SELECT COUNT(*) FROM ar_invoices")},
        {"type": "root_masters",     "label": "Master Data",          "icon": "database",
         "badge": _tree_count(db, "SELECT COUNT(*) FROM customers")},
    ]
    return render(request, "reports/relationship_tree.html", {
        "roots": roots, "page_title": "Relationship Tree",
    })


@router.get("/reports/api/relationship-node")
def relationship_node(request: Request,
                      node_type: str = Query(alias="type", default=""),
                      key: str = Query(default=""),
                      db: Session = Depends(get_db)):
    require_area(request, "relationship")
    """Children of one tree node. Read-only, safe fallbacks everywhere."""
    t_, k = node_type, key
    out: list[dict] = []

    # ---------------- module roots ----------------
    if t_ == "root_sales":
        for r in _tree_rows(db, """
            SELECT order_no, customer_name, status FROM customer_orders
            ORDER BY id DESC LIMIT 25"""):
            out.append({"type": "order", "key": r["order_no"],
                        "label": f"{r['order_no']} — {r['customer_name']}",
                        "badge": r["status"], "url": f"/production/orders/{r['order_no']}"})
    elif t_ == "root_production":
        for r in _tree_rows(db, """
            SELECT DISTINCT order_no FROM bom_lines ORDER BY order_no DESC LIMIT 25"""):
            out.append({"type": "order_bom", "key": r["order_no"],
                        "label": f"BOM · {r['order_no']}",
                        "badge": _tree_count(db, "SELECT COUNT(*) FROM bom_lines WHERE order_no=:o", {"o": r["order_no"]}),
                        "url": f"/production/orders/{r['order_no']}"})
    elif t_ == "root_inventory":
        for r in _tree_rows(db, """
            SELECT COALESCE(main_category,'Uncategorized') AS cat, COUNT(*) AS n
            FROM ingredients GROUP BY COALESCE(main_category,'Uncategorized')
            ORDER BY n DESC LIMIT 30"""):
            out.append({"type": "inv_category", "key": r["cat"],
                        "label": r["cat"], "badge": r["n"]})
    elif t_ == "root_procurement":
        for r in _tree_rows(db, """
            SELECT po_no, COALESCE(supplier_name,'') AS supplier, COALESCE(status,'') AS status
            FROM purchase_orders ORDER BY id DESC LIMIT 25"""):
            out.append({"type": "po", "key": r["po_no"],
                        "label": f"{r['po_no']} — {r['supplier']}",
                        "badge": r["status"], "url": f"/procurement/po/{r['po_no']}"})
    elif t_ == "root_finance":
        for r in _tree_rows(db, """
            SELECT invoice_no, COALESCE(customer_name,'') AS customer,
                   COALESCE(status,'Draft') AS status, COALESCE(order_no,'') AS order_no
            FROM ar_invoices ORDER BY id DESC LIMIT 25"""):
            out.append({"type": "ar_invoice", "key": r["invoice_no"],
                        "label": f"{r['invoice_no']} — {r['customer']}",
                        "badge": r["status"], "url": "/finance"})
        if not out:
            out.append({"type": "leaf", "key": "", "leaf": True,
                        "label": "No AR invoices yet — Finance module next phase", "badge": ""})
    elif t_ == "root_masters":
        for label, typ, sql in [
            ("Customers", "master_customers", "SELECT COUNT(*) FROM customers"),
            ("Suppliers", "master_suppliers", "SELECT COUNT(*) FROM suppliers"),
            ("Recipes",   "master_recipes",   "SELECT COUNT(*) FROM recipes"),
            ("Brands",    "leaf",             "SELECT COUNT(*) FROM brands"),
        ]:
            out.append({"type": typ, "key": "", "label": label,
                        "badge": _tree_count(db, sql), "leaf": typ == "leaf",
                        "url": {"Customers": "/customers", "Suppliers": "/suppliers",
                                "Recipes": "/recipes", "Brands": "/brands"}[label]})

    # ---------------- order drill-down (the SAP document flow) ----------------
    elif t_ == "order":
        o = {"o": k}
        stages = [
            ("order_lines",  "Order Lines",    "SELECT COUNT(*) FROM order_lines WHERE order_no=:o", f"/production/orders/{k}"),
            ("order_bom",    "Production BOM", "SELECT COUNT(*) FROM bom_lines WHERE order_no=:o", f"/production/orders/{k}"),
            ("order_store",  "Store Issues",   "SELECT COUNT(*) FROM store_issuance_lines WHERE order_no=:o", "/production/store-issuance"),
            ("order_kitchen","Kitchen Moves",  "SELECT COUNT(*) FROM kitchen_section_transactions WHERE order_no=:o", "/production/orders"),
            ("order_qc",     "QC Checks",      "SELECT COUNT(*) FROM qc_checks WHERE order_no=:o", "/qc"),
            ("order_pack",   "Packing/Dispatch","SELECT COUNT(*) FROM packing_dispatch WHERE order_no=:o", "/packing"),
            ("order_ar",     "AR Invoice",     "SELECT COUNT(*) FROM ar_invoices WHERE order_no=:o", "/finance"),
        ]
        for typ, label, sql, url in stages:
            n = _tree_count(db, sql, o)
            out.append({"type": typ, "key": k, "label": label, "badge": n,
                        "url": url, "leaf": n == 0})
    elif t_ == "order_lines":
        for r in _tree_rows(db, """
            SELECT COALESCE(recipe_name, recipe_code,'') AS recipe,
                   COALESCE(portions, quantity, 0) AS qty
            FROM order_lines WHERE order_no=:o ORDER BY id LIMIT 100""", {"o": k}):
            out.append({"type": "leaf", "key": "", "leaf": True,
                        "label": r["recipe"], "badge": f"{r['qty']} portions"})
    elif t_ == "order_bom":
        for r in _tree_rows(db, """
            SELECT COALESCE(inventory_code,'') AS code, COALESCE(item_name, ingredient_name,'') AS item,
                   COALESCE(required_qty, quantity, 0) AS qty, COALESCE(uom,'') AS uom
            FROM bom_lines WHERE order_no=:o ORDER BY id LIMIT 200""", {"o": k}):
            out.append({"type": "inv_item", "key": r["code"], "leaf": not r["code"],
                        "label": f"{r['item']} ({r['code']})" if r["code"] else r["item"],
                        "badge": f"{r['qty']} {r['uom']}".strip()})
    elif t_ == "order_store":
        for r in _tree_rows(db, """
            SELECT COALESCE(inventory_code,'') AS code, COALESCE(item_name,'') AS item,
                   COALESCE(issued_qty, quantity, 0) AS qty
            FROM store_issuance_lines WHERE order_no=:o ORDER BY id LIMIT 200""", {"o": k}):
            out.append({"type": "inv_item", "key": r["code"], "leaf": not r["code"],
                        "label": f"{r['item']} ({r['code']})" if r["code"] else r["item"],
                        "badge": r["qty"]})
    elif t_ in ("order_kitchen", "order_qc", "order_pack", "order_ar"):
        table = {"order_kitchen": ("kitchen_section_transactions", "section", "txn_type"),
                 "order_qc": ("qc_checks", "check_point", "status"),
                 "order_pack": ("packing_dispatch", "dispatch_status", "vehicle_no"),
                 "order_ar": ("ar_invoices", "invoice_no", "status")}[t_]
        tbl, c1, c2 = table
        for r in _tree_rows(db, f"""
            SELECT COALESCE({c1},'') AS a, COALESCE({c2},'') AS b
            FROM {tbl} WHERE order_no=:o ORDER BY id DESC LIMIT 50""", {"o": k}):
            out.append({"type": "leaf", "key": "", "leaf": True,
                        "label": r["a"] or "—", "badge": r["b"]})

    # ---------------- inventory drill-down ----------------
    elif t_ == "inv_category":
        for r in _tree_rows(db, """
            SELECT inventory_code, COALESCE(item_name, ingredient_name,'') AS item,
                   COALESCE(current_stock,0) AS stock, COALESCE(uom,'') AS uom
            FROM ingredients WHERE COALESCE(main_category,'Uncategorized')=:c
            ORDER BY item LIMIT 100""", {"c": k}):
            out.append({"type": "inv_item", "key": r["inventory_code"],
                        "label": f"{r['item']} ({r['inventory_code']})",
                        "badge": f"{r['stock']} {r['uom']}".strip()})
    elif t_ == "inv_item":
        # where-used + movements: the item's links across the system
        used = _tree_count(db, "SELECT COUNT(*) FROM recipe_ingredients WHERE inventory_code=:c", {"c": k})
        moves = _tree_count(db, "SELECT COUNT(*) FROM inventory_transactions WHERE inventory_code=:c", {"c": k})
        po = _tree_count(db, "SELECT COUNT(*) FROM purchase_order_lines WHERE inventory_code=:c", {"c": k})
        out = [
            {"type": "inv_item_recipes", "key": k, "label": "Used in Recipes", "badge": used, "leaf": used == 0},
            {"type": "leaf", "key": "", "label": "Ledger Movements", "badge": moves, "leaf": True,
             "url": f"/inventory/ledger/{k}"},
            {"type": "leaf", "key": "", "label": "On Purchase Orders", "badge": po, "leaf": True, "url": "/procurement"},
        ]
    elif t_ == "inv_item_recipes":
        for r in _tree_rows(db, """
            SELECT DISTINCT COALESCE(r.recipe_name, r.name, '') AS recipe
            FROM recipe_ingredients ri JOIN recipes r ON r.id = ri.recipe_id
            WHERE ri.inventory_code=:c LIMIT 100""", {"c": k}):
            out.append({"type": "leaf", "key": "", "leaf": True, "label": r["recipe"], "badge": ""})

    # ---------------- procurement drill-down ----------------
    elif t_ == "po":
        lines = _tree_count(db, "SELECT COUNT(*) FROM purchase_order_lines WHERE po_no=:p", {"p": k})
        grns = _tree_count(db, "SELECT COUNT(*) FROM grns WHERE po_no=:p", {"p": k})
        out = [
            {"type": "po_lines", "key": k, "label": "PO Lines", "badge": lines, "leaf": lines == 0},
            {"type": "po_grns", "key": k, "label": "Goods Receipts (GRN)", "badge": grns, "leaf": grns == 0},
        ]
    elif t_ == "po_lines":
        for r in _tree_rows(db, """
            SELECT COALESCE(inventory_code,'') AS code, COALESCE(item_name,'') AS item,
                   COALESCE(quantity,0) AS qty
            FROM purchase_order_lines WHERE po_no=:p ORDER BY line_no LIMIT 100""", {"p": k}):
            out.append({"type": "inv_item", "key": r["code"], "leaf": not r["code"],
                        "label": f"{r['item']} ({r['code']})" if r["code"] else r["item"], "badge": r["qty"]})
    elif t_ == "po_grns":
        for r in _tree_rows(db, """
            SELECT grn_no, COALESCE(status,'') AS status FROM grns
            WHERE po_no=:p ORDER BY id DESC LIMIT 50""", {"p": k}):
            out.append({"type": "leaf", "key": "", "leaf": True,
                        "label": r["grn_no"], "badge": r["status"]})

    # ---------------- masters drill-down ----------------
    elif t_ == "master_customers":
        for r in _tree_rows(db, """
            SELECT customer_code, customer_name FROM customers ORDER BY customer_name LIMIT 100"""):
            out.append({"type": "customer", "key": r["customer_code"],
                        "label": r["customer_name"], "badge": r["customer_code"]})
    elif t_ == "customer":
        n = _tree_count(db, """
            SELECT COUNT(*) FROM customer_orders co
            JOIN customers c ON (co.customer_name = c.customer_name OR co.customer_no = c.customer_code)
            WHERE c.customer_code=:c""", {"c": k})
        out.append({"type": "customer_orders", "key": k, "label": "Orders", "badge": n, "leaf": n == 0})
    elif t_ == "customer_orders":
        for r in _tree_rows(db, """
            SELECT co.order_no, co.status FROM customer_orders co
            JOIN customers c ON (co.customer_name = c.customer_name OR co.customer_no = c.customer_code)
            WHERE c.customer_code=:c ORDER BY co.id DESC LIMIT 25""", {"c": k}):
            out.append({"type": "order", "key": r["order_no"],
                        "label": r["order_no"], "badge": r["status"],
                        "url": f"/production/orders/{r['order_no']}"})
    elif t_ == "master_suppliers":
        for r in _tree_rows(db, """
            SELECT COALESCE(supplier_code,'') AS code, supplier_name FROM suppliers
            ORDER BY supplier_name LIMIT 100"""):
            out.append({"type": "leaf", "key": "", "leaf": True,
                        "label": r["supplier_name"], "badge": r["code"]})
    elif t_ == "master_recipes":
        for r in _tree_rows(db, """
            SELECT id, COALESCE(recipe_name, name,'') AS recipe, COALESCE(status,'') AS status
            FROM recipes ORDER BY recipe LIMIT 100"""):
            out.append({"type": "recipe", "key": str(r["id"]),
                        "label": r["recipe"], "badge": r["status"]})
    elif t_ == "recipe":
        for r in _tree_rows(db, """
            SELECT COALESCE(inventory_code,'') AS code, COALESCE(ingredient_name, item_name,'') AS item,
                   COALESCE(quantity,0) AS qty, COALESCE(uom,'') AS uom
            FROM recipe_ingredients WHERE recipe_id=:r ORDER BY id LIMIT 100""", {"r": k}):
            out.append({"type": "inv_item", "key": r["code"], "leaf": not r["code"],
                        "label": f"{r['item']} ({r['code']})" if r["code"] else r["item"],
                        "badge": f"{r['qty']} {r['uom']}".strip()})

    if not out:
        out = [{"type": "leaf", "key": "", "leaf": True, "label": "No linked documents", "badge": ""}]
    return _JSON({"children": out})
